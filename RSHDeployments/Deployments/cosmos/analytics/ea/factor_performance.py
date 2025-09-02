#
# Functional factors
#
# Author: Yun Chen
# Copyright: Indigo Dao, LLC
# Date: 2022
#
import numbers

import numpy as np
import util.routines as rt
import stats.routine.filters as fil
import os.path
import classes.root as root
import pandas as pd
import util.utilities as util
from util.utilities import display
from util.utilities import display
# from analytics.va.variance_analysis import aggregate_period_contributions
# from analytics.va.variance_analysis import save_and_merge_contributions
# from analytics.va.variance_analysis import merge_contributions
# from analytics.va.variance_analysis import filter_contributions_by_dates
import openpyxl as pxl
import openpyxl.styles as styles
import dataloader.market_data as md
import dataloader.portfolio as port
from util.intersect import *
import factors.portfolio as pc


def factor_diagnostic(start_date, end_date, factors, universe, value_types='DESCRIPTOR',
                      reb_frequency='DAILY', diagnostics_flag=False,
                      calendar_str=None):
    if calendar_str is None or not isinstance(calendar_str, str):
        calendar_str = 'GL'
    bus_days = util.load_business_days(calendar_str, start_date, end_date)
    if len(bus_days) == 0:
        print(f"No valid business days according to {calendar_str}")
        return None
    if universe is None:
        print(f"No valid universe")
        return
    if factors is None:
        print(f"No valid factors")
        return None
    if isinstance(factors, str):
        factors = np.array([factors])
    if len(factors) == 0:
        print(f"Zero length in list of factors")
        return None
    if reb_frequency is None:
        reb_frequency = 'DAILY'
    if not isinstance(reb_frequency, str):
        reb_frequency = 'DAILY'
    reb_frequency = reb_frequency.strip().upper()
    if value_types is None:
        value_types = 'DESCRIPTOR'
    if isinstance(value_types, str):
        value_types = np.array([value_types])
    if len(factors) != len(value_types):
        if len(value_types) == 1:
            value_types = np.full((len(factors), 1), value_types[0])
        else:
            raise Exception('Mismatch dimension in factors and corresponding value types')
    wt_flag = 'EQUAL'
    # universe
    pos_days = util.previous_business_days(bus_days, calendar_str)
    reb_days = util.load_business_days(calendar_str, pos_days[0], pos_days[-1], reb_frequency)
    print(f"loading universe: {universe} between {pos_days[0]} and {pos_days[-1]}")
    univ = port.get_positions(pos_days[0], pos_days[-1], universe, calendar_str)

    cache_location = os.path.join(util.default_output_location('reports'), 'diagnostics')
    if not os.path.exists(cache_location):
        os.makedirs(cache_location)
        print(f"diagnostics cache location created: {cache_location}")
    sector = 'COSMOS_SECTOR'
    quintiles = {}
    deciles = {}
    for i, f in enumerate(factors):
        directory = os.path.join(cache_location, f"{f}", f"{value_types[i]}", f"{universe}", f"{reb_frequency}")
        if not os.path.join(directory):
            os.makedirs(directory)
            print(f"{f} cache location created: {directory}")
        quintiles[f] = filter_universe(reb_days, univ, f, value_types[i], 5, directory)
        print(f"finished quintile for {f}")
        # deciles[f] = filter_universe(reb_days, univ, f, value_types[i], 10, directory)
        # print(f"finished decile for {f}")
        # sector sort
        # industry sort
        # factor double sort

    if diagnostics_flag:
        print(f"Generating {wt_flag} weights for {universe}")
        wts = port.get_portfolio_weights(pos_days[0], pos_days[-1], universe, calendar_str=calendar_str,
                                         weight_flag='EQUAL')
        ret = md.get_returns(bus_days[0], bus_days[-1], univ.columns.to_numpy(), calendar_str)
        bench_ret = port.w_prime_r(wts, ret, calendar_str)
        print(f"Benchmark {universe} {wt_flag} total performance {bus_days[0]} - {bus_days[-1]}: "
              f"{(np.prod(1+bench_ret.to_numpy())-1)*100:.2f} %")
        for f in factors:
            qr = diagnostic(quintiles[f]['buckets'], quintiles[f]['factor values'], ret, bench_ret,
                            wt_flag, calendar_str=calendar_str)
    return qr, bench_ret


def filter_universe(days, por, factor, value_type, bin_size, cache_location=None):
    """
    loading values of a given for a factor on days, bucket stocks into percentile based on number of bin size
    'missing' bin is set up too
    :param days: rebalance days
    :param por:
    :param factor:
    :param value_type:
    :param bin_size:
    :param cache_location:
    :return:
    """
    if bin_size <= 0:
        bin_size = 1
    por_days = por.index.to_numpy()
    if not os.path.exists(cache_location):
        os.makedirs(cache_location)
        print(f"Created {cache_location}")
    if isinstance(factor, root.Root):
        factor_name = factor.name
    else:
        factor_name = factor
    days, i1, i2 = intersect(days, por.index.to_numpy())
    if len(days) == 0:
        return None
    factor = root.load_object(factor)
    buckets = {}
    for i in range(bin_size):
        buckets[i] = pd.DataFrame(0, index=por.index, columns=por.columns)
    buckets['missing'] = pd.DataFrame(0, index=por.index, columns=por.columns)
    values = pd.DataFrame(np.nan, index=por.index, columns=por.columns)
    directory = os.path.join(cache_location, f"{bin_size}")
    if not os.path.exists(directory):
        os.makedirs(directory)
        print(f"Created {directory}")
    reb_days = np.array([])
    for key in buckets.keys():
        file = os.path.join(directory, f"{key}.qd")
        if not os.path.exists(file):
            reb_days = days
            break
        data = util.load_data(file)
        missing = np.setdiff1d(por_days, data.index.to_numpy())
        if len(missing) == 0:
            continue
        k_days = np.array([])
        for j in range(len(days)):
            m = missing[missing >= days[j]]
            if j < len(days) - 1:
                m = m[m < days[j+1]]
            if len(m) == 0:
                continue
            k_days = np.append(k_days, days[j])
        if len(k_days) == 0:
            continue
        reb_days = np.union1d(reb_days, k_days)

    # caching portfolio buckets
    for dx, d in enumerate(reb_days):
        # --------------------------------------------------------
        # each rebalance days: bucket stocks, compute diagnostics
        # --------------------------------------------------------
        u = por.columns[por.loc[d] != 0].to_numpy()
        b = factor.load_values(value_type, d, d, u)
        if b is None:
            print(f"{d}: no valid {value_type} for factor {factor.name}")
            continue
        bt = b.transpose()
        bt['bin'] = pd.qcut(bt[d], bin_size, labels=range(bin_size))
        p_days = por.index[por.index >= d]
        if dx < len(reb_days) - 1:
            p_days = p_days[p_days < reb_days[dx + 1]]
        for i in range(bin_size):
            sec_ids = bt.index[bt['bin'] == i].to_numpy()
            if len(sec_ids) == 0:
                print(f"{util.current_time()}: {factor.name}: {value_type}: {d}: bucket - {i}: no stocks")
                continue
            bf = pd.DataFrame(1, index=p_days, columns=sec_ids)
            file = os.path.join(directory, f"{i}.qd")
            if os.path.exists(file):
                data = util.load_data(file)
                index = np.sort(np.union1d(data.index, bf.index))
                ids = np.union1d(data.columns, bf.columns)
                df = pd.DataFrame(0.0, index=index, columns=ids)
                df.update(data)
                df.update(bf)
                good_index = np.where(df.sum(axis=0) != 0)[0]
                df = df.iloc[:, good_index]
                util.save_data(df, file)
                print(f"{factor_name}: {d}: {bin_size} buckets: {i} merged new {len(p_days)} days records {len(bf.columns)} stocks "
                      f"onto {len(data.index)} days")
                del (df, good_index)
            else:
                util.save_data(bf, file)
                print(f"{factor_name}: {d}: bucket :{i}: {len(p_days)} days records {len(bf.columns)} stocks saved")
        # missing
        sec_ids = bt.index[pd.isnull(bt['bin'])].to_numpy()
        mf = pd.DataFrame(1, index=p_days, columns=sec_ids)
        file = os.path.join(directory, f"missing.qd")
        if os.path.exists(file):
            data = util.load_data(file)
            index = np.sort(np.union1d(data.index, mf.index))
            ids = np.union1d(data.columns, mf.columns)
            df = pd.DataFrame(0.0, index=index, columns=ids)
            df.update(data)
            df.update(mf)
            good_index = np.where(df.sum(axis=0) == 0)[0]
            df = df.iloc[:, good_index]
            util.save_data(df, file)
            print(f"{factor_name}: {d}: missing: merged new {len(p_days)} days records {len(mf.columns)} stocks "
                  f"onto {len(data.index)} days")
        else:
            util.save_data(mf, file)
            print(f"{factor_name}: {d}: missing : saved: {len(mf.columns)} stocks")

    for dx, d in enumerate(days):
        u = por.columns[por.loc[d] != 0].to_numpy()
        b = factor.load_values(value_type, d, d, u)
        p_days = por.index[por.index >= d]
        if dx < len(days) - 1:
            p_days = p_days[p_days < days[dx + 1]]

        for i in range(bin_size):
            file = os.path.join(directory, f"{i}.qd")
            data = util.load_data(file)
            sec_ids, i1, i2 = intersect(por.columns, data.columns)
            buckets[i].loc[p_days, sec_ids] = np.tile(data.loc[d, sec_ids], (len(p_days), 1))
        file = os.path.join(directory, f"missing.qd")
        data = util.load_data(file)
        sec_ids, i1, i2 = intersect(por.columns, data.columns)
        if len(sec_ids) > 0:
            buckets['missing'].loc[p_days, sec_ids] = np.tile(data.loc[d, sec_ids], (len(p_days), 1))
        # --------------------------------------------------------
        # each rebalance days: record factor values
        # --------------------------------------------------------
        values.loc[p_days, b.columns.to_numpy()] = np.tile(b.loc[d, b.columns], (len(p_days), 1))
    for i in buckets.keys():
        good_index = np.where(buckets[i].sum(axis=0) != 0)[0]
        buckets[i] = buckets[i].iloc[:, good_index]
        del good_index
    good_index = np.where(buckets['missing'].sum(axis=0) != 0)[0]
    buckets['missing'] = buckets['missing'].iloc[:, good_index]
    del good_index
    print(f"{util.current_time()}: {factor.name}: {value_type}: {bin_size} (+1) buckets: {days[0]} - {days[-1]}")
    result = {'buckets': buckets, 'factor values': values}
    return result


def diagnostic(buckets, factor_values, ret, bench_ret=None, wt_flag='EQUAL', calendar_str=None):
    if calendar_str is None:
        calendar_str = 'GL'
    if wt_flag is None:
        wt_flag = 'EQUAL'
    wt_flag = wt_flag.strip().upper()
    managed_returns = {}

    for key in buckets.keys():
        try:
            wts = port.calculate_weights(buckets[key], wt_flag, calendar_str)
            managed_returns[key] = port.w_prime_r(wts, ret, calendar_str)
        except ValueError:
            print(ValueError)

    return managed_returns


def get_bucket_names(universes, factors, value_types, bin_sizes, create_flag=False,
                     exclude_middle=False, add_to_universe=False, freq='MONTHEND', composite_flag=False,
                     weight_flag='EQUAL'):
    """

    :param universes:
    :param factors:
    :param value_types:
    :param bin_sizes:
    :param create_flag:
    :param exclude_middle:
    :param add_to_universe:
    :param freq:
    :param composite_flag: False
    :param weight_flag: EQUAL
    :return:
    """
    names = np.array([])
    if not isinstance(universes, (list, np.ndarray)):
        universes = np.array([universes])

    if not isinstance(factors, (list, np.ndarray)):
        factors = np.array([factors])
    if not isinstance(value_types, (list, np.ndarray)):
        value_types = np.array([value_types])
    if not isinstance(value_types, (list, np.ndarray)):
        value_types = np.array([value_types])
    if not isinstance(bin_sizes, (list, np.ndarray)):
        bin_sizes = np.array([bin_sizes])
    if weight_flag is None:
        weight_flag = 'EQUAL'

    for uix, u in enumerate(universes):
        v_type = value_types[uix]
        v_type = v_type.strip()
        if v_type in ['DESCRIPTOR', 'DESCRIPTORS']:
            v_type = 'D'
        elif v_type in ['EXPOSURE', 'EXPOSURES']:
            v_type = 'E'
        if not isinstance(freq, str):
            freq = 'DAILY'
        freq = freq.upper().strip()
        u_obj = root.load_object(u)
        sec_type = u_obj.security_type
        f_name = factors[uix]
        f_name = f_name.strip()
        bin_size = bin_sizes[uix]
        name = f"{u.strip()}_{f_name}_{v_type}_{bin_size}"
        for j in range(bin_size):
            if exclude_middle:
                if 0 < j < bin_size-1:
                    continue
            name_j = f"{name}_{j+1}"
            names = np.append(names, name_j)
            if root.exists(name_j):
                obj = root.load_object(name_j)
                print(f"{name_j} ({type(obj)}) already created")
                del obj
                continue
            if create_flag:
                print(f"Creating object: {name_j}")
                obj = pc.FilteredPortfolio(name_j)
                obj.universe = u
                obj.calendar = u_obj.calendar
                obj.factor_universes = np.array([u])
                obj.add_factor(f_name, root.Life(19000101, 99991231), value_types[uix], 'PORTFOLIO')
                low = j / bin_size * 100
                high = (j + 1) / bin_size * 100
                if j == 0:
                    obj.filters = np.array([lambda x: fil.percentile_filter(x, low, high, True, True)])
                else:
                    obj.filters = np.array([lambda x: fil.percentile_filter(x, low, high, False, True)])
                obj.descriptor_frequency = freq
                obj.security_type = sec_type
                obj.composite_flag = composite_flag
                obj.weighting_method = weight_flag
                root.save_object(obj, create_flag)
    if add_to_universe and create_flag:
        util.add_composites(names, save_flag=True)
    return names


def return_statistics(ret, ben=None, mkt=None, calendar_str='US', indicator='cosmos_us', regime_only=False, freq='DAILY'):
    if isinstance(ret, pd.Series):
        ret = ret.to_frame()
    freq = freq.strip().upper()
    if freq in ('MONTH', 'MONTHLY', 'MONTHEND'):
        af = 12
    elif freq in ('QUARTER', 'QUARTEREND', 'QUARTERLY'):
        af = 4
    elif freq in ('DAILY', 'DAY'):
        af = 252
    elif freq in ('YEAR', 'YEARLY', 'ANNUAL', 'ANNUALLY'):
        af = 1
    else:
        raise ValueError(f"wrong frequency: {freq}")
    s_day = ret.index[0]
    e_day = ret.index[-1]
    data = {'names': None, 'dates': ret.index.to_numpy(), 'total returns': None,
            'annualized returns': None, 'annual actual returns': None,
            'volatilities': None, 'information ratios': None,
            'draw down': None, 'draw down periods': None, 'annual returns': None,
            'annual volatilities': None, 'annual information ratios': None,
            'monthly returns': None, 'semiannual returns': None, 'business regimes': None,
            'horizons': None, 'correlations': None}
    result = {'managed': data.copy(), 'benchmark': data.copy(), 'active': data.copy(), 'market': data.copy(),
              'ts': None}
    names = ret.columns.to_numpy()
    if ben is None:
        ben = pd.DataFrame(0.0, index=ret.index, columns=['CASH'])
    if isinstance(ben, pd.Series):
        ben = ben.to_frame()
    if mkt is None:
        mkt = ben
    if isinstance(mkt, pd.Series):
        mkt = mkt.to_frame()
    bmk = ben.columns.to_numpy()
    result['dates'] = ret.index.to_numpy()
    result['managed']['names'] = names
    result['benchmark']['names'] = bmk
    result['market']['names'] = mkt.columns.to_numpy()
    result['active']['names'] = np.array(names)
    t = len(ret.index)
    r = ret.to_numpy()
    b = np.tile(ben.to_numpy(), (1, r.shape[1]))
    m = mkt.to_numpy()
    # handle some time series being shorter in time history
    # match index returns with that series
    lengths = np.full((1, len(names)),len(ret.index))
    for i in range(r.shape[1]):
        ix = np.argmax(~np.isnan(r[:, i]))
        if ix > 0:
            b[:ix, i] = np.NAN
            lengths[0, i] = t - ix
    a = r - b
    dates = ret.index.to_numpy()
    months = util.month(dates)
    years = util.year(dates)
    unique_years = np.unique(years)
    unique_months = np.unique(months)
    result['ts'] = {'managed': ret, 'benchmark': ben, 'active': pd.DataFrame(a, columns=names, index=dates),
                    'market': mkt}
    # --------------------------------------------
    # period return statistics
    # total returns
    managed = result['managed']
    benchmark = result['benchmark']
    active = result['active']
    market = result['market']
    managed['total returns'] = pd.DataFrame(np.nanprod(1+r, axis=0, keepdims=True) - 1, index=['values'],
                                            columns=names)
    benchmark['total returns'] = pd.DataFrame(np.nanprod(1+b, axis=0, keepdims=True) - 1, index=['values'],
                                              columns=names)
    active['total returns'] = managed['total returns'] - benchmark['total returns']
    market['total returns'] = np.prod(1 + m) - 1
    managed['volatilities'] = pd.DataFrame(np.nanstd(r, axis=0, keepdims=True) * np.sqrt(af), index=['values'],
                                           columns=names)
    benchmark['volatilities'] = pd.DataFrame(np.nanstd(b, axis=0, keepdims=True) * np.sqrt(af), index=['values'],
                                             columns=names)
    active['volatilities'] = pd.DataFrame(np.nanstd(a, axis=0, keepdims=True) * np.sqrt(af), index=['values'],
                                          columns=names)
    managed['draw down'] = pd.DataFrame(columns=names, index=['values', 'start', 'end', 'duration'])
    benchmark['draw down'] = pd.DataFrame(columns=[bmk], index=['values', 'start', 'end', 'duration'])
    active['draw down'] = pd.DataFrame(columns=names, index=['values', 'start', 'end', 'duration'])
    benchmark['volatilities'][benchmark['volatilities'] == 0.0] = np.nan
    managed['beta'] = pd.DataFrame(columns=names, index=['beta'])
    good_index = np.where(pd.notnull(mkt))[0]
    for ix in range(r.shape[1]):
        try:
            r_index = np.where(pd.notnull(r[:, ix]))[0]
            gx = np.intersect1d(good_index, r_index)
            if len(gx) == 0:
                print(f"{names[ix]} no valid series overlapping market")
                continue
            beta = np.linalg.lstsq(mkt.iloc[gx].to_numpy(), r[gx, ix], rcond=None)
            managed['beta'].iloc[0, ix] = beta[0][0]
        except ValueError as ve:
            print(ve)
            print(f"{names[ix]}: failed in beta computation")
        except Exception as ee:
            print(ee)
            print(f"{names[ix]}: failed in beta computation")
    benchmark['beta'] = pd.DataFrame(1.0, columns=names, index=['beta'])
    try:
        r_index = np.where(pd.notnull(b[:, 0]))[0]
        gx = np.intersect1d(good_index, r_index)
        if len(gx) == 0:
            print(f"benchmark: no valid series overlapping market")
        else:
            beta = np.linalg.lstsq(mkt.iloc[gx].to_numpy(), b[gx, 0], rcond=None)
            benchmark['beta'].loc['beta', names] = beta[0][0]
    except ValueError as ve:
        print(ve)
        print(f"benchmark: failed in beta computation")
    except Exception as ee:
        print(ee)
        print(f"benchmark: failed in beta computation")
    active['beta'] = managed['beta'] - benchmark['beta']
    for name in names:
        try:
            dd, periods, duration = rt.maximum_drawdown(ret[name])
            managed['draw down'].loc['values', name] = dd
            managed['draw down'].loc['start', name] = periods[0]
            managed['draw down'].loc['end', name] = periods[1]
            managed['draw down'].loc['duration', name] = duration
            del (dd, periods, duration)
            dd, periods, duration = rt.maximum_drawdown_relative(ret[name], ben)
            active['draw down'].loc['values', name] = dd
            active['draw down'].loc['start', name] = periods[0]
            active['draw down'].loc['end', name] = periods[1]
            active['draw down'].loc['duration', name] = duration
            del (dd, periods, duration)
        except ValueError as ve:
            print(ve)
            print(f"Unable to assess drawdown for {name}")
    dd, periods, duration = rt.maximum_drawdown(ben)
    benchmark['draw down'].loc['values', bmk] = dd
    benchmark['draw down'].loc['start', bmk] = periods[0]
    benchmark['draw down'].loc['end', bmk] = periods[1]
    benchmark['draw down'].loc['duration', bmk] = duration
    del (dd, periods, duration)
    # correlations
    cr = pd.DataFrame(a, columns=names)
    cr = cr.dropna()
    if not cr.empty:
        active['correlations'] = cr.corr()
    # regime statistics
    if indicator is None or not isinstance(indicator, str):
        indicator = 'cosmos_us'
    indicator = indicator.strip()
    indicator_ref = md.get_indicator_references()
    if indicator not in list(indicator_ref['indicators']):
        print(f"Unsupported business regimes")
        return None
    result['indicator'] = indicator
    regimes = md.get_regimes(s_day, e_day, indicator, calendar_str)
    rg_days = np.intersect1d(regimes.index, ret.index)
    regimes = regimes.loc[rg_days]
    unique_regimes = np.unique(regimes['regimes'])
    unique_regime_values = np.unique(regimes['values'])
    regime_ref = md.get_regime_reference(indicator)
    all_regimes = np.unique(regime_ref['values'])
    managed['regime returns'] = pd.DataFrame(index=all_regimes, columns=names)
    managed['regime volatilities'] = pd.DataFrame(index=all_regimes, columns=names)
    managed['regime information ratios'] = pd.DataFrame(index=all_regimes, columns=names)
    managed['regime betas'] = pd.DataFrame(index=all_regimes, columns=names)
    benchmark['regime returns'] = pd.DataFrame(index=all_regimes, columns=names)
    benchmark['regime volatilities'] = pd.DataFrame(index=all_regimes, columns=names)
    benchmark['regime information ratios'] = pd.DataFrame(index=all_regimes, columns=names)
    benchmark['regime betas'] = pd.DataFrame(1.0, index=all_regimes, columns=names)
    active['regime returns'] = pd.DataFrame(index=all_regimes, columns=names)
    active['regime volatilities'] = pd.DataFrame(index=all_regimes, columns=names)
    active['regime information ratios'] = pd.DataFrame(index=all_regimes, columns=names)
    active['regime betas'] = pd.DataFrame(index=all_regimes, columns=names)
    market['regime returns'] = pd.DataFrame(index=all_regimes, columns=names)
    market['regime volatilities'] = pd.DataFrame(index=all_regimes, columns=names)
    market['regime information ratios'] = pd.DataFrame(index=all_regimes, columns=names)
    market['regime betas'] = pd.DataFrame(1.0, index=all_regimes, columns=names)
    for regime in unique_regime_values:
        index = np.where(regimes['values'] == regime)[0]
        if len(index) == 0:
            continue
        mr = r[index, :]
        mb = b[index, :]
        ma = a[index, :]
        mm = m[index]
        tr = np.sum(~np.isnan(mr), axis=0)
        tb = np.sum(~np.isnan(mb), axis=0)
        managed['regime returns'].loc[regime] = np.nanprod(1+mr, axis=0, keepdims=True)**(af/tr) - 1
        managed['regime volatilities'].loc[regime] = np.nanstd(mr, axis=0) * np.sqrt(af)
        benchmark['regime returns'].loc[regime] = np.nanprod(1+mb, axis=0, keepdims=True)**(af/tb) - 1
        benchmark['regime volatilities'].loc[regime] = np.nanstd(mb, axis=0) * np.sqrt(af)
        active['regime volatilities'].loc[regime] = np.nanstd(ma, axis=0) * np.sqrt(af)
        market['regime returns'].loc[regime] = np.nanprod(1+mm, axis=0, keepdims=True)**(af/tr) - 1
        market['regime volatilities'].loc[regime] = np.nanstd(mm, axis=0) * np.sqrt(af)
        for ix in range(mr.shape[1]):
            mx = np.where(np.logical_and(pd.notnull(mb[:,ix]), pd.notnull(mr[:, ix])))[0]
            if len(mx) <= 3:
                print(f"{regime}: {names[ix]} no sufficient data: beta calculation")
                continue
            beta = np.linalg.lstsq(np.atleast_2d(mm[mx]), mr[mx, ix], rcond=None)
            managed['regime betas'].loc[regime, names[ix]] = beta[0][0]
            beta = np.linalg.lstsq(np.atleast_2d(mm[mx]), mb[mx, ix], rcond=None)
            benchmark['regime betas'].loc[regime, names[ix]] = beta[0][0]
    active['regime betas'] = managed['regime betas'] - benchmark['regime betas']
    managed['regime volatilities'][managed['regime volatilities'] == 0.0] = np.nan
    benchmark['regime volatilities'][benchmark['regime volatilities'] == 0.0] = np.nan
    active['regime volatilities'][active['regime volatilities'] == 0.0] = np.nan
    managed['regime information ratios'] = managed['regime returns'] / managed['regime volatilities']
    benchmark['regime information ratios'] = benchmark['regime returns'] / benchmark['regime volatilities']
    active['regime returns'] = managed['regime returns'] - benchmark['regime returns']
    active['regime information ratios'] = active['regime returns'] / active['regime volatilities']
    market['regime information ratios'] = market['regime returns'] / market['regime volatilities']
    if regime_only:
        return result
    # ------------------------------
    # horizons
    horizon_index = np.array(['YTD', '1 Yr', '3 Yr', '5 Yr', '7 Yr', '10 Yr', '20 Yr', 'Since Inception'])
    horizon = pd.DataFrame(np.nan, index=horizon_index, columns=names, dtype='float64')
    managed['horizons'] = {'returns': horizon.copy(), 'volatilities': horizon.copy(),
                           'information ratios': horizon.copy(), 'hit rates': horizon.copy()}
    benchmark['horizons'] = {'returns': horizon.copy(), 'volatilities': horizon.copy(),
                             'information ratios': horizon.copy(), 'hit rates': horizon.copy()}
    active['horizons'] = {'returns': horizon.copy(), 'volatilities': horizon.copy(),
                          'information ratios': horizon.copy(), 'hit rates': horizon.copy()}
    market['horizons'] = {'returns': horizon.copy(), 'volatilities': horizon.copy(),
                          'information ratios': horizon.copy(), 'hit rates': horizon.copy()}
    # YTD
    index = np.where(years == years[-1])[0]
    ts = r[index, :]
    y = 'YTD'
    n_days = len(index)
    # managed['horizons']['returns'].loc[y] = np.nanprod(1 + ts, 0) ** (af / n_days) - 1
    managed['horizons']['returns'].loc[y] = np.nanprod(1 + ts, 0) - 1
    managed['horizons']['volatilities'].loc[y] = np.nanstd(ts, 0) * np.sqrt(af)
    managed['horizons']['information ratios'].loc[y] = managed['horizons']['returns'].loc[y] \
                                                       / managed['horizons']['volatilities'].loc[y]
    managed['horizons']['hit rates'].loc[y] = np.sum(ts > 0, 0) / pd.notnull(ts).sum(axis=0)
    ts = b[index, :]
    # benchmark['horizons']['returns'].loc[y] = np.nanprod(1 + ts, 0) ** (af / n_days) - 1
    benchmark['horizons']['returns'].loc[y] = np.nanprod(1 + ts, 0) - 1
    benchmark['horizons']['volatilities'].loc[y] = np.nanstd(ts, 0) * np.sqrt(af)
    if (benchmark['horizons']['volatilities'].loc[y] == 0).any():
        benchmark['horizons']['volatilities'].loc[y] = np.nan
    benchmark['horizons']['information ratios'].loc[y] = benchmark['horizons']['returns'].loc[y] \
                                                         / benchmark['horizons']['volatilities'].loc[y]
    benchmark['horizons']['hit rates'].loc[y] = np.sum(ts > 0, 0) / pd.notnull(ts).sum(axis=0)
    ts = a[index, :]
    # active['horizons']['returns'].loc[y] = np.nanprod(1 + ts, 0) ** (af / n_days) - 1
    active['horizons']['returns'].loc[y] = managed['horizons']['returns'].loc[y] - \
        benchmark['horizons']['returns'].loc[y]
    active['horizons']['volatilities'].loc[y] = np.nanstd(ts, 0) * np.sqrt(af)
    active['horizons']['information ratios'].loc[y] = active['horizons']['returns'].loc[y] \
                                                      / active['horizons']['volatilities'].loc[y]
    active['horizons']['hit rates'].loc[y] = np.sum(ts > 0, 0) / pd.notnull(ts).sum(axis=0)

    ts = m[index]
    # market['horizons']['returns'].loc[y] = np.nanprod(1 + ts, 0) ** (af / n_days) - 1
    market['horizons']['returns'].loc[y] = np.nanprod(1 + ts, 0) - 1
    market['horizons']['volatilities'].loc[y] = np.nanstd(ts, 0) * np.sqrt(af)
    market['horizons']['information ratios'].loc[y] = market['horizons']['returns'].loc[y] \
                                                       / market['horizons']['volatilities'].loc[y]
    market['horizons']['hit rates'].loc[y] = np.sum(ts > 0, 0) / pd.notnull(ts).sum(axis=0)
    del (y, n_days)
    # 1 Yr, 3 Yr, 5 Yr, 7 Yr, 10 Yr, 20 Yr
    n_years = np.array(['1 Yr', '3 Yr', '5 Yr', '7 Yr', '10 Yr', '20 Yr'])
    yf = np.array([1, 3, 5, 7, 10, 20])
    for ix, y in enumerate(n_years):
        n_days = af * yf[ix]
        if len(dates) >= n_days:
            ts = r[-n_days:, :]
            managed['horizons']['returns'].loc[y] = np.nanprod(1 + ts, 0)**(af/n_days) - 1
            managed['horizons']['volatilities'].loc[y] = np.nanstd(ts, 0) * np.sqrt(af)
            managed['horizons']['information ratios'].loc[y] = managed['horizons']['returns'].loc[y] \
                                                               / managed['horizons']['volatilities'].loc[y]
            managed['horizons']['hit rates'].loc[y] = np.sum(ts > 0, 0) / pd.notnull(ts).sum(axis=0)
            ts = b[-n_days:, :]
            benchmark['horizons']['returns'].loc[y] = np.nanprod(1 + ts, 0) ** (af / n_days) - 1
            benchmark['horizons']['volatilities'].loc[y] = np.nanstd(ts, 0) * np.sqrt(af)
            if (benchmark['horizons']['volatilities'].loc[y] == 0).any():
                benchmark['horizons']['volatilities'].loc[y] = np.nan
            benchmark['horizons']['information ratios'].loc[y] = benchmark['horizons']['returns'].loc[y] \
                                                               / benchmark['horizons']['volatilities'].loc[y]
            benchmark['horizons']['hit rates'].loc[y] = np.sum(ts > 0, 0) / pd.notnull(ts).sum(axis=0)
            ts = a[-n_days:, :]
            # active['horizons']['returns'].loc[y] = np.nanprod(1 + ts, 0) ** (af / n_days) - 1
            active['horizons']['returns'].loc[y] = managed['horizons']['returns'].loc[y] - \
                benchmark['horizons']['returns'].loc[y]
            active['horizons']['volatilities'].loc[y] = np.nanstd(ts, 0) * np.sqrt(af)
            active['horizons']['information ratios'].loc[y] = active['horizons']['returns'].loc[y] \
                                                               / active['horizons']['volatilities'].loc[y]
            active['horizons']['hit rates'].loc[y] = np.sum(ts > 0, 0) / pd.notnull(ts).sum(axis=0)
            ts = m[-n_days:, :]
            market['horizons']['returns'].loc[y] = np.nanprod(1 + ts, 0) ** (af / n_days) - 1
            market['horizons']['volatilities'].loc[y] = np.nanstd(ts, 0) * np.sqrt(af)
            market['horizons']['information ratios'].loc[y] = market['horizons']['returns'].loc[y] \
                                                               / market['horizons']['volatilities'].loc[y]
            market['horizons']['hit rates'].loc[y] = np.sum(ts > 0, 0) / pd.notnull(ts).sum(axis=0)
    # Since Inception
    ts = r
    y = 'Since Inception'
    n_days = ts.shape[0]
    managed['horizons']['returns'].loc[y] = np.nanprod(1 + ts, 0) ** (af / n_days) - 1
    managed['horizons']['volatilities'].loc[y] = np.nanstd(ts, 0) * np.sqrt(af)
    managed['horizons']['information ratios'].loc[y] = managed['horizons']['returns'].loc[y] \
                                                       / managed['horizons']['volatilities'].loc[y]
    managed['horizons']['hit rates'].loc[y] = np.sum(ts > 0, 0) / pd.notnull(ts).sum(axis=0)
    ts = b
    benchmark['horizons']['returns'].loc[y] = np.nanprod(1 + ts, 0) ** (af / n_days) - 1
    benchmark['horizons']['volatilities'].loc[y] = np.nanstd(ts, 0) * np.sqrt(af)
    if (benchmark['horizons']['volatilities'].loc[y] == 0).any():
        benchmark['horizons']['volatilities'].loc[y] = np.nan
    benchmark['horizons']['information ratios'].loc[y] = benchmark['horizons']['returns'].loc[y] \
                                                         / benchmark['horizons']['volatilities'].loc[y]
    benchmark['horizons']['hit rates'].loc[y] = np.sum(ts > 0, 0) / pd.notnull(ts).sum(axis=0)
    ts = a
    # active['horizons']['returns'].loc[y] = np.nanprod(1 + ts, 0) ** (af / n_days) - 1
    active['horizons']['returns'].loc[y] = managed['horizons']['returns'].loc[y] - \
        benchmark['horizons']['returns'].loc[y]
    active['horizons']['volatilities'].loc[y] = np.nanstd(ts, 0) * np.sqrt(af)
    active['horizons']['information ratios'].loc[y] = active['horizons']['returns'].loc[y] \
                                                      / active['horizons']['volatilities'].loc[y]
    active['horizons']['hit rates'].loc[y] = np.sum(ts > 0, 0) / pd.notnull(ts).sum(axis=0)
    ts = m
    market['horizons']['returns'].loc[y] = np.nanprod(1 + ts, 0) ** (af / n_days) - 1
    market['horizons']['volatilities'].loc[y] = np.nanstd(ts, 0) * np.sqrt(af)
    market['horizons']['information ratios'].loc[y] = market['horizons']['returns'].loc[y] \
                                                       / market['horizons']['volatilities'].loc[y]
    market['horizons']['hit rates'].loc[y] = np.sum(ts > 0, 0) / pd.notnull(ts).sum(axis=0)
    del (ts, y, n_days)
    # annualized returns
    managed['annualized returns'] = pd.DataFrame((1+managed['total returns']) ** (af/lengths) - 1,
                                                 index=['values'], columns=names)
    benchmark['annualized returns'] = pd.DataFrame((1+benchmark['total returns']) ** (af/lengths) - 1,
                                                   index=['values'], columns=names)
    active['annualized returns'] = managed['annualized returns'] - benchmark['annualized returns']
    market['annualized returns'] = np.prod(1+market['total returns']) ** (af/len(mkt.index)) - 1
    managed['information ratios'] = managed['annualized returns'] / managed['volatilities']
    benchmark['information ratios'] = benchmark['annualized returns'] / benchmark['volatilities']
    active['information ratios'] = active['annualized returns'] / active['volatilities']
    # yearly returns
    managed['annual returns'] = pd.DataFrame(index=unique_years, columns=names)
    managed['annual actual returns'] = pd.DataFrame(index=unique_years, columns=names)
    managed['annual volatilities'] = pd.DataFrame(index=unique_years, columns=names)
    managed['annual information ratios'] = pd.DataFrame(index=unique_years, columns=names)
    managed['annual betas'] = pd.DataFrame(index=unique_years, columns=names)
    benchmark['annual returns'] = pd.DataFrame(index=unique_years, columns=names)
    benchmark['annual actual returns'] = pd.DataFrame(index=unique_years, columns=names)
    benchmark['annual volatilities'] = pd.DataFrame(index=unique_years, columns=names)
    benchmark['annual information ratios'] = pd.DataFrame(index=unique_years, columns=names)
    benchmark['annual betas'] = pd.DataFrame(1.0, index=unique_years, columns=names)
    active['annual returns'] = pd.DataFrame(index=unique_years, columns=names)
    active['annual actual returns'] = pd.DataFrame(index=unique_years, columns=names)
    active['annual volatilities'] = pd.DataFrame(index=unique_years, columns=names)
    active['annual information ratios'] = pd.DataFrame(index=unique_years, columns=names)
    active['annual betas'] = pd.DataFrame(index=unique_years, columns=names)
    for year in unique_years:
        index = np.where(years == year)[0]
        if len(index) == 0:
            continue
        if len(index) == 1:
            print(f"only one day observation found for year {year}: skipping")
            continue
        yr = r[index, :]
        yb = b[index, :]
        ya = a[index, :]
        ym = m[index]
        tr = np.sum(~np.isnan(yr), axis=0)
        tb = np.sum(~np.isnan(yb), axis=0)
        tr = tr.astype('float64')
        tb = tb.astype('float64')
        tr[tr == 0] = np.nan
        tb[tb == 0] = np.nan
        managed['annual returns'].loc[year] = np.nanprod(1+yr, axis=0, keepdims=True)**(af/tr) - 1
        managed['annual actual returns'].loc[year] = np.nanprod(1 + yr, axis=0, keepdims=True) - 1
        managed['annual volatilities'].loc[year] = np.nanstd(yr, axis=0) * np.sqrt(af)
        benchmark['annual returns'].loc[year] = np.nanprod(1+yb, axis=0, keepdims=True)**(af/tb) - 1
        benchmark['annual actual returns'].loc[year] = np.nanprod(1 + yb, axis=0, keepdims=True) - 1
        benchmark['annual volatilities'].loc[year] = np.nanstd(yb, axis=0) * np.sqrt(af)
        active['annual volatilities'].loc[year] = np.nanstd(ya, axis=0) * np.sqrt(af)
        for ix in range(yr.shape[1]):
            yx = np.where(np.logical_and(pd.notnull(yb[:, ix]), pd.notnull(yr[:, ix])))[0]
            if len(yx) < 3:
                print(f"{year}: {names[ix]}: no sufficient overlapping data; beta calculations")
                continue
            beta = np.linalg.lstsq(np.atleast_2d(ym[yx]), yr[yx, ix], rcond=None)
            managed['annual betas'].loc[year, names[ix]] = beta[0][0]
            beta = np.linalg.lstsq(np.atleast_2d(ym[yx]), yb[yx, ix], rcond=None)
            benchmark['annual betas'].loc[year, names[ix]] = beta[0][0]
    active['annual betas'] = managed['annual betas'] - benchmark['annual betas']
    managed['annual volatilities'][managed['annual volatilities'] == 0.0] = np.nan
    benchmark['annual volatilities'][benchmark['annual volatilities'] == 0.0] = np.nan
    active['annual volatilities'][active['annual volatilities'] == 0.0] = np.nan
    managed['annual information ratios'] = managed['annual returns'] / managed['annual volatilities']
    benchmark['annual information ratios'] = benchmark['annual returns'] / benchmark['annual volatilities']
    active['annual returns'] = managed['annual returns'] - benchmark['annual returns']
    active['annual actual returns'] = managed['annual actual returns'] - benchmark['annual actual returns']
    active['annual information ratios'] = active['annual returns'] / active['annual volatilities']

    # monthly returns
    #
    all_months = range(1, 13)
    managed['monthly returns'] = pd.DataFrame(index=all_months, columns=names)
    managed['monthly volatilities'] = pd.DataFrame(index=all_months, columns=names)
    managed['monthly information ratios'] = pd.DataFrame(index=all_months, columns=names)
    benchmark['monthly returns'] = pd.DataFrame(index=all_months, columns=names)
    benchmark['monthly volatilities'] = pd.DataFrame(index=all_months, columns=names)
    benchmark['monthly information ratios'] = pd.DataFrame(index=all_months, columns=names)
    active['monthly returns'] = pd.DataFrame(index=all_months, columns=names)
    active['monthly volatilities'] = pd.DataFrame(index=all_months, columns=names)
    active['monthly information ratios'] = pd.DataFrame(index=all_months, columns=names)
    for month in unique_months:
        index = np.where(months == month)[0]
        if len(index) == 0:
            continue
        mr = r[index, :]
        mb = b[index, :]
        ma = a[index, :]
        tr = np.sum(~np.isnan(mr), axis=0)
        tb = np.sum(~np.isnan(mb), axis=0)
        tr = tr.astype('float64')
        tb = tb.astype('float64')
        tr[tr == 0] = np.nan
        tb[tb == 0] = np.nan
        managed['monthly returns'].loc[month] = np.nanprod(1+mr, axis=0, keepdims=True)**(af/tr) - 1
        managed['monthly volatilities'].loc[month] = np.nanstd(mr, axis=0) * np.sqrt(af)
        benchmark['monthly returns'].loc[month] = np.nanprod(1+mb, axis=0, keepdims=True)**(af/tb) - 1
        benchmark['monthly volatilities'].loc[month] = np.nanstd(mb, axis=0) * np.sqrt(af)
        active['monthly volatilities'].loc[month] = np.nanstd(ma, axis=0) * np.sqrt(af)
    managed['monthly volatilities'][managed['monthly volatilities'] == 0] = np.nan
    benchmark['monthly volatilities'][benchmark['monthly volatilities'] == 0] = np.nan
    active['monthly volatilities'][active['monthly volatilities'] == 0] = np.nan
    managed['monthly information ratios'] = managed['monthly returns'] / managed['monthly volatilities']
    benchmark['monthly information ratios'] = benchmark['monthly returns'] / benchmark['monthly volatilities']
    active['monthly returns'] = managed['monthly returns'] - benchmark['monthly returns']
    active['monthly information ratios'] = active['monthly returns'] / active['monthly volatilities']

    # quarterly returns
    #
    all_quarters = range(1, 5)
    managed['quarterly returns'] = pd.DataFrame(index=all_quarters, columns=names)
    managed['quarterly volatilities'] = pd.DataFrame(index=all_quarters, columns=names)
    managed['quarterly information ratios'] = pd.DataFrame(index=all_quarters, columns=names)
    benchmark['quarterly returns'] = pd.DataFrame(index=all_quarters, columns=names)
    benchmark['quarterly volatilities'] = pd.DataFrame(index=all_quarters, columns=names)
    benchmark['quarterly information ratios'] = pd.DataFrame(index=all_quarters, columns=names)
    active['quarterly returns'] = pd.DataFrame(index=all_quarters, columns=names)
    active['quarterly volatilities'] = pd.DataFrame(index=all_quarters, columns=names)
    active['quarterly information ratios'] = pd.DataFrame(index=all_quarters, columns=names)
    for quarter in all_quarters:
        index = np.where(np.isin(months, (quarter - 1) * 3 + np.array([1, 2, 3])))[0]
        if len(index) == 0:
            continue
        mr = r[index, :]
        mb = b[index, :]
        ma = a[index, :]
        tr = np.sum(~np.isnan(mr), axis=0)
        tb = np.sum(~np.isnan(mb), axis=0)
        tr = tr.astype('float64')
        tb = tb.astype('float64')
        tr[tr == 0] = np.nan
        tb[tb == 0] = np.nan
        managed['quarterly returns'].loc[quarter] = np.nanprod(1+mr, axis=0, keepdims=True)**(af/tr) - 1
        managed['quarterly volatilities'].loc[quarter] = np.nanstd(mr, axis=0) * np.sqrt(af)
        benchmark['quarterly returns'].loc[quarter] = np.nanprod(1+mb, axis=0, keepdims=True)**(af/tb) - 1
        benchmark['quarterly volatilities'].loc[quarter] = np.nanstd(mb, axis=0) * np.sqrt(af)
        active['quarterly volatilities'].loc[quarter] = np.nanstd(ma, axis=0) * np.sqrt(af)
    managed['quarterly volatilities'][managed['quarterly volatilities'] == 0] = np.nan
    benchmark['quarterly volatilities'][benchmark['quarterly volatilities'] == 0] = np.nan
    active['quarterly volatilities'][active['quarterly volatilities'] == 0] = np.nan
    managed['quarterly information ratios'] = managed['quarterly returns'] / managed['quarterly volatilities']
    benchmark['quarterly information ratios'] = benchmark['quarterly returns'] / benchmark['quarterly volatilities']
    active['quarterly returns'] = managed['quarterly returns'] - benchmark['quarterly returns']
    active['quarterly information ratios'] = active['quarterly returns'] / active['quarterly volatilities']

    # half-yearly returns
    #
    half_years = range(1, 3)
    managed['semiannual returns'] = pd.DataFrame(index=half_years, columns=names)
    managed['semiannual volatilities'] = pd.DataFrame(index=half_years, columns=names)
    managed['semiannual information ratios'] = pd.DataFrame(index=half_years, columns=names)
    benchmark['semiannual returns'] = pd.DataFrame(index=half_years, columns=names)
    benchmark['semiannual volatilities'] = pd.DataFrame(index=half_years, columns=names)
    benchmark['semiannual information ratios'] = pd.DataFrame(index=half_years, columns=names)
    active['semiannual returns'] = pd.DataFrame(index=half_years, columns=names)
    active['semiannual volatilities'] = pd.DataFrame(index=half_years, columns=names)
    active['semiannual information ratios'] = pd.DataFrame(index=half_years, columns=names)
    for half in half_years:
        index = np.where(np.isin(months, (half - 1) * 6 + np.array([1, 2, 3, 4, 5, 6])))[0]
        if len(index) == 0:
            continue
        mr = r[index, :]
        mb = b[index, :]
        ma = a[index, :]
        tr = np.sum(~np.isnan(mr), axis=0)
        tb = np.sum(~np.isnan(mb), axis=0)
        tr = tr.astype('float64')
        tb = tb.astype('float64')
        tr[tr == 0] = np.nan
        tb[tb == 0] = np.nan
        managed['semiannual returns'].loc[half] = np.nanprod(1+mr, axis=0, keepdims=True)**(af/tr) - 1
        managed['semiannual volatilities'].loc[half] = np.nanstd(mr, axis=0) * np.sqrt(af)
        benchmark['semiannual returns'].loc[half] = np.nanprod(1+mb, axis=0, keepdims=True)**(af/tb) - 1
        benchmark['semiannual volatilities'].loc[half] = np.nanstd(mb, axis=0) * np.sqrt(af)
        active['semiannual volatilities'].loc[half] = np.nanstd(ma, axis=0) * np.sqrt(af)
    managed['semiannual volatilities'][managed['semiannual volatilities'] == 0] = np.nan
    benchmark['semiannual volatilities'][benchmark['semiannual volatilities'] == 0] = np.nan
    active['semiannual volatilities'][active['semiannual volatilities'] == 0] = np.nan
    managed['semiannual information ratios'] = managed['semiannual returns'] / managed['semiannual volatilities']
    benchmark['semiannual information ratios'] = benchmark['semiannual returns'] / benchmark['semiannual volatilities']
    active['semiannual returns'] = managed['semiannual returns'] - benchmark['semiannual returns']
    active['semiannual information ratios'] = active['semiannual returns'] / active['semiannual volatilities']

    return result


def get_bucket_returns(start_date, end_date, buckets, calendar_str=None):
    bus_days = util.load_business_days(calendar_str, start_date, end_date)
    if len(bus_days) == 0:
        return None
    if isinstance(buckets, str):
        buckets = np.array([buckets])
    buckets = np.sort(np.unique(buckets))
    df = pd.DataFrame(np.nan, index=bus_days, columns=buckets)
    for bucket in buckets:
        obj = root.load_object(bucket)
        if obj is None:
            print(f"Bucket: {bucket} not set up properly; skipping loading returns")
            continue
        r = obj.get_returns(bus_days[0], bus_days[-1], calendar_str)
        if r is None:
            obj.compute_returns(bus_days[0], bus_days[-1], True)
            r = obj.get_returns(bus_days[0], bus_days[-1], calendar_str)
        else:
            dates = r.index[np.where(pd.isnull(r))[0]].to_numpy()
            if len(dates) > 0:
                try:
                    obj.compute_returns(dates[0], dates[-1], True)
                    r = obj.get_returns(bus_days[0], bus_days[-1], calendar_str)
                except ValueError:
                    print(f'Unable to compute returns for {obj.name}')
                    continue
                # for d in dates:
                #     try:
                #         obj.compute_returns(d, d, True)
                #     except ValueError:
                #         print(f"{d}: cannot compute returns for {obj.name}")
        df.update(r)
    return df


def regression_ts(ret, factor_returns, bmk=None, look_back=63, lag=0, r_lower=0, r_higher=100,
                  fr_lower=0, fr_higher=100, vol_matching=True):
    """
    time series regression of return matrix against a set of factor returns

    Parameters
    ----------
    ret: DataFrame, T X N
    factor_returns: DataFrame, T x N
    bmk: benchmark returns, DataFrame, T x N, default None
    look_back: number of days used to compute regression, default 63 days
    lag: return series lag
    r_lower: lower bound for return filter, percentage, min 0, max 100
    r_higher: lower bound for return filter, percentage, min 0, max 100
    fr_lower: lower bound for factor return filter, percentage, min 0, max 100
    fr_higher: higher bound for factor return filter, percentage, min 0, max 100
    vol_matching: default True, match to independent variable series vol level by projected series's vol

    Returns
    -------

    """
    if isinstance(ret, list):
        ret = np.array(ret)
    if ret.ndim == 1:
        ret = ret.reshape((len(ret), 1))
    if isinstance(factor_returns, list):
        factor_returns = np.array(factor_returns)
    if factor_returns.ndim == 1:
        factor_returns = factor_returns.reshape((len(factor_returns), 1))
    factor_returns = rt.winsorize(factor_returns, fr_lower, fr_higher)
    if bmk is None:
        act = ret.astype('float64').copy()
    else:
        if isinstance(bmk, list):
            bmk = np.array(bmk)
        if bmk.ndim == 1:
            bmk = bmk.reshape((len(bmk), 1))
        act = ret - np.tile(bmk, (1, ret.shape[1]))
    act = rt.winsorize(act, r_lower, r_higher)
    if not isinstance(look_back, numbers.Number):
        raise ValueError(f"look back must be a number")
    if look_back <= 0:
        raise ValueError(f"look back must be a positive number")
    if lag is None or not isinstance(lag, numbers.Number) or lag < 0:
        lag = 0
    if look_back + lag > ret.shape[0]:
        display(f"Not long enough of timeseries for regression exposures")
        return None
    num = ret.shape[0] - look_back - lag + 1
    result = [None] * ret.shape[1]
    for i in range(ret.shape[1]):
        result[i] = np.empty((num, factor_returns.shape[1]))
    for i in range(num):
        m = act[i:i + look_back, :]
        f = factor_returns[i:i + look_back, :]
        temp_val = np.concatenate((m, f), axis=1)
        ix = np.where(pd.isnull(temp_val).sum(axis=1) == 0)[0]
        if len(ix) < 3:
            util.display(f"No. {i}-th: No sufficient data for regression: continue")
            continue
        m = m[ix, :]
        f = f[ix, :]
        del temp_val, ix
        b = np.linalg.lstsq(f, m, rcond=None)
        b = b[0]
        for j in range(ret.shape[1]):
            try:
                result[j][i, :] = np.atleast_2d(b[:, j])
                if vol_matching:
                    fac = np.nanstd(m[:, j])
                    fac = fac / np.nanstd(np.matmul(f, result[j][[i], :].T))
                    result[j][i, :] = fac * result[j][i, :]
            except Exception as ee:
                display(ee)
                display(f" Error: Time: {i}; Regression {j}")
        if np.mod(i, 500) == 0:
            display(f"Time No. {i} regression")
    return result


def factor_exposure_ts(ret, factor_group, print_report=False, file=None, calendar_str='US', look_back=126, lag=0,
                       r_lower=0, r_higher=100, fr_lower=0, fr_higher=100, vol_matching=True, factors=None):
    """
    Return timeseries based exposure timeseries estimation. Using OLS regression with a flexible look-back periods
    Parameters
    ----------
    ret: DataFrame, T x N dimension
    factor_group: either factor group's name, or factor returns DataFrame, T x N dimension
    print_report: default False
    file: default None, alternative output file name
    calendar_str
    look_back
    lag
    r_lower
    r_higher
    fr_lower
    fr_higher
    vol_matching
    factors: default None

    Returns
    -------

    """
    dates = ret.index.to_numpy()
    dates = util.load_business_days(calendar_str, dates[0], dates[-1])
    ret = ret.loc[dates]
    if isinstance(factor_group, pd.DataFrame):
        r = factor_group.copy().astype('float64')
    else:
        fg = root.load_object(factor_group)
        r = fg.load_factor_returns(dates[0], dates[-1])
        r = r['values'][0]['values']
        display(f"{len(r.index)} X {len(r.columns)} factor returns ({fg.name}): loaded")
    r = r.loc[dates]
    if factors is not None:
        if isinstance(factors, str):
            factors = np.array([factors])
        elif isinstance(factors, list):
            factors = np.array(factors)
        if len(factors) == 0:
            display(f"Requested factors is empty: return None")
            return None
        ix = np.where(np.isin(r.columns, factors))[0]
        if len(ix) == 0:
            display(f"No factors requested found in factor return series: returning None")
            return None
        r = r[r.columns[ix]]
        display(f"{len(ix)} out of {len(factors)} factors found in the factor return series")

    ets = regression_ts(ret.to_numpy(), r.to_numpy().astype('float64'), look_back=look_back, lag=lag,
                        r_lower=r_lower, r_higher=r_higher, fr_lower=fr_lower, fr_higher=fr_higher)
    days = dates[look_back + lag - 1:]
    if print_report:
        if file is None:
            file = os.path.join(util.default_output_location('reports'), 'tmp',
                                f'factor_exposure.{dates[0].strftime(util.yyyymmdd_format)}'
                                f'.{dates[-1].strftime(util.yyyymmdd_format)}.{util.clock()}.xlsx')
        for ix, c in enumerate(ret.columns):
            sheet = c
            exist = os.path.exists(file)
            try:
                df = pd.DataFrame(ets[ix], index=days, columns=r.columns)
                if exist:
                    with pd.ExcelWriter(file, engine='openpyxl', mode='a', if_sheet_exists="replace") as writer:
                        df.to_excel(writer, sheet_name=sheet)
                else:
                    with pd.ExcelWriter(file, engine='openpyxl', mode='w') as writer:
                        df.to_excel(writer, sheet_name=sheet)
                display(f"Successfully added {sheet} to {file}")
            except IOError as ioe:
                display(ioe)
                display(f"Unable to output to {sheet} due to IO Error")
            except Exception as ee:
                display(ee)
                display(f"Unable to output to {sheet} due to IO Error")

    result = []
    for ix, z in enumerate(ets):
        result = result + [pd.DataFrame(z, index=dates[look_back + lag - 1:], columns=r.columns)]
    return result

# --------------------------------------------------------------
#
# Exposure Attributions
#
# --------------------------------------------------------------


def exposure_attribution(start_date, end_date, portfolio, benchmark, factors, calendar_str=None,
                         managed_wt_flag=None, benchmark_wt_flag=None, caching_frequency='MONTHEND',
                         factor_frequency='MONTHEND', buckets=5, print_report=False, email_address=None, prod=False):
    """

    :param start_date:
    :param end_date:
    :param portfolio:
    :param benchmark:
    :param factors:
    :param calendar_str:
    :param managed_wt_flag:
    :param benchmark_wt_flag:
    :param caching_frequency:
    :param factor_frequency:
    :param buckets:
    :param print_report:
    :param email_address:
    :param prod:
    :return:
    """
    if calendar_str is None or not isinstance(calendar_str, str):
        calendar_str = 'GL'
    if print_report is None or not isinstance(print_report, bool):
        print_report = False
    if email_address is None or not isinstance(email_address, str):
        email_address = None
    bus_days = util.load_business_days(calendar_str, start_date, end_date)
    if len(bus_days) == 0:
        print(f"no valid business requested, returning ...")
        return None
    if portfolio is None:
        print(f"No valid portfolio requested, returning ...")
        return None
    if benchmark is None:
        print(f"No valid benchmark requested, returning ...")
        return None
    if factors is None or len(factors) == 0:
        print(f"No valid grouping factors")
        return None
    if managed_wt_flag is None or not isinstance(managed_wt_flag, str):
        managed_wt_flag = port.get_default_weighting_method(portfolio)
    if benchmark_wt_flag is None or not isinstance(benchmark_wt_flag, str):
        benchmark_wt_flag = port.get_default_weighting_method(benchmark)
    if caching_frequency is None:
        caching_frequency = 'MONTHEND'
    if not isinstance(caching_frequency, str):
        caching_frequency = 'MONTHEND'
    caching_frequency = caching_frequency.upper().strip()
    if prod:
        env = 'PROD'
    else:
        env = 'DEV'
    output_location = os.path.join(util.default_output_location('reports', env), 'ea',
                                   f"{portfolio}", f"{benchmark}")
    if not util.exists(output_location):
        util.makedirs(output_location)
        print(f"created: {output_location}")
    df = ea(start_date, end_date, portfolio, benchmark, factors, calendar_str=calendar_str,
            managed_wt_flag=managed_wt_flag, benchmark_wt_flag=benchmark_wt_flag, output_location=output_location,
            caching_frequency=caching_frequency, factor_frequency=factor_frequency, buckets=buckets)
    result = aggregate_period_contributions(df)
    result['managed portfolio'] = portfolio
    result['managed weighting method'] = managed_wt_flag
    result['benchmark portfolio'] = benchmark
    result['benchmark weighting method'] = benchmark_wt_flag
    if print_report:
        export_ea_report(result, output_location)
    return result


def ea(start_date, end_date, portfolio, benchmark, factors, calendar_str=None,
       managed_wt_flag=None, benchmark_wt_flag=None, output_location=None,
       caching_frequency='MONTHEND', factor_frequency='MONTHEND', value_type='EXPOSURE',
       buckets=5, prod=False):
    """
    read from cache or compute if missing
    :param start_date:
    :param end_date:
    :param portfolio:
    :param benchmark:
    :param factors:
    :param calendar_str:
    :param managed_wt_flag:
    :param benchmark_wt_flag:
    :param output_location: caching location
    :param caching_frequency: default 'MONTHEND
    :param factor_frequency: default 'MONTHEND
    :param value_type: default 'EXPOSURE
    :param buckets: default 5
    :param prod: default False
    :return:
    """
    if isinstance(factors, str):
        factors = np.array([factors])
    bus_days = util.load_business_days(calendar_str, start_date, end_date)
    bus_days = bus_days[bus_days <= util.prior_day(calendar_str)]
    if len(bus_days) == 0:
        print(f"Requested dates all after current day; returning")
        return None
    if caching_frequency is None:
        caching_frequency = 'MONTHEND'
    if not isinstance(caching_frequency, str):
        caching_frequency = 'MONTHEND'
    caching_frequency = caching_frequency.upper().strip()
    if factors is None:
        print(f"No valid factor for exposure attributions provided")
        return None
    cache_days = util.get_period_start_end(bus_days[0], bus_days[-1], caching_frequency, calendar_str)
    if output_location is None:
        if prod:
            env = 'PROD'
        else:
            env = 'DEV'
        output_location = os.path.join(util.default_output_location('reports', env), 'ea', f"{portfolio}", f"{benchmark}")
    if not os.path.exists(output_location):
        os.makedirs(output_location)
        print(f"Created EA report caching directory: {output_location}")
    for k in cache_days.index:
        cache_start = cache_days.loc[k, 'from']
        cache_end = cache_days.loc[k, 'to']
        b_days = bus_days[np.logical_and(bus_days >= cache_start, bus_days <= cache_end)]
        s_dates = np.full((len(factors), 1), b_days[0])
        e_dates = np.full((len(factors), 1), b_days[-1])
        compute_flags = np.full((len(factors), 1), True)
        for idx, fac in enumerate(factors):
            file = os.path.join(output_location, f"{portfolio}_vs_{benchmark}"
                                                 f"_{managed_wt_flag}_{benchmark_wt_flag}"
                                                 f"_{fac}"
                                                 f"_{cache_start.strftime(util.yyyymmdd_format)}"
                                                 f"_{cache_end.strftime(util.yyyymmdd_format)}.qd")
            if util.exists(file):
                cache = util.load_data(file)
                if cache is not None and isinstance(cache, dict) and 'dates' in cache:
                    missing_dates = np.setdiff1d(b_days, cache['dates'])
                    if len(missing_dates) == 0:
                        compute_flags[idx] = False
                        continue
                    s_dates[idx] = np.min(missing_dates)
                    e_dates[idx] = np.max(missing_dates)
        if compute_flags.sum() > 0:
            r_start_date = np.min(s_dates)
            r_end_date = np.max(e_dates)
            p_start_date = util.previous_business_days(r_start_date, calendar_str)
            p_end_date = util.previous_business_days(r_end_date, calendar_str)
            por = port.get_portfolio_weights(p_start_date, p_end_date, portfolio,
                                             weight_flag=managed_wt_flag, calendar_str=calendar_str)
            bench = port.get_portfolio_weights(p_start_date, p_end_date, benchmark,
                                               weight_flag=benchmark_wt_flag, calendar_str=calendar_str)
            sec_ids = np.union1d(por.columns.to_numpy(), bench.columns.to_numpy())
            ret = md.get_returns(r_start_date, r_end_date, sec_ids, calendar_str=calendar_str)
            for idx, fac in enumerate(factors):
                if not compute_flags[idx]:
                    continue
                f_obj = root.load_object(fac)
                if f_obj.group_flag:
                    snap = f_obj.snapshot(expand_flag=True)
                    groups = snap['factors']
                s_date = s_dates[idx][0]
                e_date = e_dates[idx][0]
                group_result = compute_factor_contributions(s_date, e_date, por, bench, ret, fac, buckets, calendar_str)
                display(f"Exposure Attribution: {fac}: {portfolio} vs {benchmark} "
                        f"result computed for: {s_date} - {e_date}")
                file = os.path.join(output_location, f"{portfolio}_vs_{benchmark}"
                                                     f"_{managed_wt_flag}_{benchmark_wt_flag}"
                                                     f"_{fac}"
                                                     f"_{cache_start.strftime(util.yyyymmdd_format)}"
                                                     f"_{cache_end.strftime(util.yyyymmdd_format)}.qd")
                try:
                    group_result = save_and_merge_contributions(group_result, file)
                    print(f"Exposure Attributions {fac}: {portfolio} vs {benchmark} "
                          f"result cached to: {cache_start} - {cache_end}\n{file}")
                except IOError as ioe:
                    display(ioe)
                    util.save_data(group_result, file)
                except ValueError as ve:
                    display(ve)
                    util.save_data(group_result, file)
    result = {'groups': factors, 'values': np.full((len(factors), 1), None)}
    for k in cache_days.index:
        cache_start = cache_days.loc[k, 'from']
        cache_end = cache_days.loc[k, 'to']
        for idx, fac in enumerate(factors):
            file = os.path.join(output_location, f"{portfolio}_vs_{benchmark}"
                                                 f"_{managed_wt_flag}_{benchmark_wt_flag}"
                                                 f"_{fac}"
                                                 f"_{cache_start.strftime(util.yyyymmdd_format)}"
                                                 f"_{cache_end.strftime(util.yyyymmdd_format)}.qd")
            try:
                data = util.load_data(file)
                result['values'][idx] = merge_contributions(result['values'][idx], data)
                print(f"Exposure Attribution {fac}: {portfolio} vs {benchmark} "
                      f"loaded result from cache: {cache_start} - {cache_end}")
            except IOError:
                print(IOError)
                print(f"Unable to load results for {fac}: {cache_start} - {cache_end}")
    for idx, fac in enumerate(factors):
        result['values'][idx] = filter_contributions_by_dates(result['values'][idx], bus_days)
    return result


def compute_factor_contributions(start_date, end_date, por, ben, ret, fac, buckets,
                                 calendar_str=None, factor_frequency='MONTHEND', value_type='EXPOSURE',
                                 universe=33):
    bus_days = util.load_business_days(calendar_str, start_date, end_date)
    pos_days = util.previous_business_days(bus_days, calendar_str)
    r = ret.loc[bus_days]
    p = por.loc[pos_days]
    b = ben.loc[pos_days]
    a = port.active_portfolio(p, b)
    sec_ids = a.columns.to_numpy()

    # managed contributions
    rr, pp = port.align_portfolio_and_return(p, r, calendar_str, sec_ids)
    mc = pd.DataFrame(pp.to_numpy() * rr.to_numpy(), index=bus_days, columns=sec_ids)  # managed contributions
    # benchmark contributions
    rr, bb = port.align_portfolio_and_return(b, r, calendar_str, sec_ids)
    bc = pd.DataFrame(bb.to_numpy() * rr.to_numpy(), index=bus_days, columns=sec_ids)  # benchmark contributions
    # active contributions
    rr, aw = port.align_portfolio_and_return(a, r, calendar_str, sec_ids)
    ac = pd.DataFrame(aw.to_numpy() * rr.to_numpy(), index=bus_days, columns=sec_ids)  # active contributions

    # load exposures
    fac_obj = root.load_object(fac)
    if fac_obj is None:
        print(f"No valid factor for grouping method supplied")
        return None
    snapshots = fac_obj.snapshot(expand_flag=True)
    if fac_obj.group_flag:
        groups = snapshots['factors']
    else:
        if universe is None:
            universe = fac_obj.universe
        groups = []
        for ix in range(buckets):
            groups = groups + [f"{fac_obj.name}_{ix+1}"]
    groups = np.sort(np.array(groups))
    groups = np.concatenate((groups, ['NA']))
    mr = pd.DataFrame(np.nansum(mc, axis=1), index=bus_days, columns=['values'])  # managed returns
    br = pd.DataFrame(np.nansum(bc, axis=1), index=bus_days, columns=['values'])  # benchmark returns
    ar = pd.DataFrame(np.nansum(ac, axis=1), index=bus_days, columns=['values'])  # active returns
    mg = pd.DataFrame(0, index=bus_days, columns=groups)  # managed group returns
    bg = pd.DataFrame(0, index=bus_days, columns=groups)  # benchmark group returns
    ag = pd.DataFrame(0, index=bus_days, columns=groups)  # active group returns
    mgc = pd.DataFrame(0, index=bus_days, columns=groups)  # managed group return contributions
    bgc = pd.DataFrame(0, index=bus_days, columns=groups)  # benchmark group return contributions
    agc = pd.DataFrame(0, index=bus_days, columns=groups)  # active group return contributions
    ss = pd.DataFrame(0, index=bus_days, columns=groups)  # stock selections
    aa = pd.DataFrame(0, index=bus_days, columns=groups)  # asset allocation
    ie = pd.DataFrame(0, index=bus_days, columns=groups)  # interactive effects
    ww = pd.DataFrame(0, index=pos_days, columns=groups)  # managed weights in each group
    wb = pd.DataFrame(0, index=pos_days, columns=groups)  # benchmark weights in each group
    wa = pd.DataFrame(0, index=pos_days, columns=groups)  # active weights in each group

    for idx, d in enumerate(pos_days):
        rd = bus_days[idx]
        b_ret = br.loc[rd, 'values']
        v_d = util.previous_day(rd, calendar_str, factor_frequency)
        sids = a.columns[np.where(a.loc[d] != 0)[0]].to_numpy()
        if universe is not None:
            univ = port.get_cached_positions(v_d, v_d, universe)
            univ = univ.columns.to_numpy()
            uids = np.union1d(sids, univ)
        else:
            univ = None
            uids = sids
        group_secs = dict.fromkeys(groups)
        gids = np.array([])
        if fac_obj.group_flag:
            exposures = fac_obj.load_exposures(v_d, sids)
            for g in groups:
                if g == 'NA':
                    continue
                if g not in exposures.columns:
                    continue
                s = exposures.index[np.where(exposures[[g]] == 1)[0]].to_numpy()
                gids = np.union1d(s, gids)
                group_secs[g] = s
        else:
            exposures = fac_obj.load_values(value_type, v_d, v_d, uids)
            exposures = exposures.T
            mids = np.intersect1d(sids, exposures.index)
            if univ is None:
                eids = mids
            else:
                eids = np.intersect1d(univ, exposures.index)
            ep = exposures.loc[mids]
            for ix in range(buckets):
                g = f"{fac_obj.name}_{ix+1}"
                lower = np.nanpercentile(exposures.loc[eids], ix/buckets*100)
                upper = np.nanpercentile(exposures.loc[eids], (ix + 1)/buckets*100)
                if ix < buckets - 1:
                    gx = np.where(np.logical_and(ep >= lower, ep < upper))[0]
                else:
                    gx = np.where(np.logical_and(ep >= lower, ep <= upper))[0]
                s = ep.index[gx].to_numpy()
                group_secs[g] = s
                gids = np.union1d(s, gids)
        group_secs['NA'] = np.setdiff1d(sids, gids)

        for g in groups:
            s = group_secs[g]
            g_ac = ac.loc[rd, s].sum()  # contribution to active returns
            g_bc = bc.loc[rd, s].sum()  # contribution to benchmark returns
            g_mc = mc.loc[rd, s].sum()  # contributions to managed returns
            mgc.loc[rd, g] = g_mc
            bgc.loc[rd, g] = g_bc
            agc.loc[rd, g] = g_ac
            ww.loc[d, g] = pp.loc[d, s].sum()
            wb.loc[d, g] = bb.loc[d, s].sum()
            wa.loc[d, g] = ww.loc[d, g] - wb.loc[d, g]
            if ww.loc[d, g] != 0:
                r_m = g_mc / ww.loc[d, g]
            else:
                r_m = np.nan
            if wb.loc[d, g] != 0:
                r_b = g_bc / wb.loc[d, g]
            else:
                r_b = np.nan
            asset_allocation = wa.loc[d, g] * (r_b - b_ret)
            if np.isnan(r_b):
                stock_selection = wb.loc[d, g] * r_m - g_bc
            else:
                stock_selection = wb.loc[d, g] * (r_m - r_b)
            interaction = g_ac - np.nansum(asset_allocation) - np.nansum(stock_selection)
            ss.loc[rd, g] = stock_selection
            aa.loc[rd, g] = asset_allocation
            ie.loc[rd, g] = interaction
            mg.loc[rd, g] = r_m
            bg.loc[rd, g] = r_b
            ag.loc[rd, g] = r_m - r_b
        error = ar.loc[rd, 'values'] - ss.loc[rd].sum() - aa.loc[rd].sum() - ie.loc[rd].sum()
        print(f"{rd}: {fac}: Total Active {ar.loc[rd,'values']*100:.2f} %: SS {ss.loc[rd].sum()*100:.2f} %, "
              f"AA {aa.loc[rd].sum()*100:.2f} %, IE {ie.loc[rd].sum()*100:.2f} % ---"
              f"Error {error*100: .3f} %")

    result = {'dates': bus_days,
              'position dates': pos_days,
              'managed returns': mr,
              'benchmark returns': br,
              'active returns': ar,
              'managed weights': pp,
              'benchmark weights': bb,
              'active weights': pp - bb,
              'managed contributions': mc,
              'benchmark contributions': bc,
              'active contributions': ac,
              'groups': groups,
              'managed group weights': ww,
              'benchmark group weights': wb,
              'active group weights': wa,
              'managed group contributions': mgc,
              'benchmark group contributions': bgc,
              'active group contributions': agc,
              'managed group returns': mg,
              'benchmark group returns': bg,
              'active group returns': ag,
              'stock selections': ss,
              'asset allocations': aa,
              'interaction effects': ie
              }
    return result


def compute_group_contributions(start_date, end_date, por, ben, ret, group, calendar_str=None):
    bus_days = util.load_business_days(calendar_str, start_date, end_date)
    pos_days = util.previous_business_days(bus_days, calendar_str)
    r = ret.loc[bus_days]
    p = por.loc[pos_days]
    b = ben.loc[pos_days]
    a = port.active_portfolio(p, b)
    sec_ids = a.columns.to_numpy()

    # managed contributions
    rr, pp = port.align_portfolio_and_return(p, r, calendar_str, sec_ids)
    mc = pd.DataFrame(pp.to_numpy() * rr.to_numpy(), index=bus_days, columns=sec_ids)  # managed contributions
    # benchmark contributions
    rr, bb = port.align_portfolio_and_return(b, r, calendar_str, sec_ids)
    bc = pd.DataFrame(bb.to_numpy() * rr.to_numpy(), index=bus_days, columns=sec_ids)  # benchmark contributions
    # active contributions
    rr, aa = port.align_portfolio_and_return(a, r, calendar_str, sec_ids)
    ac = pd.DataFrame(aa.to_numpy() * rr.to_numpy(), index=bus_days, columns=sec_ids)  # active contributions

    # load exposures
    group_obj = root.load_object(group)
    if group_obj is None:
        display(f"No valid grouping method supplied")
        return None
    snapshots = group_obj.snapshot(expand_flag=True)
    groups = snapshots['factors']
    groups = np.sort(groups)
    groups = np.append(groups, ['NA'])

    mr = pd.DataFrame(np.nansum(mc, axis=1), index=bus_days, columns=['values'])  # managed returns
    br = pd.DataFrame(np.nansum(bc, axis=1), index=bus_days, columns=['values'])  # benchmark returns
    ar = pd.DataFrame(np.nansum(ac, axis=1), index=bus_days, columns=['values'])  # active returns
    mg = pd.DataFrame(0, index=bus_days, columns=groups)  # managed group returns
    bg = pd.DataFrame(0, index=bus_days, columns=groups)  # benchmark group returns
    ag = pd.DataFrame(0, index=bus_days, columns=groups)  # active group returns
    mgc = pd.DataFrame(0, index=bus_days, columns=groups)  # managed group return contributions
    bgc = pd.DataFrame(0, index=bus_days, columns=groups)  # benchmark group return contributions
    agc = pd.DataFrame(0, index=bus_days, columns=groups)  # active group return contributions
    ss = pd.DataFrame(0, index=bus_days, columns=groups)  # stock selections
    aa = pd.DataFrame(0, index=bus_days, columns=groups)  # asset allocation
    ie = pd.DataFrame(0, index=bus_days, columns=groups)  # interactive effects
    ww = pd.DataFrame(0, index=pos_days, columns=groups)  # managed weights in each group
    wb = pd.DataFrame(0, index=pos_days, columns=groups)  # benchmark weights in each group
    wa = pd.DataFrame(0, index=pos_days, columns=groups)  # active weights in each group

    for idx, d in enumerate(pos_days):
        rd = bus_days[idx]
        b_ret = br.loc[rd, 'values']
        exposures = group_obj.load_exposures(d, sec_ids)
        for g in groups:
            if g in exposures.columns:
                s = exposures.index[exposures[g] > 0].to_numpy()
            else:
                s = exposures.index[exposures.sum(axis=1) == 0].to_numpy()
            if len(s) == 0:
                continue
            g_ac = ac.loc[rd, s].sum()  # contribution to active returns
            g_bc = bc.loc[rd, s].sum()  # contribution to benchmark returns
            g_mc = mc.loc[rd, s].sum()  # contributions to managed returns
            mgc.loc[rd, g] = g_mc
            bgc.loc[rd, g] = g_bc
            agc.loc[rd, g] = g_ac
            ww.loc[d, g] = pp.loc[d, s].sum()
            wb.loc[d, g] = bb.loc[d, s].sum()
            wa.loc[d, g] = ww.loc[d, g] - wb.loc[d, g]
            if ww.loc[d, g] != 0:
                r_m = g_mc / ww.loc[d, g]
            else:
                r_m = np.nan
            if wb.loc[d, g] != 0:
                r_b = g_bc / wb.loc[d, g]
            else:
                r_b = np.nan
            asset_allocation = wa.loc[d, g] * (r_b - b_ret)
            if np.isnan(r_b):
                stock_selection = wb.loc[d, g] * r_m - g_bc
            else:
                stock_selection = wb.loc[d, g] * (r_m - r_b)
            interaction = g_ac - np.nansum(asset_allocation) - np.nansum(stock_selection)
            ss.loc[rd, g] = stock_selection
            aa.loc[rd, g] = asset_allocation
            ie.loc[rd, g] = interaction
            mg.loc[rd, g] = r_m
            bg.loc[rd, g] = r_b
            ag.loc[rd, g] = r_m - r_b
        error = ar.loc[rd, 'values'] - ss.loc[rd].sum() - aa.loc[rd].sum() - ie.loc[rd].sum()
        display(f"{rd}: {group}: Total Active {ar.loc[rd,'values']*100:.2f} %: SS {ss.loc[rd].sum()*100:.2f} %, "
                f"AA {aa.loc[rd].sum()*100:.2f} %, IE {ie.loc[rd].sum()*100:.2f} % ---"
                f"Error {error*100: .3f} %")

    result = {'dates': bus_days,
              'position dates': pos_days,
              'managed returns': mr,
              'benchmark returns': br,
              'active returns': ar,
              'managed weights': pp,
              'benchmark weights': bb,
              'active weights': aa,
              'managed contributions': mc,
              'benchmark contributions': bc,
              'active contributions': ac,
              'groups': groups,
              'managed group weights': ww,
              'benchmark group weights': wb,
              'active group weights': wa,
              'managed group contributions': mgc,
              'benchmark group contributions': bgc,
              'active group contributions': agc,
              'managed group returns': mg,
              'benchmark group returns': bg,
              'active group returns': ag,
              'stock selections': ss,
              'asset allocations': aa,
              'interaction effects': ie
              }
    return result


def merge_contributions(data, result):
    if isinstance(data, np.ndarray):
        data = data[0]  # using first element only
    if data is None:
        return result
    d_dates = data['dates']
    r_dates = result['dates']
    dates = np.sort(np.union1d(d_dates, r_dates))
    q = {'dates': dates}
    pd_dates = data['position dates']
    pr_dates = result['position dates']
    pos_dates = np.sort(np.union1d(pd_dates, pr_dates))
    q['position dates'] = pos_dates
    fields = ['managed weights', 'benchmark weights', 'active weights',
              'managed contributions', 'benchmark contributions', 'active contributions']
    for f in fields:
        df = data[f].combine_first(result[f])
        df.fillna(0, inplace=True)
        q[f] = df
        del df

    fields = ['managed returns', 'benchmark returns', 'active returns']
    for f in fields:
        df = pd.DataFrame(dtype='float64', index=dates, columns=['values'])
        df.loc[d_dates, 'values'] = data[f].loc[d_dates, 'values']
        df.loc[r_dates, 'values'] = result[f].loc[r_dates, 'values']
        q[f] = df
        del df

    d_groups = data['groups']
    r_groups = result['groups']
    groups = np.sort(np.union1d(d_groups, r_groups))
    q['groups'] = groups
    fields = ['stock selections', 'asset allocations', 'interaction effects',
              'managed group returns', 'benchmark group returns',
              'active group returns', 'managed group contributions',
              'benchmark group contributions', 'active group contributions']
    for f in fields:
        df = pd.DataFrame(0.0, index=dates, columns=groups)
        df.loc[d_dates, d_groups] = data[f].loc[d_dates, d_groups]
        df.loc[r_dates, r_groups] = result[f].loc[r_dates, r_groups]
        q[f] = df
        del df

    fields = ['managed group weights', 'benchmark group weights', 'active group weights']
    for f in fields:
        df = pd.DataFrame(0.0, index=pos_dates, columns=groups)
        df.loc[pd_dates, d_groups] = data[f].loc[pd_dates, d_groups]
        df.loc[pr_dates, r_groups] = result[f].loc[pr_dates, r_groups]
        q[f] = df
        del df
    return q


def save_and_merge_contributions(result, file):
    if not os.path.exists(file):
        util.save_data(result, file)
        display(f"variance analysis successfully saved to \n{file}")
        return result
    data = util.load_data(file)
    q = merge_contributions(data, result)
    util.save_data(q, file)
    display(f"Merged {len(result['dates'])} days to prior {len(data['dates'])} days results")
    display(f"variance analysis successfully saved to \n{file}")
    return q


def filter_contributions_by_dates(data, bus_days=None):
    if bus_days is None:
        return data
    if isinstance(bus_days, str):
        bus_days = np.array([bus_days])
    if isinstance(data, np.ndarray):
        data = data[0]
    bus_days = np.unique(bus_days)
    # dates and position dates
    days, ia, ib = intersect(data['dates'], bus_days)
    data['dates'] = days
    data['position dates'] = data['position dates'][ia]
    # the rest
    for key in data.keys():
        if key in ['groups', 'dates', 'position dates']:
            continue
        df = data[key]
        if isinstance(df, pd.DataFrame):
            if 'weights' in key:
                dates = df.index[df.index.isin(data['position dates'])]
            else:
                dates = df.index[df.index.isin(bus_days)]
            df = df.loc[dates]
        else:
            display(f"Not supported type")
        data[key] = df
    return data


def aggregate_period_contributions(data):
    groups = data['groups']
    data['period'] = np.full((len(groups), 1), None)
    data['risk'] = np.full((len(groups), 1), None)
    for idx, group in enumerate(groups):
        p = {}
        risk = {}
        val = data['values'][idx][0]
        p['sec_ids'] = val['active contributions'].columns.to_numpy()
        p['managed average weights'] = val['managed weights'].mean(axis=0).to_numpy()
        p['benchmark average weights'] = val['benchmark weights'].mean(axis=0).to_numpy()
        p['active average weights'] = p['managed average weights'] - p['benchmark average weights']
        # managed russell linking factor
        m_link = rt.russell_linking_factors(val['managed returns'].to_numpy())
        b_link = rt.russell_linking_factors(val['benchmark returns'].to_numpy())
        a_link = rt.russell_linking_factors(val['active returns'].to_numpy())
        d_link = m_link - b_link
        num_of_groups = val['asset allocations'].shape[1]
        risk['managed volatility'] = np.nanstd(val['managed returns'].to_numpy().astype('float64'))*np.sqrt(252)
        risk['benchmark volatility'] = np.nanstd(val['benchmark returns'].to_numpy().astype('float64'))*np.sqrt(252)
        risk['tracking error'] = np.nanstd(val['active returns'].to_numpy().astype('float64')) * np.sqrt(252)
        p['groups'] = val['groups']
        risk['groups'] = val['groups']
        p['managed return'] = np.matmul(m_link.T, val['managed returns'].to_numpy())
        p['benchmark return'] = np.matmul(b_link.T, val['benchmark returns'].to_numpy())
        p['active return'] = p['managed return'] - p['benchmark return']
        val['stock selections'][np.isnan(val['stock selections'])] = 0
        p['stock selection'] = np.matmul(m_link.T, val['stock selections'].to_numpy())
        val['asset allocations'][np.isnan(val['asset allocations'])] = 0
        p['asset allocation'] = np.matmul(m_link.T, val['asset allocations'].to_numpy()) +\
                                np.matmul(d_link.T, val['benchmark returns'].to_numpy()) / num_of_groups
        p['interaction effect'] = np.matmul(m_link.T, val['interaction effects'].to_numpy())
        p['managed contribution'] = np.matmul(m_link.T, val['managed contributions'].to_numpy())
        p['benchmark contribution'] = np.matmul(m_link.T, val['benchmark contributions'].to_numpy())
        p['active contribution'] = p['managed contribution'] - p['benchmark contribution']
        p['managed group contribution'] = np.matmul(m_link.T, val['managed group contributions'].to_numpy())
        p['benchmark group contribution'] = np.matmul(b_link.T, val['benchmark group contributions'].to_numpy())
        p['active group contribution'] = p['asset allocation'] + p['stock selection'] + p['interaction effect']
        contrib = val['managed group contributions'].to_numpy()
        c2 = np.nansum(np.matmul(contrib.T, contrib), axis=1).T
        c2 = c2.reshape((1, len(c2)))
        r2 = np.nansum(val['managed returns'].to_numpy() * val['managed returns'].to_numpy())
        risk['managed rcte'] = c2 / r2
        contrib = val['benchmark group contributions'].to_numpy()
        c2 = np.nansum(np.matmul(contrib.T, contrib), axis=1).T
        c2 = c2.reshape((1, len(c2)))
        r2 = np.nansum(val['benchmark returns'].to_numpy() * val['benchmark returns'].to_numpy())
        risk['benchmark rcte'] = c2 / r2
        contrib = val['active group contributions'].to_numpy()
        c2 = np.nansum(np.matmul(contrib.T, contrib), axis=1).T
        c2 = c2.reshape((1, len(c2)))
        r2 = np.nansum(val['active returns'].to_numpy() * val['active returns'].to_numpy())
        risk['active rcte'] = c2 / r2
        p['managed group weight average'] = np.nanmean(val['managed group weights'], axis=0)
        p['managed group weight std'] = np.nanstd(val['managed group weights'], axis=0)
        p['benchmark group weight average'] = np.nanmean(val['benchmark group weights'], axis=0)
        p['benchmark group weight std'] = np.nanstd(val['benchmark group weights'], axis=0)
        p['active group weight average'] = np.nanmean(val['active group weights'], axis=0)
        p['active group weight std'] = np.nanstd(val['active group weights'], axis=0)
        data['period'][idx] = p
        data['risk'][idx] = risk
    return data


def export_ea_report(data, output_location):
    managed = data['managed portfolio']
    if isinstance(managed, int):
        m_ref = md.get_portfolio_references(managed)
        managed = m_ref['ReferenceCode'].iloc[0]
    benchmark = data['benchmark portfolio']
    if isinstance(benchmark, int):
        b_ref = md.get_portfolio_references(benchmark)
        benchmark = b_ref['ReferenceCode'].iloc[0]
    man_wt_flag = data['managed weighting method']
    ben_wt_flag = data['benchmark weighting method']
    v = data['values'][0][0]
    dates = v['dates']
    groups = data['groups']
    sorted_groups = np.array(groups)
    sorted_groups.sort()
    file_name = f"{data['managed portfolio']}_vs_{data['benchmark portfolio']}_{man_wt_flag}_{ben_wt_flag}"
    for sg in sorted_groups:
        file_name = file_name + f"_{sg}"
    file_name = file_name + f"_{dates[0].strftime(util.yyyymmdd_format)}" \
                            f"_{dates[-1].strftime(util.yyyymmdd_format)}.xlsx"
    file = os.path.join(output_location, file_name)
    vp = data['period']
    risk = data['risk']
    if os.path.exists(file):
        util.preserve_file(file)

    # printing
    wb = pxl.Workbook()
    for idx, group in enumerate(groups):
        g_obj = root.load_object(group)
        name = g_obj.name
        p = vp[idx][0]
        r = risk[idx][0]
        if idx == 0:
            sheet = wb.active
            sheet.title = name
        else:
            sheet = wb.create_sheet(name, 2 * idx)
        detail = wb.create_sheet(f"Detail - {name}", 2 * idx + 1)
        display(f"exporting exposure attributions on: '{name}'")
        row = 1
        col = 1
        sheet.cell(row, col).value = 'Exposure Attributions - Realized Risk Analysis'
        sheet.cell(row, col).font = styles.Font(bold=True, color='0000FF')
        sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 4)

        row = row + 1
        sheet.cell(row, col).value = 'Start Date'
        sheet.cell(row, col+1).value = dates[0].strftime(util.YY_MM_DD_format)
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = 'End Date'
        sheet.cell(row, col+1).value = dates[-1].strftime(util.YY_MM_DD_format)
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = 'Factor'
        sheet.cell(row, col+1).value = name
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = 'Managed'
        sheet.cell(row, col+1).value = managed
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col + 1).font = styles.Font(bold=True)
        row = row + 1
        sheet.cell(row, col).value = 'Benchmark'
        sheet.cell(row, col+1).value = benchmark
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+1).font = styles.Font(bold=True)
        row = row + 3
        sheet.cell(row, col).value = 'Summary'
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col).fill = fill_pale_green
        sheet.cell(row, col).border = border_bottom
        sheet.cell(row, col+1).value = 'Return'
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+1).fill = fill_pale_green
        sheet.cell(row, col+1).border = border_bottom
        sheet.cell(row, col+2).value = 'Risk'
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+2).fill = fill_pale_green
        sheet.cell(row, col+2).border = border_bottom
        row = row + 1
        sheet.cell(row, col).value = managed
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+1).value = p['managed return'][0][0]
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+2).value = r['managed volatility']
        sheet.cell(row, col+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = benchmark
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+1).value = p['benchmark return'][0][0]
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+2).value = r['benchmark volatility']
        sheet.cell(row, col+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = 'Active'
        sheet.cell(row, col+1).value = p['active return'][0][0]
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+2).value = r['tracking error']
        sheet.cell(row, col+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='center')
        if abs(sheet.cell(row, col + 2).value) > 0.05:
            sheet.cell(row, col + 2).font = styles.Font(bold=True, color='FF0000')
            sheet.cell(row, col + 2).fill = fill_yellow

        groups = p['groups']

        row = row + 2
        sheet.cell(row, col).value = 'Weight'
        sheet.cell(row, col).fill = fill_pale_green
        sheet.cell(row, col).border = border_bottom
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+1).value = 'Total'
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+1).fill = fill_pale_green
        sheet.cell(row, col+1).border = border_bottom
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+2).value = g
            sheet.cell(row, col+gdx+2).alignment = styles.Alignment(horizontal='center')
            sheet.cell(row, col+gdx+2).fill = fill_pale_green
            sheet.cell(row, col+gdx+2).border = border_bottom
        row = row + 1
        sheet.cell(row, col).value = managed
        sheet.cell(row, col+1).value = np.nansum(p['managed group weight average'])
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+2).value = p['managed group weight average'][gdx]
            sheet.cell(row, col+gdx+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col+gdx+2).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = benchmark
        sheet.cell(row, col+1).value = np.nansum(p['benchmark group weight average'])
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+2).value = p['benchmark group weight average'][gdx]
            sheet.cell(row, col+gdx+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col+gdx+2).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = 'Active'
        sheet.cell(row, col+1).value = np.nansum(p['active group weight average'])
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+2).value = p['active group weight average'][gdx]
            sheet.cell(row, col+gdx+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col+gdx+2).alignment = styles.Alignment(horizontal='center')
            if abs(sheet.cell(row, col+gdx+2).value) > 0.05:
                sheet.cell(row, col+gdx+2).font = styles.Font(bold=True, color='FF0000')
                sheet.cell(row, col+gdx+2).fill = fill_yellow
        # performance
        row = row + 2
        sheet.cell(row, col).value = 'Performance'
        sheet.cell(row, col).fill = fill_pale_green
        sheet.cell(row, col).border = border_bottom
        sheet.cell(row, col+1).value = 'Total'
        sheet.cell(row, col+1).fill = fill_pale_green
        sheet.cell(row, col+1).border = border_bottom
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+2).value = g
            sheet.cell(row, col+gdx+2).alignment = styles.Alignment(horizontal='center')
            sheet.cell(row, col+gdx+2).fill = fill_pale_green
            sheet.cell(row, col+gdx+2).border = border_bottom
        row = row + 1
        sheet.cell(row, col).value = 'Stock Selection'
        sheet.cell(row, col+1).value = np.nansum(p['stock selection']) + np.nansum(p['interaction effect'])
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+2).value = np.nansum(p['stock selection'][0][gdx]) + \
                                               np.nansum(p['interaction effect'][0][gdx])
            sheet.cell(row, col+gdx+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col+gdx+2).alignment = styles.Alignment(horizontal='center')
            if sheet.cell(row, col + gdx + 2).value > 0.02:
                sheet.cell(row, col + gdx + 2).font = styles.Font(bold=True, color='0000FF')
                sheet.cell(row, col + gdx + 2).fill = fill_green
            if sheet.cell(row, col + gdx + 2).value < -0.02:
                sheet.cell(row, col + gdx + 2).font = styles.Font(bold=True, color='FF0000')
                sheet.cell(row, col + gdx + 2).fill = fill_yellow
        row = row + 1
        sheet.cell(row, col).value = 'Asset Allocation'
        sheet.cell(row, col+1).value = np.nansum(p['asset allocation'])
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+2).value = np.nansum(p['asset allocation'][0][gdx])
            sheet.cell(row, col+gdx+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col+gdx+2).alignment = styles.Alignment(horizontal='center')
            if sheet.cell(row, col + gdx + 2).value > 0.02:
                sheet.cell(row, col + gdx + 2).font = styles.Font(bold=True, color='0000FF')
                sheet.cell(row, col + gdx + 2).fill = fill_green
            if sheet.cell(row, col + gdx + 2).value < -0.02:
                sheet.cell(row, col + gdx + 2).font = styles.Font(bold=True, color='FF0000')
                sheet.cell(row, col + gdx + 2).fill = fill_yellow

        # row = row + 1
        # sheet.cell(row, col).value = 'Interaction'
        # sheet.cell(row, col+1).value = np.nansum(p['interaction effect'])
        # sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        # sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        # for gdx, g in enumerate(groups):
        #     sheet.cell(row, col+gdx+2).value = np.nansum(p['interaction effect'][0][gdx])
        #     sheet.cell(row, col+gdx+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        #     sheet.cell(row, col+gdx+2).alignment = styles.Alignment(horizontal='center')
        #     if sheet.cell(row, col + gdx + 2).value > 0.02:
        #         sheet.cell(row, col + gdx + 2).font = styles.Font(bold=True, color='0000FF')
        #         sheet.cell(row, col + gdx + 2).fill = fill_green
        #     if sheet.cell(row, col + gdx + 2).value < -0.02:
        #         sheet.cell(row, col + gdx + 2).font = styles.Font(bold=True, color='FF0000')
        #         sheet.cell(row, col + gdx + 2).fill = fill_yellow
        row = row + 1
        sheet.cell(row, col).value = 'Total'
        sheet.cell(row, col+1).value = np.nansum(p['interaction effect'])\
                                       +np.nansum(p['asset allocation'])\
                                       +np.nansum(p['stock selection'])
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+2).value = np.nansum(p['active group contribution'][0][gdx])
            sheet.cell(row, col+gdx+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col+gdx+2).alignment = styles.Alignment(horizontal='center')
            if sheet.cell(row, col+gdx+2).value > 0.02:
                sheet.cell(row, col+gdx+2).font = styles.Font(bold=True, color='0000FF')
                sheet.cell(row, col+gdx+2).fill = fill_green
            if sheet.cell(row, col+gdx+2).value < -0.02:
                sheet.cell(row, col+gdx+2).font = styles.Font(bold=True, color='FF0000')
                sheet.cell(row, col+gdx+2).fill = fill_yellow

        # risk decompositions
        row = row + 2
        sheet.cell(row, col).value = 'RCTE'
        sheet.cell(row, col).fill = fill_pale_green
        sheet.cell(row, col).border = border_bottom
        sheet.cell(row, col+1).value = 'Total'
        sheet.cell(row, col+1).fill = fill_pale_green
        sheet.cell(row, col+1).border = border_bottom
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+2).value = g
            sheet.cell(row, col+gdx+2).alignment = styles.Alignment(horizontal='center')
            sheet.cell(row, col+gdx+2).fill = fill_pale_green
            sheet.cell(row, col+gdx+2).border = border_bottom
        row = row + 1
        sheet.cell(row, col).value = managed
        sheet.cell(row, col+1).value = np.nansum(r['managed rcte'])
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+2).value = np.nansum(r['managed rcte'][0][gdx])
            sheet.cell(row, col+gdx+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col+gdx+2).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = benchmark
        sheet.cell(row, col+1).value = np.nansum(r['benchmark rcte'])
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+2).value = np.nansum(r['benchmark rcte'][0][gdx])
            sheet.cell(row, col+gdx+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col+gdx+2).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = 'Active'
        sheet.cell(row, col+1).value = np.nansum(r['active rcte'])
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+2).value = np.nansum(r['active rcte'][0][gdx])
            sheet.cell(row, col+gdx+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col+gdx+2).alignment = styles.Alignment(horizontal='center')
            if sheet.cell(row, col+gdx+2).value > 0.15:
                sheet.cell(row, col+gdx+2).font = styles.Font(bold=True, color='FF0000')
                sheet.cell(row, col+gdx+2).fill = fill_yellow
            if sheet.cell(row, col+gdx+2).value < -0.15:
                sheet.cell(row, col+gdx+2).font = styles.Font(bold=True, color='0000FF')
                sheet.cell(row, col+gdx+2).fill = fill_green
            # ---------------------------------------------------------------
            #
            #           Detail
            #
            # ---------------------------------------------------------------
            display('*' * 100)
            display(f"exporting factor attributions detail on: '{name}'")
            display('*' * 100)
            por_types = ['Portfolio', 'Benchmark', 'Active']
            portfolios = [data['managed portfolio'], data['benchmark portfolio'], '-']
            row = 1
            col = 1
            columns = ['ID', 'Ticker', 'Cusip', 'Sedol', 'Issuer', 'Exchange', 'Currency', 'Sector', 'Industry',
                       'Weight', 'Contribution', 'Weight', 'Contribution', 'Weight', 'Contribution']
            col = col + len(columns) - 6
            detail.cell(row, col).value = portfolios[0]  # result['managed_portfolio']
            detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
            detail.cell(row, col).font = styles.Font(bold=True, color='0000FF')
            detail.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
            col = col + 2
            detail.cell(row, col).value = portfolios[1]  # result['benchmark_portfolio']
            detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
            detail.cell(row, col).font = styles.Font(bold=True, color='0000FF')
            detail.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
            col = col + 2
            detail.cell(row, col).value = 'Active'
            detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
            detail.cell(row, col).font = styles.Font(bold=True, color='0000FF')
            detail.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
            row = row + 1
            col = 1
            for cc in columns:
                detail.cell(row, col).value = cc
                detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
                col = col + 1
            sec_ids = p['sec_ids']
            ref = md.get_references(sec_ids)
            cusips = md.get_cusips(sec_ids)
            sedols = md.get_sedols(sec_ids)
            ind = md.get_rbics_classification(sec_ids)
            tickers = md.get_tickers(sec_ids)
            for jx, s in enumerate(sec_ids):
                row = row + 1
                col = 1
                detail.cell(row, col).value = s
                zx = np.where(tickers['sec_id'] == s)[0]
                if len(zx) > 0:
                    detail.cell(row, col + 1).value = tickers.loc[tickers.index[zx[-1]], 'ticker_region']
                zx = np.where(cusips['sec_id'] == s)[0]
                if len(zx) > 0:
                    detail.cell(row, col + 2).value = cusips.loc[cusips.index[zx[-1]], 'cusip']
                zx = np.where(sedols['sec_id'] == s)[0]
                if len(zx) > 0:
                    detail.cell(row, col + 3).value = sedols.loc[sedols.index[zx[-1]], 'sedol']
                zx = np.where(ref['sec_id'] == s)[0]
                if len(zx) > 0:
                    detail.cell(row, col + 4).value = ref.loc[ref.index[zx[-1]], 'name']
                    detail.cell(row, col + 5).value = ref.loc[ref.index[zx[-1]], 'exchange']
                    detail.cell(row, col + 6).value = ref.loc[ref.index[zx[-1]], 'currency']
                zx = np.where(ind['sec_id'] == s)[0]
                if len(zx) > 0:
                    detail.cell(row, col + 7).value = ind.loc[ind.index[zx[-1]], 'l1_name']
                    detail.cell(row, col + 8).value = ind.loc[ind.index[zx[-1]], 'l3_name']
                detail.cell(row, col + 9).value = p['managed average weights'][jx]
                detail.cell(row, col + 9).number_format = styles.numbers.FORMAT_PERCENTAGE_00
                detail.cell(row, col + 9).alignment = styles.Alignment(horizontal='center')
                detail.cell(row, col + 10).value = p['managed contribution'][0, jx]
                detail.cell(row, col + 10).number_format = styles.numbers.FORMAT_PERCENTAGE_00
                detail.cell(row, col + 10).alignment = styles.Alignment(horizontal='center')
                detail.cell(row, col + 11).value = p['benchmark average weights'][jx]
                detail.cell(row, col + 11).number_format = styles.numbers.FORMAT_PERCENTAGE_00
                detail.cell(row, col + 11).alignment = styles.Alignment(horizontal='center')
                detail.cell(row, col + 12).value = p['benchmark contribution'][0, jx]
                detail.cell(row, col + 12).number_format = styles.numbers.FORMAT_PERCENTAGE_00
                detail.cell(row, col + 12).alignment = styles.Alignment(horizontal='center')
                detail.cell(row, col + 13).value = p['active average weights'][jx]
                detail.cell(row, col + 13).number_format = styles.numbers.FORMAT_PERCENTAGE_00
                detail.cell(row, col + 13).alignment = styles.Alignment(horizontal='center')
                detail.cell(row, col + 14).value = p['active contribution'][0, jx]
                detail.cell(row, col + 14).number_format = styles.numbers.FORMAT_PERCENTAGE_00
                detail.cell(row, col + 14).alignment = styles.Alignment(horizontal='center')

    wb.save(file)
    display(f"Successfully variance analysis report to {file}")


# styles


fill_pale_green = styles.PatternFill("solid", start_color='EEFFEE')
fill_yellow = styles.PatternFill("solid", start_color='FFFF66')
fill_green = styles.PatternFill("solid", start_color='99FF99')
font_blue = styles.Font(color=styles.colors.BLUE)
font_blue_bold = styles.Font(bold=True, color=styles.colors.BLUE)
font_red_bold = styles.Font(bold=True, color='CC0000')
border_bottom = styles.Border(bottom=styles.Side(color='000000', border_style='double'))
