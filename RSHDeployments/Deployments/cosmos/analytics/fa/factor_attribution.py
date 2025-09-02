#
# factor attribution
#
# Author: Yun Chen
# Copyright: Indigo Dao, LLC
# Date: September 15, 2022
#
import numpy as np
import pandas as pd
import util.utilities as util
from util.utilities import display
import os
from util.intersect import *
import classes.root as root
import factors.portfolio as pt
import dataloader.portfolio as port
import dataloader.market_data as md
import util.routines as rt
import openpyxl as pxl
import openpyxl.styles as styles


def factor_attribution(start_date, end_date, portfolio, benchmark, factor_group, calendar_str='US',
                       managed_wt_flag=None, benchmark_wt_flag=None, caching_frequency='MONTHEND',
                       print_report=False, email_address=None, fwd_fill_days=0, prod=False):
    """
    Factor Attribution: attribute managed and active returns by factor models
    :param start_date:
    :param end_date:
    :param portfolio:
    :param benchmark:
    :param factor_group: string
    :param calendar_str: [optional] default of factor model
    :param managed_wt_flag:
    :param benchmark_wt_flag:
    :param caching_frequency:
    :param print_report:
    :param email_address:
    :param fwd_fill_days: default 0
    :param prod: default False
    :return:

    Author    : Yun Chen
    Copyright : Indigo Dao, LLC
    Date      : August 1, 2022
    """
    if factor_group is None or not isinstance(factor_group, str):
        display(f"No valid factor model: nothing to be done, returning")
        return None
    fg = root.load_object(factor_group)
    if fg is None:
        display(f"No valid factor model: nothing to be done, returning")
        return None
    if calendar_str is None or not isinstance(calendar_str, str):
        calendar_str = fg.calendar
    if print_report is None or not isinstance(print_report, bool):
        print_report = False
    if email_address is None or not isinstance(email_address, str):
        email_address = None
    bus_days = util.load_business_days(calendar_str, start_date, end_date)
    if len(bus_days) == 0:
        display(f"no valid business requested, returning ...")
        return None
    if portfolio is None:
        display(f"No valid portfolio requested, returning ...")
        return None
    if benchmark is None:
        display(f"No valid benchmark requested, returning ...")
        return None
    if managed_wt_flag is None or not isinstance(managed_wt_flag, str):
        managed_wt_flag = port.get_default_weighting_method(portfolio)
    if benchmark_wt_flag is None or not isinstance(benchmark_wt_flag, str):
        benchmark_wt_flag = port.get_default_weighting_method(benchmark)
    if caching_frequency is None:
        caching_frequency = 'MONTHEND'
    if not isinstance(caching_frequency, str):
        caching_frequency = 'MONTHEND'
    caching_frequency = caching_frequency.upper().strip()
    if prod:
        env = 'PROD'
    else:
        env = 'DEV'
    output_location = os.path.join(util.default_output_location('reports', env), 'fa',
                                   f"{portfolio}", f"{benchmark}")
    if not os.path.exists(output_location):
        os.makedirs(output_location)
        display(f"created: {output_location}")
    if fwd_fill_days is None:
        fwd_fill_days = 0
    df = fa(start_date, end_date, portfolio, benchmark, factor_group, calendar_str=calendar_str,
            managed_wt_flag=managed_wt_flag, benchmark_wt_flag=benchmark_wt_flag, output_location=output_location,
            caching_frequency=caching_frequency, fwd_fill_days=fwd_fill_days)
    result = aggregate_period_contributions(df)
    result['managed portfolio'] = portfolio
    result['managed weighting method'] = managed_wt_flag
    result['benchmark portfolio'] = benchmark
    result['benchmark weighting method'] = benchmark_wt_flag
    if print_report:
        export_fa_report(result, output_location)
    return result


def fa(start_date, end_date, portfolio, benchmark, factor_groups, calendar_str=None,
       managed_wt_flag=None, benchmark_wt_flag=None, output_location=None,
       caching_frequency='MONTHEND', fwd_fill_days=0):
    """
    read from cache or compute if missing
    :param start_date:
    :param end_date:
    :param portfolio:
    :param benchmark:
    :param factor_groups:
    :param calendar_str: [optional] default 'GL'
    :param managed_wt_flag:
    :param benchmark_wt_flag:
    :param output_location: caching location
    :param caching_frequency: default 'MONTHEND
    :param fwd_fill_days: default 0
    :return:

    Author: Yun Chen
    Copyright: Indigo Dao, LLC
    Date: September 16, 2022
    """
    if isinstance(factor_groups, str):
        factor_groups = np.array([factor_groups])
    bus_days = util.load_business_days(calendar_str, start_date, end_date)
    bus_days = bus_days[bus_days <= util.prior_day(calendar_str)]
    if len(bus_days) == 0:
        display(f"Requested dates all after current day; returning")
        return None
    if caching_frequency is None:
        caching_frequency = 'MONTHEND'
    if not isinstance(caching_frequency, str):
        caching_frequency = 'MONTHEND'
    caching_frequency = caching_frequency.upper().strip()
    if factor_groups is None:
        display(f"No valid factor group(s) provided")
        return None
    if fwd_fill_days is None:
        fwd_fill_days = 0
    cache_days = util.get_period_start_end(bus_days[0], bus_days[-1], caching_frequency, calendar_str)
    if output_location is None:
        output_location = os.path.join(util.default_output_location('reports'), 'fa',
                                       f"{portfolio}", f"{benchmark}")
    if not os.path.exists(output_location):
        os.makedirs(output_location)
        display(f"Created FA report caching directory: {output_location}")
    for k in cache_days.index:
        cache_start = cache_days.loc[k, 'from']
        cache_end = cache_days.loc[k, 'to']
        b_days = bus_days[np.logical_and(bus_days >= cache_start, bus_days <= cache_end)]
        s_dates = np.full((len(factor_groups), 1), b_days[0])
        e_dates = np.full((len(factor_groups), 1), b_days[-1])
        compute_flags = np.full((len(factor_groups), 1), True)
        for idx, group in enumerate(factor_groups):
            file = os.path.join(output_location, f"{portfolio}_vs_{benchmark}"
                                                 f"_{managed_wt_flag}_{benchmark_wt_flag}"
                                                 f"_{group}"
                                                 f"_{cache_start.strftime(util.yyyymmdd_format)}"
                                                 f"_{cache_end.strftime(util.yyyymmdd_format)}.qd")
            if util.exists(file):
                cache = util.load_data(file)
                if cache is not None and isinstance(cache, dict) and 'dates' in cache:
                    missing_dates = np.setdiff1d(b_days, cache['dates'])
                    if len(missing_dates) == 0:
                        compute_flags[idx] = False
                        continue
                    s_dates[idx] = np.min(missing_dates)
                    e_dates[idx] = np.max(missing_dates)
        if compute_flags.sum() > 0:
            r_start_date = np.min(s_dates)
            r_end_date = np.max(e_dates)
            p_start_date = util.previous_business_days(r_start_date, calendar_str)
            p_end_date = util.previous_business_days(r_end_date, calendar_str)
            por = port.get_portfolio_weights(p_start_date, p_end_date, portfolio,
                                             weight_flag=managed_wt_flag, calendar_str=calendar_str,
                                             forward_fill_days=fwd_fill_days)
            bench = port.get_portfolio_weights(p_start_date, p_end_date, benchmark,
                                               weight_flag=benchmark_wt_flag, calendar_str=calendar_str,
                                               forward_fill_days=fwd_fill_days)
            sec_ids = np.union1d(por.columns.to_numpy(), bench.columns.to_numpy())
            ret = md.get_returns(r_start_date, r_end_date, sec_ids, calendar_str=calendar_str)
            for idx, group in enumerate(factor_groups):
                if not compute_flags[idx]:
                    continue
                s_date = s_dates[idx][0]
                e_date = e_dates[idx][0]
                group_result = compute_group_contributions(s_date, e_date, por, bench, ret, group, calendar_str)
                display(f"Factor Attribution {group}: {portfolio} vs {benchmark} "
                        f"result computed for: {s_date} - {e_date}")
                file = os.path.join(output_location, f"{portfolio}_vs_{benchmark}"
                                                     f"_{managed_wt_flag}_{benchmark_wt_flag}"
                                                     f"_{group}"
                                                     f"_{cache_start.strftime(util.yyyymmdd_format)}"
                                                     f"_{cache_end.strftime(util.yyyymmdd_format)}.qd")
                try:
                    group_result = save_and_merge_contributions(group_result, file)
                    display(f"Factor Attribution {group}: {portfolio} vs {benchmark} "
                            f"result cached to: {cache_start} - {cache_end}")
                except IOError as ioe:
                    display(f"IO Error: {cache_start} - {cache_end}: {ioe}")
                    util.save_data(group_result, file)
                except ValueError as ve:
                    display(f"Value Error: {cache_start} - {cache_end}: {ve}")
                    util.save_data(group_result, file)
                except Exception as ee:
                    display(f"Exception: {cache_start} - {cache_end}: {ee}")
                    util.save_data(group_result, file)
    result = {'factor_groups': factor_groups, 'values': np.full((len(factor_groups), 1), None)}
    for k in cache_days.index:
        cache_start = cache_days.loc[k, 'from']
        cache_end = cache_days.loc[k, 'to']
        for idx, group in enumerate(factor_groups):
            file = os.path.join(output_location, f"{portfolio}_vs_{benchmark}"
                                                 f"_{managed_wt_flag}_{benchmark_wt_flag}"
                                                 f"_{group}"
                                                 f"_{cache_start.strftime(util.yyyymmdd_format)}"
                                                 f"_{cache_end.strftime(util.yyyymmdd_format)}.qd")
            try:
                data = util.load_data(file)
                result['values'][idx] = merge_contributions(result['values'][idx], data)
                display(f"Factor Attributions {group}: {portfolio} vs {benchmark} "
                        f"loaded result from cache: {cache_start} - {cache_end}")
            except IOError as ioe:
                display(ioe)
                display(f"I/O Error: Unable to load results for {group}: {cache_start} - {cache_end}")
            except Exception as eee:
                display(eee)
                display(f"Exception: Unable to load results for {group}: {cache_start} - {cache_end}")
    for idx, group in enumerate(factor_groups):
        result['values'][idx] = filter_contributions_by_dates(result['values'][idx], bus_days)
    return result


def compute_group_contributions(start_date, end_date, por, ben, ret, group, calendar_str=None):
    bus_days = util.load_business_days(calendar_str, start_date, end_date)
    pos_days = util.previous_business_days(bus_days, calendar_str)
    r = ret.loc[bus_days]
    p = por.loc[pos_days]
    b = ben.loc[pos_days]
    a = port.active_portfolio(p, b)
    sec_ids = a.columns.to_numpy()

    # managed contributions
    rr, pp = port.align_portfolio_and_return(p, r, calendar_str, sec_ids)
    mc = pd.DataFrame(pp.to_numpy() * rr.to_numpy(), index=bus_days, columns=sec_ids)  # managed contributions
    # benchmark contributions
    rr, bb = port.align_portfolio_and_return(b, r, calendar_str, sec_ids)
    bc = pd.DataFrame(bb.to_numpy() * rr.to_numpy(), index=bus_days, columns=sec_ids)  # benchmark contributions
    # active contributions
    rr, aw = port.align_portfolio_and_return(a, r, calendar_str, sec_ids)
    ac = pd.DataFrame(aw.to_numpy() * rr.to_numpy(), index=bus_days, columns=sec_ids)  # active contributions

    # load exposures
    group_obj = root.load_object(group)
    if group_obj is None:
        display(f"No valid factor group supplied")
        raise ValueError(f"No valid factor group provided: check factor group name and definition")
    snapshots = group_obj.snapshot(expand_flag=True)
    groups = snapshots['factors']
    types = snapshots['factor_types']
    themes = snapshots['factor_themes']

    # load factor returns
    fr = group_obj.load_factor_returns(bus_days[0], bus_days[-1], calendar_str=calendar_str)
    fr = fr['values'][0]

    # load residual returns
    res = group_obj.load_residuals(bus_days[0], bus_days[-1], sec_ids, calendar_str=calendar_str)
    res = res['values'][0]

    # managed residual contributions
    rs, ps = port.align_portfolio_and_return(p, res, calendar_str, sec_ids)
    mrc = pd.DataFrame(ps.to_numpy() * rs.to_numpy(), index=bus_days, columns=sec_ids)
    del (rs, ps)
    # benchmark residual contributions
    rs, ps = port.align_portfolio_and_return(b, res, calendar_str, sec_ids)
    brc = pd.DataFrame(ps.to_numpy() * rs.to_numpy(), index=bus_days, columns=sec_ids)
    del (rs, ps)
    # active residual contributions
    rs, ps = port.align_portfolio_and_return(a, res, calendar_str, sec_ids)
    arc = pd.DataFrame(ps.to_numpy() * rs.to_numpy(), index=bus_days, columns=sec_ids)
    del (rs, ps)

    mr = pd.DataFrame(np.nansum(mc, axis=1), index=bus_days, columns=['values'])  # managed returns
    br = pd.DataFrame(np.nansum(bc, axis=1), index=bus_days, columns=['values'])  # benchmark returns
    ar = pd.DataFrame(np.nansum(ac, axis=1), index=bus_days, columns=['values'])  # active returns
    me = pd.DataFrame(np.nansum(ac, axis=1), index=bus_days, columns=['values'])  # error in managed returns
    be = pd.DataFrame(np.nansum(ac, axis=1), index=bus_days, columns=['values'])  # error in benchmark returns
    ae = pd.DataFrame(np.nansum(ac, axis=1), index=bus_days, columns=['values'])  # error in active returns
    mg = pd.DataFrame(0, index=bus_days, columns=groups)  # managed group return contributions
    bg = pd.DataFrame(0, index=bus_days, columns=groups)  # benchmark group return contributions
    ag = pd.DataFrame(0, index=bus_days, columns=groups)  # active group return contributions
    # ss = pd.DataFrame(0, index=bus_days, columns=groups)  # stock selections
    # aa = pd.DataFrame(0, index=bus_days, columns=groups)  # asset allocation
    # ie = pd.DataFrame(0, index=bus_days, columns=groups)  # interactive effects
    wm = pd.DataFrame(0, index=pos_days, columns=groups)  # managed weights in each group
    wb = pd.DataFrame(0, index=pos_days, columns=groups)  # benchmark weights in each group
    wa = pd.DataFrame(0, index=pos_days, columns=groups)  # active weights in each group

    for idx, ed in enumerate(pos_days):
        rd = bus_days[idx]
        exposures = group_obj.load_exposures(ed, sec_ids)
        ids = np.intersect1d(sec_ids, exposures.index)
        factors = np.intersect1d(groups, exposures.columns)
        factors = np.intersect1d(factors, fr['values'].columns)
        vm = np.atleast_2d(pp.loc[ed, ids])
        vb = np.atleast_2d(bb.loc[ed, ids])
        va = np.atleast_2d(aa.loc[ed, ids])
        wm.loc[ed, factors] = np.matmul(vm, exposures.loc[ids, factors].to_numpy())[0, :]
        wb.loc[ed, factors] = np.matmul(vb, exposures.loc[ids, factors].to_numpy())[0, :]
        wa.loc[ed, factors] = np.matmul(va, exposures.loc[ids, factors].to_numpy())[0, :]

        mg.loc[rd, factors] = wm.loc[ed, factors] * fr['values'].loc[rd, factors]
        bg.loc[rd, factors] = wb.loc[ed, factors] * fr['values'].loc[rd, factors]
        ag.loc[rd, factors] = wa.loc[ed, factors] * fr['values'].loc[rd, factors]

        me.loc[rd, 'values'] = mr.loc[rd, 'values'] - mg.loc[rd].sum() - mrc.loc[rd].sum()
        be.loc[rd, 'values'] = br.loc[rd, 'values'] - bg.loc[rd].sum() - brc.loc[rd].sum()
        ae.loc[rd, 'values'] = ar.loc[rd, 'values'] - ag.loc[rd].sum() - arc.loc[rd].sum()
        display(f"{rd}: {group}: Total Active {ar.loc[rd,'values']:.2%}: Factor {ag.loc[rd].sum():.2%}, "
                f"Residual: {arc.loc[rd].sum():.2%} --- Error {ae.loc[rd, 'values']: .3%}")

    result = {'dates': bus_days,
              'position dates': pos_days,
              'managed weights': pp,
              'benchmark weights': bb,
              'active weights': aw,
              'managed returns': mr,
              'benchmark returns': br,
              'active returns': ar,
              'groups': groups,
              'types': types,
              'themes': themes,
              'managed group weights': wm,
              'benchmark group weights': wb,
              'active group weights': wa,
              'managed group contributions': mg,
              'benchmark group contributions': bg,
              'active group contributions': ag,
              'managed contributions': mc,
              'benchmark contributions': bc,
              'active contributions': ac,
              'managed factor contributions': mc - mrc,
              'benchmark factor contributions': bc - brc,
              'active factor contributions': ac - arc,
              'managed residual contributions': mrc,
              'benchmark residual contributions': brc,
              'active residual contributions': arc,
              'managed error': me,
              'benchmark error': be,
              'active error': ae
              }
    return result


def merge_contributions(data, result):
    if isinstance(data, np.ndarray):
        data = data[0]  # using first element only
    if data is None:
        return result
    d_dates = data['dates']
    r_dates = result['dates']
    dates = np.sort(np.union1d(d_dates, r_dates))
    q = {'dates': dates}
    pd_dates = data['position dates']
    pr_dates = result['position dates']
    pos_dates = np.sort(np.union1d(pd_dates, pr_dates))
    q['position dates'] = pos_dates

    fields = ['managed returns', 'benchmark returns', 'active returns', 'managed error', 'benchmark error',
              'active error']
    for f in fields:
        df = pd.DataFrame(dtype='float64', index=dates, columns=['values'])
        df.loc[d_dates, 'values'] = data[f].loc[d_dates, 'values']
        df.loc[r_dates, 'values'] = result[f].loc[r_dates, 'values']
        q[f] = df
        del df

    d_f = pd.DataFrame(zip(data['groups'], data['types'], data['themes']), columns=['groups', 'types', 'themes'])
    r_f = pd.DataFrame(zip(result['groups'], result['types'], result['themes']), columns=['groups', 'types', 'themes'])
    q_f = pd.concat((d_f, r_f), axis=0, ignore_index=True)
    q_f.drop_duplicates(keep='first', ignore_index=True, inplace=True)
    q['groups'] = q_f['groups'].to_numpy()
    q['types'] = q_f['types'].to_numpy()
    q['themes'] = q_f['themes'].to_numpy()
    groups = q['groups']
    d_groups = d_f['groups'].to_numpy()
    r_groups = r_f['groups'].to_numpy()
    del (d_f, r_f, q_f)
    # d_groups = data['groups']
    # r_groups = result['groups']
    # groups = np.sort(np.union1d(d_groups, r_groups))
    # q['groups'] = groups

    fields = ['managed weights', 'benchmark weights', 'active weights',
              'managed contributions', 'benchmark contributions', 'active contributions',
              'managed factor contributions', 'benchmark factor contributions',
              'active factor contributions', 'managed residual contributions',
              'benchmark residual contributions', 'active residual contributions']
    for f in fields:
        df = data[f].combine_first(result[f])
        df.fillna(0, inplace=True)
        q[f] = df
        del df
    fields = ['managed group contributions',
              'benchmark group contributions', 'active group contributions']
    for f in fields:
        df = pd.DataFrame(0.0, index=dates, columns=groups)
        df.loc[d_dates, d_groups] = data[f].loc[d_dates, d_groups]
        df.loc[r_dates, r_groups] = result[f].loc[r_dates, r_groups]
        q[f] = df
        del df

    fields = ['managed group weights', 'benchmark group weights', 'active group weights']
    for f in fields:
        df = pd.DataFrame(0.0, index=pos_dates, columns=groups)
        df.loc[pd_dates, d_groups] = data[f].loc[pd_dates, d_groups]
        df.loc[pr_dates, r_groups] = result[f].loc[pr_dates, r_groups]
        q[f] = df
        del df
    return q


def save_and_merge_contributions(result, file):
    if not os.path.exists(file):
        util.save_data(result, file)
        display(f"factor attribution successfully saved to \n{file}")
        return result
    data = util.load_data(file)
    q = merge_contributions(data, result)
    util.save_data(q, file)
    display(f"Merged {len(result['dates'])} days to prior {len(data['dates'])} days results")
    display(f"factor attribution successfully saved to \n{file}")
    return q


def filter_contributions_by_dates(data, bus_days=None):
    if bus_days is None:
        return data
    if isinstance(bus_days, str):
        bus_days = np.array([bus_days])
    if isinstance(data, np.ndarray):
        data = data[0]
    bus_days = np.unique(bus_days)
    # dates and position dates
    days, ia, ib = intersect(data['dates'], bus_days)
    data['dates'] = days
    data['position dates'] = data['position dates'][ia]
    # the rest
    for key in data.keys():
        if key in ['groups', 'dates', 'position dates']:
            continue
        df = data[key]
        if isinstance(df, pd.DataFrame):
            if 'weights' in key:
                dates = df.index[df.index.isin(data['position dates'])]
            else:
                dates = df.index[df.index.isin(bus_days)]
            df = df.loc[dates]
        else:
            display(f"Not supported type")
        data[key] = df
    return data


def aggregate_period_contributions(data):
    groups = data['factor_groups']
    data['period'] = np.full((len(groups), 1), None)
    data['risk'] = np.full((len(groups), 1), None)
    for idx, group in enumerate(groups):
        p = {}
        risk = {}
        val = data['values'][idx][0]
        p['sec_ids'] = val['active residual contributions'].columns.to_numpy()
        p['managed average weights'] = val['managed weights'].mean(axis=0).to_numpy()
        p['benchmark average weights'] = val['benchmark weights'].mean(axis=0).to_numpy()
        p['active average weights'] = p['managed average weights'] - p['benchmark average weights']
        # managed russell linking factor
        m_link = rt.russell_linking_factors(val['managed returns'].to_numpy())
        b_link = rt.russell_linking_factors(val['benchmark returns'].to_numpy())
        a_link = rt.russell_linking_factors(val['active returns'].to_numpy())
        d_link = m_link - b_link
        # num_of_groups = val['asset allocations'].shape[1]
        risk['managed volatility'] = np.nanstd(val['managed returns'].to_numpy().astype('float64'))*np.sqrt(252)
        risk['benchmark volatility'] = np.nanstd(val['benchmark returns'].to_numpy().astype('float64'))*np.sqrt(252)
        risk['tracking error'] = np.nanstd(val['active returns'].to_numpy().astype('float64')) * np.sqrt(252)
        p['groups'] = val['groups']
        p['types'] = val['types']
        p['themes'] = val['themes']
        risk['groups'] = val['groups']
        num = len(val['groups'])
        p['managed return'] = np.matmul(m_link.T, val['managed returns'].to_numpy())
        p['benchmark return'] = np.matmul(b_link.T, val['benchmark returns'].to_numpy())
        p['active return'] = p['managed return'] - p['benchmark return']
        p['managed group contribution'] = np.matmul(m_link.T, val['managed group contributions'].to_numpy())
        # val['managed residual contributions'].fillna(0, inplace=True)
        p['managed residual contribution'] = np.matmul(m_link.T, val['managed residual contributions'].to_numpy())
        p['managed contribution'] = np.matmul(m_link.T, val['managed contributions'].to_numpy())
        p['managed factor contribution'] = np.matmul(m_link.T, val['managed factor contributions'].to_numpy())
        # val['managed error'] = val['managed returns'].sum(axis=1).to_numpy() - \
        #     val['managed group contributions'].sum(axis=1).to_numpy() - \
        #     val['managed residual contributions'].sum(axis=1).to_numpy()
        # val['managed error'] = pd.DataFrame(np.atleast_2d(val['managed error']).T, index=val['managed returns'].index)
        p['managed error'] = p['managed return'] - p['managed group contribution'].sum(axis=1) - \
            p['managed residual contribution'].sum(axis=1)
        p['benchmark group contribution'] = np.matmul(b_link.T, val['benchmark group contributions'].to_numpy())
        val['benchmark residual contributions'].fillna(0, inplace=True)
        p['benchmark residual contribution'] = np.matmul(m_link.T, val['benchmark residual contributions'].to_numpy())
        p['benchmark contribution'] = np.matmul(m_link.T, val['benchmark contributions'].to_numpy())
        p['benchmark factor contribution'] = np.matmul(m_link.T, val['benchmark factor contributions'].to_numpy())
        # val['benchmark error'] = val['benchmark returns'].sum(axis=1).to_numpy() - \
        #     val['benchmark group contributions'].sum(axis=1).to_numpy() - \
        #     val['benchmark residual contributions'].sum(axis=1).to_numpy()
        # val['benchmark error'] = pd.DataFrame(np.atleast_2d(val['benchmark error']).T,
        #                                       index=val['managed returns'].index)
        p['benchmark error'] = p['benchmark return'] - \
            p['benchmark group contribution'].sum(axis=1) - \
            p['benchmark residual contribution'].sum(axis=1)
        p['active group contribution'] = p['managed group contribution'] - p['benchmark group contribution']
        p['active residual contribution'] = p['managed residual contribution'] - p['benchmark residual contribution']
        p['active contribution'] = p['managed contribution'] - p['benchmark contribution']
        p['active factor contribution'] = p['managed factor contribution'] - p['benchmark factor contribution']
        val['active error'] = val['managed error'] - val['benchmark error']
        p['active error'] = p['managed error'] - p['benchmark error']
        vf = val['managed residual contributions'].sum(axis=1).to_frame('residual')
        contrib = pd.concat((val['managed group contributions'], vf), axis=1)
        contrib = pd.concat((contrib, val['managed error']), axis=1)
        contrib = contrib.to_numpy()
        c2 = np.nansum(np.matmul(contrib.T, contrib), axis=1).T
        c2 = c2.reshape((1, len(c2)))
        r2 = np.nansum(val['managed returns'].to_numpy() * val['managed returns'].to_numpy())
        risk['managed group rcte'] = c2[0, :num] / r2
        risk['managed residual rcte'] = c2[0, num] / r2
        risk['managed error rcte'] = c2[0, -1] / r2
        vf = val['benchmark residual contributions'].sum(axis=1).to_frame('residual')
        contrib = pd.concat((val['benchmark group contributions'], vf), axis=1)
        contrib = pd.concat((contrib, val['benchmark error']), axis=1)
        contrib = contrib.to_numpy()
        c2 = np.nansum(np.matmul(contrib.T, contrib), axis=1).T
        c2 = c2.reshape((1, len(c2)))
        r2 = np.nansum(val['benchmark returns'].to_numpy() * val['benchmark returns'].to_numpy())
        risk['benchmark group rcte'] = c2[0, :num] / r2
        risk['benchmark residual rcte'] = c2[0, num] / r2
        risk['benchmark error rcte'] = c2[0, -1] / r2
        vf = val['active residual contributions'].sum(axis=1).to_frame('residual')
        contrib = pd.concat((val['active group contributions'], vf), axis=1)
        contrib = pd.concat((contrib, val['active error']), axis=1)
        contrib = contrib.to_numpy()
        c2 = np.nansum(np.matmul(contrib.T, contrib), axis=1).T
        c2 = c2.reshape((1, len(c2)))
        r2 = np.nansum(val['active returns'].to_numpy() * val['active returns'].to_numpy())
        risk['active group rcte'] = c2[0, :num] / r2
        risk['active residual rcte'] = c2[0, num] / r2
        risk['active error rcte'] = c2[0, -1] / r2

        p['managed group weight average'] = np.nanmean(val['managed group weights'], axis=0)
        p['managed group weight std'] = np.nanstd(val['managed group weights'], axis=0)
        p['benchmark group weight average'] = np.nanmean(val['benchmark group weights'], axis=0)
        p['benchmark group weight std'] = np.nanstd(val['benchmark group weights'], axis=0)
        p['active group weight average'] = np.nanmean(val['active group weights'], axis=0)
        p['active group weight std'] = np.nanstd(val['active group weights'], axis=0)
        data['period'][idx] = p
        data['risk'][idx] = risk
    return data


def export_fa_report(data, output_location):
    managed = data['managed portfolio']
    if isinstance(managed, int):
        m_ref = md.get_portfolio_references(managed)
        managed = m_ref['ReferenceCode'].iloc[0]
    benchmark = data['benchmark portfolio']
    if isinstance(benchmark, int):
        b_ref = md.get_portfolio_references(benchmark)
        benchmark = b_ref['ReferenceCode'].iloc[0]
    man_wt_flag = data['managed weighting method']
    ben_wt_flag = data['benchmark weighting method']
    v = data['values'][0][0]
    dates = v['dates']
    groups = data['factor_groups']
    sorted_groups = np.array(groups)
    sorted_groups.sort()
    file_name = f"{data['managed portfolio']}_vs_{data['benchmark portfolio']}_{man_wt_flag}_{ben_wt_flag}"
    for sg in sorted_groups:
        file_name = file_name + f"_{sg}"
    file_name = file_name + f"_{dates[0].strftime(util.yyyymmdd_format)}" \
                            f"_{dates[-1].strftime(util.yyyymmdd_format)}.xlsx"
    file = os.path.join(output_location, file_name)
    vp = data['period']
    risk = data['risk']
    if os.path.exists(file):
        util.preserve_file(file)

    # printing

    wb = pxl.Workbook()
    for idx, group in enumerate(groups):
        g_obj = root.load_object(group)
        name = g_obj.name
        p = vp[idx][0]
        r = risk[idx][0]
        if idx == 0:
            sheet = wb.active
            sheet.title = name
        else:
            sheet = wb.create_sheet(name, 2 * idx)
        detail = wb.create_sheet(f"Detail - {name}", 2 * idx + 1)
        display(f"exporting factor attributions on: '{name}'")
        row = 1
        col = 1
        sheet.cell(row, col).value = 'Factor Attributions - Realized Risk Analysis'
        sheet.cell(row, col).font = styles.Font(bold=True, color='0000FF')
        sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 4)
        row = row + 1
        sheet.cell(row, col).value = 'Start Date'
        sheet.cell(row, col+1).value = dates[0].strftime(util.YY_MM_DD_format)
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = 'End Date'
        sheet.cell(row, col+1).value = dates[-1].strftime(util.YY_MM_DD_format)
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = 'Factor Model'
        sheet.cell(row, col+1).value = name
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = 'Managed'
        sheet.cell(row, col+1).value = managed
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+1).font = styles.Font(bold=True)
        row = row + 1
        sheet.cell(row, col).value = 'Benchmark'
        sheet.cell(row, col+1).value = benchmark
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+1).font = styles.Font(bold=True)
        row = row + 3
        sheet.cell(row, col).value = 'Summary'
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col).fill = fill_pale_green
        sheet.cell(row, col).border = border_bottom
        sheet.cell(row, col+1).value = 'Return'
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+1).fill = fill_pale_green
        sheet.cell(row, col+1).border = border_bottom
        sheet.cell(row, col+2).value = 'Risk'
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+2).fill = fill_pale_green
        sheet.cell(row, col+2).border = border_bottom
        row = row + 1
        sheet.cell(row, col).value = managed
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+1).value = p['managed return'][0][0]
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+2).value = r['managed volatility']
        sheet.cell(row, col+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = benchmark
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+1).value = p['benchmark return'][0][0]
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+2).value = r['benchmark volatility']
        sheet.cell(row, col+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = 'Active'
        sheet.cell(row, col+1).value = p['active return'][0][0]
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+2).value = r['tracking error']
        sheet.cell(row, col+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='center')
        if abs(sheet.cell(row, col+2).value) > 0.05:
            sheet.cell(row, col+2).font = styles.Font(bold=True, color='FF0000')
            sheet.cell(row, col+2).fill = fill_yellow

        groups = p['groups']
        types = p['types']
        themes = p['themes']

        row = row + 2
        sheet.cell(row, col).value = 'Exposures'
        sheet.cell(row, col).fill = fill_pale_green
        sheet.cell(row, col).border = border_bottom
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+1).value = 'Total'
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+1).fill = fill_pale_green
        sheet.cell(row, col+1).border = border_bottom
        sheet.cell(row, col+2).value = 'Residual'
        sheet.cell(row, col+2).fill = fill_pale_green
        sheet.cell(row, col+2).border = border_bottom
        sheet.cell(row, col+3).value = 'Factor'
        sheet.cell(row, col+3).fill = fill_pale_green
        sheet.cell(row, col+3).border = border_bottom
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+4).value = g
            sheet.cell(row, col+gdx+4).alignment = styles.Alignment(horizontal='center')
            sheet.cell(row, col+gdx+4).fill = fill_dark_green
            sheet.cell(row, col+gdx+4).border = border_bottom
        row = row + 1
        sheet.cell(row, col).value = 'Type'
        sheet.cell(row, col).fill = fill_pale_green
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+1).value = ''
        sheet.cell(row, col+1).fill = fill_pale_green
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+2).value = ''
        sheet.cell(row, col+2).fill = fill_pale_green
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+3).value = ''
        sheet.cell(row, col+3).fill = fill_pale_green
        sheet.cell(row, col+3).alignment = styles.Alignment(horizontal='left')
        for gdx, g in enumerate(types):
            sheet.cell(row, col+gdx+4).value = g
            sheet.cell(row, col+gdx+4).alignment = styles.Alignment(horizontal='center')
            sheet.cell(row, col+gdx+4).fill = fill_pale_green
            sheet.cell(row, col+gdx+4).border = thin_bottom
        row = row + 1
        sheet.cell(row, col).value = 'Themes'
        sheet.cell(row, col).fill = fill_pale_green
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+1).value = ''
        sheet.cell(row, col+1).fill = fill_pale_green
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+2).value = ''
        sheet.cell(row, col+2).fill = fill_pale_green
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+3).value = ''
        sheet.cell(row, col+3).fill = fill_pale_green
        sheet.cell(row, col+3).alignment = styles.Alignment(horizontal='left')
        for gdx, g in enumerate(themes):
            sheet.cell(row, col+gdx+4).value = g
            sheet.cell(row, col+gdx+4).alignment = styles.Alignment(horizontal='center')
            sheet.cell(row, col+gdx+4).fill = fill_pale_green
            sheet.cell(row, col+gdx+4).border = thin_bottom
        row = row + 1
        sheet.cell(row, col).value = managed
        # sheet.cell(row, col+1).value = np.nansum(p['managed group weight average'])
        # sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).value = ''
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+4).value = p['managed group weight average'][gdx]
            sheet.cell(row, col+gdx+4).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col+gdx+4).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = benchmark
        # sheet.cell(row, col+1).value = np.nansum(p['benchmark group weight average'])
        # sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).value = ''
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+4).value = p['benchmark group weight average'][gdx]
            sheet.cell(row, col+gdx+4).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col+gdx+4).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = 'Active'
        # sheet.cell(row, col+1).value = np.nansum(p['active group weight average'])
        # sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).value = ''
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+4).value = p['active group weight average'][gdx]
            sheet.cell(row, col+gdx+4).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col+gdx+4).alignment = styles.Alignment(horizontal='center')
            if abs(sheet.cell(row, col+gdx+4).value) > 0.05:
                sheet.cell(row, col+gdx+4).font = styles.Font(bold=True, color='FF0000')
                sheet.cell(row, col+gdx+4).fill = fill_yellow
        # performance
        row = row + 2
        sheet.cell(row, col).value = 'Performance'
        sheet.cell(row, col).fill = fill_pale_green
        sheet.cell(row, col).border = border_bottom
        sheet.cell(row, col+1).value = 'Total'
        sheet.cell(row, col+1).fill = fill_pale_green
        sheet.cell(row, col+1).border = border_bottom
        sheet.cell(row, col+2).value = 'Residual'
        sheet.cell(row, col+2).fill = fill_pale_green
        sheet.cell(row, col+2).border = border_bottom
        sheet.cell(row, col+3).value = 'Factor'
        sheet.cell(row, col+3).fill = fill_pale_green
        sheet.cell(row, col+3).border = border_bottom
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+4).value = g
            sheet.cell(row, col+gdx+4).alignment = styles.Alignment(horizontal='center')
            sheet.cell(row, col+gdx+4).fill = fill_dark_green
            sheet.cell(row, col+gdx+4).border = border_bottom
        row = row + 1
        sheet.cell(row, col).value = 'Type'
        sheet.cell(row, col).fill = fill_pale_green
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+1).value = ''
        sheet.cell(row, col+1).fill = fill_pale_green
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+2).value = ''
        sheet.cell(row, col+2).fill = fill_pale_green
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+3).value = ''
        sheet.cell(row, col+3).fill = fill_pale_green
        sheet.cell(row, col+3).alignment = styles.Alignment(horizontal='left')
        for gdx, g in enumerate(types):
            sheet.cell(row, col+gdx+4).value = g
            sheet.cell(row, col+gdx+4).alignment = styles.Alignment(horizontal='center')
            sheet.cell(row, col+gdx+4).fill = fill_pale_green
            sheet.cell(row, col+gdx+4).border = thin_bottom
        row = row + 1
        sheet.cell(row, col).value = 'Themes'
        sheet.cell(row, col).fill = fill_pale_green
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+1).value = ''
        sheet.cell(row, col+1).fill = fill_pale_green
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+2).value = ''
        sheet.cell(row, col+2).fill = fill_pale_green
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+3).value = ''
        sheet.cell(row, col+3).fill = fill_pale_green
        for gdx, g in enumerate(themes):
            sheet.cell(row, col+gdx+4).value = g
            sheet.cell(row, col+gdx+4).alignment = styles.Alignment(horizontal='center')
            sheet.cell(row, col+gdx+4).fill = fill_pale_green
            sheet.cell(row, col+gdx+4).border = thin_bottom
        row = row + 1
        sheet.cell(row, col).value = 'Managed'
        sheet.cell(row, col+1).value = np.nansum(p['managed group contribution']) + \
            np.nansum(p['managed residual contribution'][0]) + \
            np.nansum(p['managed error'])
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+2).value = np.nansum(p['managed residual contribution'][0]) + np.nansum(p['managed error'])
        sheet.cell(row, col+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='center')
        if sheet.cell(row, col+2).value > 0.02:
            sheet.cell(row, col+2).font = styles.Font(bold=True, color='0000FF')
            sheet.cell(row, col+2).fill = fill_green
        if sheet.cell(row, col+2).value < -0.02:
            sheet.cell(row, col+2).font = styles.Font(bold=True, color='FF0000')
            sheet.cell(row, col+2).fill = fill_yellow
        sheet.cell(row, col+3).value = np.nansum(p['managed group contribution'][0])
        sheet.cell(row, col+3).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+3).alignment = styles.Alignment(horizontal='center')
        if sheet.cell(row, col+3).value > 0.02:
            sheet.cell(row, col+3).font = styles.Font(bold=True, color='0000FF')
            sheet.cell(row, col+3).fill = fill_green
        if sheet.cell(row, col+3).value < -0.02:
            sheet.cell(row, col+3).font = styles.Font(bold=True, color='FF0000')
            sheet.cell(row, col+3).fill = fill_yellow
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+4).value = p['managed group contribution'][0][gdx]
            sheet.cell(row, col+gdx+4).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col+gdx+4).alignment = styles.Alignment(horizontal='center')
            if sheet.cell(row, col + gdx + 4).value > 0.02:
                sheet.cell(row, col + gdx + 4).font = styles.Font(bold=True, color='0000FF')
                sheet.cell(row, col + gdx + 4).fill = fill_green
            if sheet.cell(row, col + gdx + 4).value < -0.02:
                sheet.cell(row, col + gdx + 4).font = styles.Font(bold=True, color='FF0000')
                sheet.cell(row, col + gdx + 4).fill = fill_yellow
        row = row + 1
        sheet.cell(row, col).value = 'Benchmark'
        sheet.cell(row, col+1).value = np.nansum(p['benchmark group contribution'])+ \
            np.nansum(p['benchmark residual contribution'][0]) + \
            np.nansum(p['benchmark error'])
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+2).value = np.nansum(p['benchmark residual contribution'][0]) + \
            np.nansum(p['benchmark error'])
        sheet.cell(row, col+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='center')
        if sheet.cell(row, col+2).value > 0.02:
            sheet.cell(row, col+2).font = styles.Font(bold=True, color='0000FF')
            sheet.cell(row, col+2).fill = fill_green
        if sheet.cell(row, col+2).value < -0.02:
            sheet.cell(row, col+2).font = styles.Font(bold=True, color='FF0000')
            sheet.cell(row, col+2).fill = fill_yellow
        sheet.cell(row, col+3).value = np.nansum(p['benchmark group contribution'][0])
        sheet.cell(row, col+3).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+3).alignment = styles.Alignment(horizontal='center')
        if sheet.cell(row, col+3).value > 0.02:
            sheet.cell(row, col+3).font = styles.Font(bold=True, color='0000FF')
            sheet.cell(row, col+3).fill = fill_green
        if sheet.cell(row, col+3).value < -0.02:
            sheet.cell(row, col+3).font = styles.Font(bold=True, color='FF0000')
            sheet.cell(row, col+3).fill = fill_yellow
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+4).value = p['benchmark group contribution'][0][gdx]
            sheet.cell(row, col+gdx+4).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col+gdx+4).alignment = styles.Alignment(horizontal='center')
            if sheet.cell(row, col+gdx+4).value > 0.02:
                sheet.cell(row, col + gdx + 4).font = styles.Font(bold=True, color='0000FF')
                sheet.cell(row, col + gdx + 4).fill = fill_green
            if sheet.cell(row, col + gdx + 4).value < -0.02:
                sheet.cell(row, col + gdx + 4).font = styles.Font(bold=True, color='FF0000')
                sheet.cell(row, col + gdx + 4).fill = fill_yellow

        row = row + 1
        sheet.cell(row, col).value = 'Total'
        sheet.cell(row, col+1).value = np.nansum(p['active group contribution']) + \
            np.nansum(p['active residual contribution'][0]) + \
            np.nansum(p['active error'])
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+2).value = np.nansum(p['active residual contribution'][0]) + np.nansum(p['active error'])
        sheet.cell(row, col+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='center')
        if sheet.cell(row, col+2).value > 0.02:
            sheet.cell(row, col+2).font = styles.Font(bold=True, color='0000FF')
            sheet.cell(row, col+2).fill = fill_green
        if sheet.cell(row, col+2).value < -0.02:
            sheet.cell(row, col+2).font = styles.Font(bold=True, color='FF0000')
            sheet.cell(row, col+2).fill = fill_yellow
        sheet.cell(row, col+3).value = np.nansum(p['active group contribution'][0])
        sheet.cell(row, col+3).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+3).alignment = styles.Alignment(horizontal='center')
        if sheet.cell(row, col+3).value > 0.02:
            sheet.cell(row, col+3).font = styles.Font(bold=True, color='0000FF')
            sheet.cell(row, col+3).fill = fill_green
        if sheet.cell(row, col+3).value < -0.02:
            sheet.cell(row, col+3).font = styles.Font(bold=True, color='FF0000')
            sheet.cell(row, col+3).fill = fill_yellow
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+4).value = p['active group contribution'][0][gdx]
            sheet.cell(row, col+gdx+4).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col+gdx+4).alignment = styles.Alignment(horizontal='center')
            if sheet.cell(row, col+gdx+4).value > 0.02:
                sheet.cell(row, col+gdx+4).font = styles.Font(bold=True, color='0000FF')
                sheet.cell(row, col+gdx+4).fill = fill_green
            if sheet.cell(row, col+gdx+4).value < -0.02:
                sheet.cell(row, col+gdx+4).font = styles.Font(bold=True, color='FF0000')
                sheet.cell(row, col+gdx+4).fill = fill_yellow

        # risk decompositions
        row = row + 2
        sheet.cell(row, col).value = 'RCTE'
        sheet.cell(row, col).fill = fill_pale_green
        sheet.cell(row, col).border = border_bottom
        sheet.cell(row, col+1).value = 'Total'
        sheet.cell(row, col+1).fill = fill_pale_green
        sheet.cell(row, col+1).border = border_bottom
        sheet.cell(row, col+2).value = 'Residual'
        sheet.cell(row, col+2).fill = fill_pale_green
        sheet.cell(row, col+2).border = border_bottom
        sheet.cell(row, col+3).value = 'Factor'
        sheet.cell(row, col+3).fill = fill_pale_green
        sheet.cell(row, col+3).border = border_bottom
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+4).value = g
            sheet.cell(row, col+gdx+4).alignment = styles.Alignment(horizontal='center')
            sheet.cell(row, col+gdx+4).fill = fill_dark_green
            sheet.cell(row, col+gdx+4).border = border_bottom
        row = row + 1
        sheet.cell(row, col).value = 'Type'
        sheet.cell(row, col).fill = fill_pale_green
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+1).value = ''
        sheet.cell(row, col+1).fill = fill_pale_green
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+2).value = ''
        sheet.cell(row, col+2).fill = fill_pale_green
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+3).value = ''
        sheet.cell(row, col+3).fill = fill_pale_green
        for gdx, g in enumerate(types):
            sheet.cell(row, col+gdx+4).value = g
            sheet.cell(row, col+gdx+4).alignment = styles.Alignment(horizontal='center')
            sheet.cell(row, col+gdx+4).fill = fill_pale_green
            sheet.cell(row, col+gdx+4).border = thin_bottom
        row = row + 1
        sheet.cell(row, col).value = 'Themes'
        sheet.cell(row, col).fill = fill_pale_green
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+1).value = ''
        sheet.cell(row, col+1).fill = fill_pale_green
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+2).value = ''
        sheet.cell(row, col+2).fill = fill_pale_green
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='left')
        sheet.cell(row, col+3).value = ''
        sheet.cell(row, col+3).fill = fill_pale_green
        for gdx, g in enumerate(themes):
            sheet.cell(row, col+gdx+4).value = g
            sheet.cell(row, col+gdx+4).alignment = styles.Alignment(horizontal='center')
            sheet.cell(row, col+gdx+4).fill = fill_pale_green
            sheet.cell(row, col+gdx+4).border = thin_bottom
        row = row + 1
        sheet.cell(row, col).value = managed
        sheet.cell(row, col+1).value = np.nansum(r['managed group rcte']) + np.nansum(r['managed residual rcte']) + \
            np.nansum(r['managed error rcte'])
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+2).value = np.nansum(r['managed residual rcte']) + np.nansum(r['managed error rcte'])
        sheet.cell(row, col+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+3).value = np.nansum(r['managed group rcte'])
        sheet.cell(row, col+3).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+3).alignment = styles.Alignment(horizontal='center')
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+4).value = np.nansum(r['managed group rcte'][gdx])
            sheet.cell(row, col+gdx+4).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col+gdx+4).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = benchmark
        sheet.cell(row, col+1).value = np.nansum(r['benchmark group rcte']) + \
            np.nansum(r['benchmark residual rcte']) + \
            np.nansum(r['benchmark error rcte'])
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+2).value = np.nansum(r['benchmark residual rcte']) + np.nansum(r['benchmark error rcte'])
        sheet.cell(row, col+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+3).value = np.nansum(r['benchmark group rcte'])
        sheet.cell(row, col+3).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+3).alignment = styles.Alignment(horizontal='center')
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+4).value = np.nansum(r['benchmark group rcte'][gdx])
            sheet.cell(row, col+gdx+4).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col+gdx+4).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = 'Active'
        sheet.cell(row, col+1).value = np.nansum(r['active group rcte']) + np.nansum(r['active residual rcte']) + \
            np.nansum(r['active error rcte'])
        sheet.cell(row, col+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+2).value = np.nansum(r['active residual rcte']) + np.nansum(r['active error rcte'])
        sheet.cell(row, col+2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+2).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col+3).value = np.nansum(r['active group rcte'])
        sheet.cell(row, col+3).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col+3).alignment = styles.Alignment(horizontal='center')
        for gdx, g in enumerate(groups):
            sheet.cell(row, col+gdx+4).value = np.nansum(r['active group rcte'][gdx])
            sheet.cell(row, col+gdx+4).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col+gdx+4).alignment = styles.Alignment(horizontal='center')
            if sheet.cell(row, col+gdx+4).value > 0.15:
                sheet.cell(row, col+gdx+4).font = styles.Font(bold=True, color='FF0000')
                sheet.cell(row, col+gdx+4).fill = fill_yellow
            if sheet.cell(row, col+gdx+4).value < -0.15:
                sheet.cell(row, col+gdx+4).font = styles.Font(bold=True, color='0000FF')
                sheet.cell(row, col+gdx+4).fill = fill_green
        # ---------------------------------------------------------------
        #
        #           Detail
        #
        # ---------------------------------------------------------------
        display('*'*100)
        display(f"exporting factor attributions detail on: '{name}'")
        display('*'*100)
        por_types = ['Portfolio', 'Benchmark', 'Active']
        portfolios = [data['managed portfolio'], data['benchmark portfolio'], '-']
        row = 1
        col = 1
        columns = ['ID', 'Ticker', 'Cusip', 'Sedol', 'Issuer', 'Exchange', 'Currency', 'Sector', 'Industry',
                   'Weight', 'Total Contribution', 'Factor', 'Residual',
                   'Weight', 'Total Contribution', 'Factor', 'Residual',
                   'Weight', 'Total Contribution', 'Factor', 'Residual']
        col = col + len(columns) - 12
        detail.cell(row, col).value = portfolios[0]   # result['managed_portfolio']
        detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
        detail.cell(row, col).font = styles.Font(bold=True, color='0000FF')
        detail.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 3)
        col = col + 4
        detail.cell(row, col).value = portfolios[1]   # result['benchmark_portfolio']
        detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
        detail.cell(row, col).font = styles.Font(bold=True, color='0000FF')
        detail.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 3)
        col = col + 4
        detail.cell(row, col).value = 'Active'
        detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
        detail.cell(row, col).font = styles.Font(bold=True, color='0000FF')
        detail.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 3)
        col = col + 4
        detail.cell(row, col).value = 'Exposures'
        detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        col = 1
        for cc in columns:
            detail.cell(row, col).value = cc
            detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
            col = col + 1
        sec_ids = p['sec_ids']
        ref = md.get_references(sec_ids)
        cusips = md.get_cusips(sec_ids)
        sedols = md.get_sedols(sec_ids)
        ind = md.get_rbics_classification(sec_ids)
        tickers = md.get_tickers(sec_ids)
        for jx, s in enumerate(sec_ids):
            row = row + 1
            col = 1
            detail.cell(row, col).value = s
            zx = np.where(tickers['sec_id'] == s)[0]
            if len(zx) > 0:
                detail.cell(row, col + 1).value = tickers.loc[tickers.index[zx[-1]], 'ticker_region']
            zx = np.where(cusips['sec_id'] == s)[0]
            if len(zx) > 0:
                detail.cell(row, col + 2).value = cusips.loc[cusips.index[zx[-1]], 'cusip']
            zx = np.where(sedols['sec_id'] == s)[0]
            if len(zx) > 0:
                detail.cell(row, col + 3).value = sedols.loc[sedols.index[zx[-1]], 'sedol']
            zx = np.where(ref['sec_id'] == s)[0]
            if len(zx) > 0:
                detail.cell(row, col + 4).value = ref.loc[ref.index[zx[-1]], 'name']
                detail.cell(row, col + 5).value = ref.loc[ref.index[zx[-1]], 'exchange']
                detail.cell(row, col + 6).value = ref.loc[ref.index[zx[-1]], 'currency']
            zx = np.where(ind['sec_id'] == s)[0]
            if len(zx) > 0:
                detail.cell(row, col + 7).value = ind.loc[ind.index[zx[-1]], 'l1_name']
                detail.cell(row, col + 8).value = ind.loc[ind.index[zx[-1]], 'l3_name']
            detail.cell(row, col + 9).value = p['managed average weights'][jx]
            detail.cell(row, col + 9).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            detail.cell(row, col + 9).alignment = styles.Alignment(horizontal='center')
            detail.cell(row, col + 10).value = p['managed contribution'][0, jx]
            detail.cell(row, col + 10).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            detail.cell(row, col + 10).alignment = styles.Alignment(horizontal='center')
            detail.cell(row, col + 11).value = p['managed factor contribution'][0, jx]
            detail.cell(row, col + 11).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            detail.cell(row, col + 11).alignment = styles.Alignment(horizontal='center')
            detail.cell(row, col + 12).value = p['managed residual contribution'][0, jx]
            detail.cell(row, col + 12).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            detail.cell(row, col + 12).alignment = styles.Alignment(horizontal='center')
            detail.cell(row, col + 13).value = p['benchmark average weights'][jx]
            detail.cell(row, col + 13).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            detail.cell(row, col + 13).alignment = styles.Alignment(horizontal='center')
            detail.cell(row, col + 14).value = p['benchmark contribution'][0, jx]
            detail.cell(row, col + 14).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            detail.cell(row, col + 14).alignment = styles.Alignment(horizontal='center')
            detail.cell(row, col + 15).value = p['benchmark factor contribution'][0, jx]
            detail.cell(row, col + 15).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            detail.cell(row, col + 15).alignment = styles.Alignment(horizontal='center')
            detail.cell(row, col + 16).value = p['benchmark residual contribution'][0, jx]
            detail.cell(row, col + 16).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            detail.cell(row, col + 16).alignment = styles.Alignment(horizontal='center')
            detail.cell(row, col + 17).value = p['active average weights'][jx]
            detail.cell(row, col + 17).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            detail.cell(row, col + 17).alignment = styles.Alignment(horizontal='center')
            detail.cell(row, col + 18).value = p['active contribution'][0, jx]
            detail.cell(row, col + 18).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            detail.cell(row, col + 18).alignment = styles.Alignment(horizontal='center')
            detail.cell(row, col + 19).value = p['active factor contribution'][0, jx]
            detail.cell(row, col + 19).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            detail.cell(row, col + 19).alignment = styles.Alignment(horizontal='center')
            detail.cell(row, col + 20).value = p['active residual contribution'][0, jx]
            detail.cell(row, col + 20).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            detail.cell(row, col + 20).alignment = styles.Alignment(horizontal='center')
    wb.save(file)
    display(f"Successfully factor attribution report to {file}")


fill_dark_green = styles.PatternFill("solid", start_color='5DBB63')
fill_pale_green = styles.PatternFill("solid", start_color='EEFFEE')
fill_yellow = styles.PatternFill("solid", start_color='FFFF66')
fill_green = styles.PatternFill("solid", start_color='99FF99')
font_blue = styles.Font(color=styles.colors.BLUE)
font_blue_bold = styles.Font(bold=True, color=styles.colors.BLUE)
font_red_bold = styles.Font(bold=True, color='CC0000')
border_bottom = styles.Border(bottom=styles.Side(color='000000', border_style='double'))
thin_bottom = styles.Border(bottom=styles.Side(color='000000', border_style='hair'))

