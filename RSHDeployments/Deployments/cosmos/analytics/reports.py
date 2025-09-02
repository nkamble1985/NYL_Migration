#
# reporting
#
# Author: Yun Chen
# Copyright: Indigo Dao, LLC
# Date: 2022
#
import calendar
import openpyxl as pxl
import openpyxl.styles as styles
import classes.root as root
import dataloader.portfolio as port
import dataloader.market_data as md
import time
import numpy as np
import pandas as pd
import util.routines as rt
from util.utilities import display
import factors.portfolio
import util.utilities as util
from util.intersect import *
import os
import analytics.ea.factor_performance as fp


def run_composite_report(bus_day, univ=None, print_report=True, file=None):

    if univ is None:
        univ = 'US1000_INDGRP_UNIV'
    calendar_str = 'US'
    bus_days = util.load_business_days(calendar_str, None, bus_day)
    bus_day = bus_days[-1]
    p = port.get_positions(bus_day, bus_day, univ, calendar_str)
    sec_ids = p.columns.to_numpy()
    r = md.get_returns(bus_days[-252], bus_day, sec_ids, calendar_str=calendar_str, security_type='QSR')
    df = pd.DataFrame(index=sec_ids, columns=['Return 1D', 'Return 5D', 'Return 21D', 'Return 63D',
                                              'Return 126D', 'Return 252D', 'Realized Volatility',
                                              'Realized Correlation', 'Predicted Beta',
                                              'Predicted Beta Lag 1M', 'Predicted Vol',
                                              'Predicted Vol Lag 1M', 'Correlation', 'Correlation Lag 1M',
                                              'P/B', 'P/B Spread', 'P/E', 'P/E Spread', 'ROE', 'ROE Spread'
                                              ])
    c, i1, i2 = intersect(sec_ids, r.columns.to_numpy())
    df.loc[c, 'Return 1D'] = r.loc[r.index[-1], c].T
    df.loc[c, 'Return 5D'] = np.prod(1+r.loc[r.index[-5:], c]).T - 1
    df.loc[c, 'Return 21D'] = np.prod(1+r.loc[r.index[-21:], c]).T - 1
    df.loc[c, 'Return 63D'] = np.prod(1+r.loc[r.index[-63:], c]).T - 1
    df.loc[c, 'Return 126D'] = np.prod(1+r.loc[r.index[-126:], c]).T - 1
    df.loc[c, 'Return 252D'] = np.prod(1+r.loc[r.index[-252:], c]).T - 1
    # v_obj = root.load_object('PredictedVolatility')
    # # present day + t - 21 Vol
    # vol = v_obj.load_values('values', bus_day, bus_day, sec_ids, composite_flag=True)
    # c, i1, i2 = intersect(sec_ids, vol.columns.to_numpy())
    # df.loc[c, 'Predicted Vol'] = vol.loc[vol.index[-1], c].T
    # vol = v_obj.load_values('values', bus_days[-21], bus_days[-21], sec_ids, composite_flag=True)
    # c, i1, i2 = intersect(sec_ids, vol.columns.to_numpy())
    # df.loc[c, 'Predicted Vol Lag 1M'] = vol.loc[vol.index[-1], c].T
    v_obj = root.load_object('US_LC_RISK_MODEL')
    # present day + t - 21 Vol
    vol = v_obj.load_values('volatility', bus_days[-1], sec_ids, composite_flag=True)
    c, i1, i2 = intersect(sec_ids, vol.index.to_numpy())
    df.loc[c, 'Predicted Vol'] = vol.loc[c, 'values']
    vol = v_obj.load_values('volatility', bus_days[-21], sec_ids, composite_flag=True)
    c, i1, i2 = intersect(sec_ids, vol.index.to_numpy())
    df.loc[c, 'Predicted Vol Lag 1M'] = vol.loc[c, 'values']
    # present day dispersion
    dispersion = v_obj.load_values('correlations', bus_days[-1], sec_ids, composite_flag=True)
    c, i1, i2 = intersect(sec_ids, dispersion.index.to_numpy())
    df.loc[c, 'Correlation'] = dispersion.loc[c, 'values']
    dispersion = v_obj.load_values('correlations', bus_days[-21], sec_ids, composite_flag=True)
    c, i1, i2 = intersect(sec_ids, dispersion.index.to_numpy())
    df.loc[c, 'Correlation Lag 1M'] = dispersion.loc[c, 'values'].T
    # realized volatility
    r_obj = root.load_object('Realized_Risk_Model')
    realized = r_obj.load_values('volatility', bus_day, sec_ids, composite_flag=True)
    c, i1, i2 = intersect(sec_ids, realized.index)
    df.loc[c, 'Realized Volatility'] = realized.loc[c, 'values']
    realized = r_obj.load_values('correlations', bus_day, sec_ids, composite_flag=True)
    c, i1, i2 = intersect(sec_ids, realized.index)
    df.loc[c, 'Realized Correlation'] = realized.loc[c, 'values']

    # present day beta
    b_obj = root.load_object('PredictedBeta')
    betas = b_obj.load_values('values', bus_day, bus_day, sec_ids, composite_flag=True)
    c, i1, i2 = intersect(sec_ids, betas.columns.to_numpy())
    df.loc[c, 'Predicted Beta'] = betas.loc[betas.index[-1], c].T
    betas = b_obj.load_values('values', bus_days[-21], bus_days[-21], sec_ids, composite_flag=True)
    c, i1, i2 = intersect(sec_ids, betas.columns.to_numpy())
    df.loc[c, 'Predicted Beta Lag 1M'] = betas.loc[betas.index[-1], c].T

    fac = ['p2b', 'p2e', 'p2s', 'ROE']
    names = ['P/B', 'P/E', 'P/S', 'ROE']
    for i, f in enumerate(fac):
        pb_obj = root.load_object(f)
        pb = pb_obj.load_values('median', bus_day, bus_day, sec_ids, composite_flag=True)
        c, i1, i2 = intersect(sec_ids, pb.columns.to_numpy())
        df.loc[c, names[i]] = pb.loc[pb.index[-1], c].T
        pb = pb_obj.load_values('std', bus_day, bus_day, sec_ids, composite_flag=True)
        c, i1, i2 = intersect(sec_ids, pb.columns.to_numpy())
        df.loc[c, f"{names[i]} Spread"] = pb.loc[pb.index[-1], c].T
    if print_report or file is not None:
        if file is None:
            location = os.path.join(util.default_output_location('reports'), 'tmp',
                                    f"{bus_day.strftime(util.yyyymmdd_format)}", 'composites')
            if not os.path.exists(location):
                os.makedirs(location)
                display(f"Created: {location}")
            file = os.path.join(location,
                                f"composite_report.{univ}.{bus_day.strftime(util.yyyymmdd_format)}.xlsx")
        df.to_excel(file)
        display(f"Successfully output composite data to {file}")


def output_stock_details(data, file_name=None, bus_day=None, calendar_str='US', sheet='Detail', prod=False):
    """
    simple function to export stock level attributes to file, padded with stock identifiers, classifications etc

    :param data: dataframe, index being regional IDs
    :param file_name:
    :param bus_day:
    :param calendar_str:
    :param sheet:
    :param prod: default False, if True output production
    :return:

    Author : Yun Chen
    Copyright: Indigo Dao, LLC
    Date: August 31, 2022
    """
    if bus_day is None:
        today = util.today()
        bus_day = util.previous_business_days(today, calendar_str)
    if not isinstance(data, pd.DataFrame):
        display(f"wrong data type: dataframe expected")
        raise ValueError(f"Wrong data type")
    sec_ids = data.index.to_numpy()
    df = pd.DataFrame(index=sec_ids)
    ref = md.get_stock_references(sec_ids)
    df = df.join(ref)
    caps = md.get_market_cap(bus_day, bus_day, sec_ids, calendar_str, base_currency='USD')
    capl = caps[1]
    capl = capl.T
    capl = capl * 1e6
    capl.rename(columns={capl.columns[0]: 'MktCap Local'}, inplace=True)
    df = df.join(capl)
    capu = caps[0]
    capu = capu.T
    capu = capu * 1e6
    capu.rename(columns={capu.columns[0]: 'MktCap USD'}, inplace=True)
    df = df.join(capu)
    df = df.join(data)

    df.reset_index(inplace=True)
    df.rename(columns={df.columns[0]: 'sec_id'}, inplace=True)

    if file_name is None:
        if prod:
            env = 'PROD'
        else:
            env = 'DEV'
        location = os.path.join(util.default_output_location('reports', env), 'tmp')
        if not util.exists(location):
            util.makedirs(location)
            display(f"created: {location}")
        file_name = os.path.join(location, f'stock_detail{time.time()}.xlsx')

    exist = os.path.exists(file_name)
    if exist:
        with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=sheet)
    else:
        with pd.ExcelWriter(file_name, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name=sheet)
    display(f"Output data to {file_name}: sheet {sheet}")
    return df


def export_portfolio(bus_day, por, calendar_str=None, location=None, recurse=False, deep=False, sheet='Detail',
                     facts=None, prod=False):
    """
    export portfolio to spreadsheet for a given day
    :param bus_day:
    :param por:
    :param calendar_str:
    :param location:
    :param recurse:
    :param deep:
    :param sheet:
    :param facts:
    :param prod: default False, if true output to production location
    :return:

    Author: Yun Chen
    Copyright: Indigo Dao, LLC
    Date: August 31, 2022
    """
    obj = root.load_object(por)
    if calendar_str is None:
        calendar_str = obj.calendar
    p = port.get_portfolio_weights(bus_day, bus_day, por, calendar_str=calendar_str, recurse=recurse,
                                   deep=deep)
    day = p.index[-1]
    data = {'sec_ids': p.columns.to_numpy(), 'Weights': p.iloc[-1, :].to_numpy()}
    if facts is not None:
        if isinstance(facts, str):
            facts = np.array([facts])
        for f in facts:
            try:
                obj = root.load_object(f)
                b = obj.load_values('DESCRIPTOR', bus_day, bus_day, data['sec_ids'])
                values = np.ones((len(data['sec_ids']), )) * np.nan
                c, i1, i2 = intersect(data['sec_ids'], b.columns)
                values[i1] = b.iloc[0, i2].to_numpy()
                del (c, i1, i2)
                data[obj.name] = values
                del values
            except ValueError as ve:
                display(ve)
                display(f)
            except Exception as ep:
                display(ep)
                display(f)
    if location is None:
        if prod:
            env = 'PROD'
        else:
            env = 'DEV'
        location = os.path.join(util.default_output_location('reports', env), 'tmp',
                                f"{day.strftime(util.yyyymmdd_format)}")
    if not os.path.exists(location):
        os.makedirs(location)
    file = os.path.join(location,
                        f'{por}.{day.strftime(util.yyyymmdd_format)}.xlsx')
    output_stock_details(data, file, bus_day=day, calendar_str=calendar_str, sheet=sheet)
    return True


def export_return_analysis(result, file, include_ts=False, calendar_str='US'):
    if util.root_directory() not in file:
        file = os.path.join(util.root_directory(), file)
    active = result['active']
    managed = result['managed']
    benchmark = result['benchmark']
    market = result['market']
    indicator = result['indicator']
    bmk = result['benchmark']['names'][0]
    mkt = result['market']['names'][0]
    wb = pxl.Workbook()
    summary = wb.active
    summary.title = 'Summary'
    corr = wb.create_sheet('Correlations', 1)
    regime = wb.create_sheet('Business Cycle', 2)
    horizons = wb.create_sheet('Horizons', 3)
    year = wb.create_sheet('Annual', 4)
    semiannual = wb.create_sheet('Semi-Annual', 5)
    quarter = wb.create_sheet('Quarter', 6)
    month = wb.create_sheet('Month', 7)
    if include_ts:
        tsm = wb.create_sheet('Managed', 9)
        tsb = wb.create_sheet('Benchmark', 10)
        tsa = wb.create_sheet('Active', 11)
        tsy = wb.create_sheet('Monthly', 8)
    # summary
    display(f"exporting return analysis 'summary'")
    row = 1
    col = 1
    summary.cell(row, col).value = 'Return Summary'
    summary.cell(row, col).font = font_blue_bold
    row = row + 1
    summary.cell(row, col).value = 'Start Date'
    summary.cell(row, col+1).value = result['dates'][0].strftime(util.YY_MM_DD_format)
    summary.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    summary.cell(row, col).value = 'End Date'
    summary.cell(row, col+1).value = result['dates'][0-1].strftime(util.YY_MM_DD_format)
    summary.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    row = row + 1
    summary.cell(row, col).value = 'Benchmark'
    summary.cell(row, col+1).value = result['benchmark']['names'][0]
    summary.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    summary.cell(row, col).value = 'Market'
    summary.cell(row, col+1).value = result['market']['names'][0]
    summary.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 2
    summary.cell(row, col+1).value = 'Annual'
    summary.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col+1).fill = fill_pale_blue
    summary.cell(row, col+2).value = 'Benchmark'
    summary.cell(row, col+2).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col+2).fill = fill_pale_blue
    summary.cell(row, col+3).value = 'Alpha'
    summary.cell(row, col+3).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 3).fill = fill_pale_blue
    summary.cell(row, col+4).value = 'Vol'
    summary.cell(row, col+4).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 4).fill = fill_pale_blue
    summary.cell(row, col+5).value = 'Vol (bmk)'
    summary.cell(row, col+5).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 5).fill = fill_pale_blue
    summary.cell(row, col+6).value = 'TE'
    summary.cell(row, col+6).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 6).fill = fill_pale_blue
    summary.cell(row, col+7).value = 'Sharpe'
    summary.cell(row, col+7).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 7).fill = fill_pale_blue
    summary.cell(row, col+8).value = 'Bmk Sharpe'
    summary.cell(row, col+8).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 8).fill = fill_pale_blue
    summary.cell(row, col+9).value = 'IR'
    summary.cell(row, col+9).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 9).fill = fill_pale_blue
    summary.cell(row, col+10).value = 'Max DD'
    summary.cell(row, col+10).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 10).fill = fill_pale_blue
    summary.cell(row, col+11).value = 'Duration'
    summary.cell(row, col+11).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 11).fill = fill_pale_blue
    summary.cell(row, col + 12).value = 'Start'
    summary.cell(row, col + 12).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 12).fill = fill_pale_blue
    summary.cell(row, col + 13).value = 'End'
    summary.cell(row, col + 13).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 13).fill = fill_pale_blue
    summary.cell(row, col + 14).value = 'Bmk Max DD'
    summary.cell(row, col + 14).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 14).fill = fill_pale_blue
    summary.cell(row, col + 15).value = 'Duration'
    summary.cell(row, col + 15).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 15).fill = fill_pale_blue
    summary.cell(row, col + 16).value = 'Start'
    summary.cell(row, col + 16).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 16).fill = fill_pale_blue
    summary.cell(row, col + 17).value = 'End'
    summary.cell(row, col + 17).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 17).fill = fill_pale_blue
    summary.cell(row, col + 18).value = 'Rel Max DD'
    summary.cell(row, col + 18).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 18).fill = fill_pale_blue
    summary.cell(row, col + 19).value = 'Duration'
    summary.cell(row, col + 19).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 19).fill = fill_pale_blue
    summary.cell(row, col + 20).value = 'Start'
    summary.cell(row, col + 20).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 20).fill = fill_pale_blue
    summary.cell(row, col + 21).value = 'End'
    summary.cell(row, col + 21).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 21).fill = fill_pale_blue
    summary.cell(row, col + 22).value = 'Beta'
    summary.cell(row, col + 22).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 22).fill = fill_pale_blue
    summary.cell(row, col + 23).value = 'Bmk Beta'
    summary.cell(row, col + 23).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 23).fill = fill_pale_blue
    summary.cell(row, col + 24).value = 'Market'
    summary.cell(row, col + 24).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 24).fill = fill_pale_blue
    summary.cell(row, col + 25).value = 'Alpha Beta Adjusted'
    summary.cell(row, col + 25).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col + 25).fill = fill_pale_blue

    for ix, name in enumerate(active['names']):
        row = row + 1
        summary.cell(row, col).value = name
        summary.cell(row, col).alignment = styles.Alignment(horizontal='left')
        summary.cell(row, col + 1).value = managed['annualized returns'].loc['values', name]
        summary.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        summary.cell(row, col + 2).value = benchmark['annualized returns'].loc['values', name]
        summary.cell(row, col + 2).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        summary.cell(row, col + 3).value = active['annualized returns'].loc['values', name]
        summary.cell(row, col + 3).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 3).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        if summary.cell(row, col + 3).value > 0.05:
            summary.cell(row, col + 3).font = font_blue_bold
            summary.cell(row, col + 3).fill = fill_pale_green
        elif summary.cell(row, col + 3).value < -0.05:
            summary.cell(row, col + 3).font = font_red_bold
            summary.cell(row, col + 3).fill = fill_yellow
        summary.cell(row, col + 4).value = managed['volatilities'].loc['values', name]
        summary.cell(row, col + 4).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 4).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        summary.cell(row, col + 5).value = benchmark['volatilities'].loc['values', name]
        summary.cell(row, col + 5).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 5).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        summary.cell(row, col + 6).value = active['volatilities'].loc['values', name]
        summary.cell(row, col + 6).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 6).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        summary.cell(row, col + 7).value = managed['information ratios'].loc['values', name]
        summary.cell(row, col + 7).alignment = styles.Alignment(horizontal='center')
        if summary.cell(row, col + 7).value > 1.0:
            summary.cell(row, col + 7).font = font_blue_bold
            summary.cell(row, col + 7).fill = fill_pale_green
        elif summary.cell(row, col + 7).value < -1.0:
            summary.cell(row, col + 7).font = font_red_bold
            summary.cell(row, col + 7).fill = fill_yellow
        summary.cell(row, col + 7).number_format = styles.numbers.FORMAT_NUMBER_00
        summary.cell(row, col + 8).value = benchmark['information ratios'].loc['values', name]
        summary.cell(row, col + 8).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 8).number_format = styles.numbers.FORMAT_NUMBER_00
        summary.cell(row, col + 9).value = active['information ratios'].loc['values', name]
        summary.cell(row, col + 9).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 9).number_format = styles.numbers.FORMAT_NUMBER_00
        summary.cell(row, col + 10).value = managed['draw down'].loc['values', name]
        summary.cell(row, col + 10).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 10).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        summary.cell(row, col + 11).value = managed['draw down'].loc['duration', name]
        summary.cell(row, col + 11).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 11).number_format = styles.numbers.FORMAT_NUMBER
        summary.cell(row, col + 12).value = managed['draw down'].loc['start', name]
        summary.cell(row, col + 12).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 12).number_format = styles.numbers.FORMAT_DATE_DDMMYY
        summary.cell(row, col + 13).value = managed['draw down'].loc['end', name]
        summary.cell(row, col + 13).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 13).number_format = styles.numbers.FORMAT_DATE_DDMMYY
        summary.cell(row, col + 14).value = benchmark['draw down'].loc['values', bmk].to_numpy(dtype='float64')[0]
        summary.cell(row, col + 14).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 14).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        summary.cell(row, col + 15).value = benchmark['draw down'].loc['duration', bmk].to_numpy(dtype='float64')[0]
        summary.cell(row, col + 15).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 15).number_format = styles.numbers.FORMAT_NUMBER
        summary.cell(row, col + 16).value = benchmark['draw down'].loc['start', bmk].to_numpy()[0]
        summary.cell(row, col + 16).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 16).number_format = styles.numbers.FORMAT_DATE_DDMMYY
        summary.cell(row, col + 17).value = benchmark['draw down'].loc['end', bmk].to_numpy()[0]
        summary.cell(row, col + 17).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 17).number_format = styles.numbers.FORMAT_DATE_DDMMYY
        summary.cell(row, col + 18).value = active['draw down'].loc['values', name]
        summary.cell(row, col + 18).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 18).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        summary.cell(row, col + 19).value = active['draw down'].loc['duration', name]
        summary.cell(row, col + 19).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 19).number_format = styles.numbers.FORMAT_NUMBER
        summary.cell(row, col + 20).value = active['draw down'].loc['start', name]
        summary.cell(row, col + 20).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 20).number_format = styles.numbers.FORMAT_DATE_DDMMYY
        summary.cell(row, col + 21).value = active['draw down'].loc['end', name]
        summary.cell(row, col + 21).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 21).number_format = styles.numbers.FORMAT_DATE_DDMMYY
        summary.cell(row, col + 22).value = managed['beta'].loc['beta', name]
        summary.cell(row, col + 22).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 22).number_format = styles.numbers.FORMAT_NUMBER_00
        summary.cell(row, col + 23).value = benchmark['beta'].loc['beta', name]
        summary.cell(row, col + 23).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 23).number_format = styles.numbers.FORMAT_NUMBER_00
        summary.cell(row, col + 24).value = market['annualized returns']
        summary.cell(row, col + 24).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 24).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        summary.cell(row, col + 25).value = managed['annualized returns'].loc['values', name] - \
                                            managed['beta'].loc['beta', name] * \
                                            market['annualized returns']
        summary.cell(row, col + 25).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col + 25).number_format = styles.numbers.FORMAT_PERCENTAGE_00

    # correlations
    display(f"exporting return analysis 'correlations'")
    row = 1
    col = 1
    corr.cell(row, col).value = 'Return Summary'
    corr.cell(row, col).font = font_blue_bold
    row = row + 1
    corr.cell(row, col).value = 'Start Date'
    corr.cell(row, col+1).value = result['dates'][0].strftime(util.YY_MM_DD_format)
    corr.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    corr.cell(row, col).value = 'End Date'
    corr.cell(row, col+1).value = result['dates'][0-1].strftime(util.YY_MM_DD_format)
    corr.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    row = row + 1
    corr.cell(row, col).value = 'Benchmark'
    corr.cell(row, col+1).value = result['benchmark']['names'][0]
    corr.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    corr.cell(row, col).value = 'Market'
    corr.cell(row, col+1).value = result['market']['names'][0]
    corr.cell(row, col+1).alignment = styles.Alignment(horizontal='center')

    row = row + 2
    correlations = result['active']['correlations']
    for ix, f_name in enumerate(correlations.columns):
        corr.cell(row, col+ix+1).value = f_name
        corr.cell(row, col+ix+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    for ix, f_name in enumerate(correlations.index):
        corr.cell(row, col).value = f_name
        corr.cell(row, col).alignment = styles.Alignment(horizontal='center')
        for jx, g_name in enumerate(correlations.index):
            corr.cell(row, col+jx+1).value = correlations.iloc[ix, jx]
            corr.cell(row, col+jx+1).number_format = styles.numbers.FORMAT_PERCENTAGE
            corr.cell(row, col+jx+1).alignment = styles.Alignment(horizontal='center')
        row = row + 1

    # horizons
    sheet = horizons
    mh = result['managed']['horizons']
    ah = result['active']['horizons']
    bh = result['benchmark']['horizons']
    mkh = result['market']['horizons']
    display(f"exporting return analysis 'horizon'")
    row = 1
    col = 1
    sheet.cell(row, col).value = 'Return Summary'
    sheet.cell(row, col).font = font_blue_bold
    row = row + 1
    sheet.cell(row, col).value = 'Start Date'
    sheet.cell(row, col+1).value = result['dates'][0].strftime(util.YY_MM_DD_format)
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    sheet.cell(row, col).value = 'End Date'
    sheet.cell(row, col+1).value = result['dates'][0-1].strftime(util.YY_MM_DD_format)
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    row = row + 1
    sheet.cell(row, col).value = 'Benchmark'
    sheet.cell(row, col+1).value = result['benchmark']['names'][0]
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    sheet.cell(row, col).value = 'Market'
    sheet.cell(row, col+1).value = result['market']['names'][0]
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')

    row = row + 1
    row = row + 1
    v_tag = 'returns'
    sheet.cell(row, col).value = 'Total Returns'
    sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, label in enumerate(ah[v_tag].index):
        sheet.cell(row, col + ix + 1).value = label
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in ah[v_tag].columns:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, label in enumerate(ah[v_tag].index):
            sheet.cell(row, col + ix + 1).value = mh['returns'].loc[label, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    sheet.cell(row, col).value = bmk
    sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
    sheet.cell(row, col).fill = fill_pale_salmon
    for ix, label in enumerate(bh[v_tag].index):
        sheet.cell(row, col + ix + 1).value = bh['returns'].loc[label, name]
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_salmon
    row = row + 1
    sheet.cell(row, col).value = mkt
    sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
    sheet.cell(row, col).fill = fill_pale_salmon
    for ix, label in enumerate(bh[v_tag].index):
        sheet.cell(row, col + ix + 1).value = mkh['returns'].loc[label, name]
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_salmon

    row = row + 1
    row = row + 1
    v_tag = 'returns'
    sheet.cell(row, col).value = 'Alphas'
    sheet.cell(row, col).alignment = styles.Alignment(horizontal='center')
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, label in enumerate(ah[v_tag].index):
        sheet.cell(row, col + ix + 1).value = label
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in ah[v_tag].columns:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, label in enumerate(ah[v_tag].index):
            sheet.cell(row, col + ix + 1).value = ah['returns'].loc[label, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
            if ah['returns'].loc[label, name] > 0.05:
                sheet.cell(row, col + ix + 1).fill = fill_pale_green
                sheet.cell(row, col + ix + 1).font = font_blue_bold
            elif ah['returns'].loc[label, name] < -0.05:
                sheet.cell(row, col + ix + 1).fill = fill_yellow
                sheet.cell(row, col + ix + 1).font = font_red_bold

    row = row + 1
    row = row + 1
    v_tag = 'volatilities'
    sheet.cell(row, col).value = 'TE'
    sheet.cell(row, col).alignment = styles.Alignment(horizontal='center')
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, label in enumerate(ah[v_tag].index):
        sheet.cell(row, col + ix + 1).value = label
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in ah[v_tag].columns:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, label in enumerate(ah[v_tag].index):
            sheet.cell(row, col + ix + 1).value = ah[v_tag].loc[label, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    row = row + 1
    v_tag = 'information ratios'
    sheet.cell(row, col).value = 'Information Ratios'
    sheet.cell(row, col).alignment = styles.Alignment(horizontal='center')
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, label in enumerate(ah[v_tag].index):
        sheet.cell(row, col + ix + 1).value = label
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in ah[v_tag].columns:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, label in enumerate(ah[v_tag].index):
            sheet.cell(row, col + ix + 1).value = ah[v_tag].loc[label, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    row = row + 1
    v_tag = 'hit rates'
    sheet.cell(row, col).value = 'Hit Rates'
    sheet.cell(row, col).alignment = styles.Alignment(horizontal='center')
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, label in enumerate(ah[v_tag].index):
        sheet.cell(row, col + ix + 1).value = label
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in ah[v_tag].columns:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, label in enumerate(ah[v_tag].index):
            sheet.cell(row, col + ix + 1).value = ah[v_tag].loc[label, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
    # annual
    sheet = year
    display(f"exporting return analysis 'annual'")
    row = 1
    col = 1
    sheet.cell(row, col).value = 'Return Summary'
    sheet.cell(row, col).font = font_blue_bold
    row = row + 1
    sheet.cell(row, col).value = 'Start Date'
    sheet.cell(row, col+1).value = result['dates'][0].strftime(util.YY_MM_DD_format)
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    sheet.cell(row, col).value = 'End Date'
    sheet.cell(row, col+1).value = result['dates'][0-1].strftime(util.YY_MM_DD_format)
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    row = row + 1
    sheet.cell(row, col).value = 'Benchmark'
    sheet.cell(row, col+1).value = result['benchmark']['names'][0]
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    sheet.cell(row, col).value = 'Market'
    sheet.cell(row, col+1).value = result['market']['names'][0]
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')

    row = row + 2
    sheet.cell(row, col).value = 'Total Returns'
    sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['annual returns'].index):
        sheet.cell(row, col + ix + 1).value = y
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['annual returns'].index):
            sheet.cell(row, col + ix + 1).value = managed['annual actual returns'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    sheet.cell(row, col).value = bmk
    sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
    sheet.cell(row, col).fill = fill_pale_salmon
    for ix, y in enumerate(benchmark['annual returns'].index):
        sheet.cell(row, col + ix + 1).value = benchmark['annual actual returns'].iloc[ix, 0]
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_salmon

    row = row + 2
    sheet.cell(row, col).value = 'Alpha'
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['annual returns'].index):
        sheet.cell(row, col + ix + 1).value = y
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue

    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['annual returns'].index):
            sheet.cell(row, col + ix + 1).value = active['annual actual returns'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
            if active['annual actual returns'].loc[y, name] > 0.05:
                sheet.cell(row, col + ix + 1).fill = fill_pale_green
                sheet.cell(row, col + ix + 1).font = font_blue_bold
            elif active['annual actual returns'].loc[y, name] < -0.05:
                sheet.cell(row, col + ix + 1).fill = fill_yellow
                sheet.cell(row, col + ix + 1).font = font_red_bold

    row = row + 2
    sheet.cell(row, col).value = 'TE'
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['annual volatilities'].index):
        sheet.cell(row, col + ix + 1).value = y
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['annual volatilities'].index):
            sheet.cell(row, col + ix + 1).value = active['annual volatilities'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
    row = row + 2
    sheet.cell(row, col).value = 'IR'
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['annual information ratios'].index):
        sheet.cell(row, col + ix + 1).value = y
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['annual information ratios'].index):
            sheet.cell(row, col + ix + 1).value = active['annual information ratios'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
            if active['annual information ratios'].loc[y, name] > 1:
                sheet.cell(row, col + ix + 1).fill = fill_pale_green
                sheet.cell(row, col + ix + 1).font = font_blue_bold
            elif active['annual information ratios'].loc[y, name] < -1:
                sheet.cell(row, col + ix + 1).fill = fill_yellow
                sheet.cell(row, col + ix + 1).font = font_red_bold

    row = row + 2
    sheet.cell(row, col).value = 'Beta'
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['annual betas'].index):
        sheet.cell(row, col + ix + 1).value = y
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['annual betas'].index):
            sheet.cell(row, col + ix + 1).value = managed['annual betas'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
    # semi-annual
    sheet = semiannual
    display(f"exporting return analysis 'semiannual'")
    row = 1
    col = 1
    sheet.cell(row, col).value = 'Return Summary'
    sheet.cell(row, col).font = font_blue_bold
    row = row + 1
    sheet.cell(row, col).value = 'Start Date'
    sheet.cell(row, col+1).value = result['dates'][0].strftime(util.YY_MM_DD_format)
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    sheet.cell(row, col).value = 'End Date'
    sheet.cell(row, col+1).value = result['dates'][0-1].strftime(util.YY_MM_DD_format)
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    row = row + 1
    sheet.cell(row, col).value = 'Benchmark'
    sheet.cell(row, col+1).value = result['benchmark']['names'][0]
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    sheet.cell(row, col).value = 'Market'
    sheet.cell(row, col+1).value = result['market']['names'][0]
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')

    row = row + 2
    sheet.cell(row, col).value = 'Total Returns'
    sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['semiannual returns'].index):
        sheet.cell(row, col + ix + 1).value = y
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['semiannual returns'].index):
            sheet.cell(row, col + ix + 1).value = managed['semiannual returns'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    sheet.cell(row, col).value = bmk
    sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
    sheet.cell(row, col).fill = fill_pale_salmon
    for ix, y in enumerate(active['semiannual returns'].index):
        sheet.cell(row, col + ix + 1).value = benchmark['semiannual returns'].loc[y, name]
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_salmon
    row = row + 2
    sheet.cell(row, col).value = 'Alpha'
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['semiannual returns'].index):
        sheet.cell(row, col + ix + 1).value = y
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['semiannual returns'].index):
            sheet.cell(row, col + ix + 1).value = active['semiannual returns'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
            if active['semiannual returns'].loc[y, name] > 0.05:
                sheet.cell(row, col + ix + 1).fill = fill_pale_green
                sheet.cell(row, col + ix + 1).font = font_blue_bold
            elif active['semiannual returns'].loc[y, name] < -0.05:
                sheet.cell(row, col + ix + 1).fill = fill_yellow
                sheet.cell(row, col + ix + 1).font = font_red_bold

    row = row + 2
    sheet.cell(row, col).value = 'TE'
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['semiannual volatilities'].index):
        sheet.cell(row, col + ix + 1).value = y
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['semiannual volatilities'].index):
            sheet.cell(row, col + ix + 1).value = active['semiannual volatilities'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
    row = row + 2
    sheet.cell(row, col).value = 'IR'
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['semiannual information ratios'].index):
        sheet.cell(row, col + ix + 1).value = y
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['semiannual information ratios'].index):
            sheet.cell(row, col + ix + 1).value = active['semiannual information ratios'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
            if active['semiannual information ratios'].loc[y, name] > 1:
                sheet.cell(row, col + ix + 1).fill = fill_pale_green
                sheet.cell(row, col + ix + 1).font = font_blue_bold
            elif active['semiannual information ratios'].loc[y, name] < -1:
                sheet.cell(row, col + ix + 1).fill = fill_yellow
                sheet.cell(row, col + ix + 1).font = font_red_bold

    # quarter
    sheet = quarter
    display(f"exporting return analysis 'quarter'")
    row = 1
    col = 1
    sheet.cell(row, col).value = 'Return Summary'
    sheet.cell(row, col).font = font_blue_bold
    row = row + 1
    sheet.cell(row, col).value = 'Start Date'
    sheet.cell(row, col+1).value = result['dates'][0].strftime(util.YY_MM_DD_format)
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    sheet.cell(row, col).value = 'End Date'
    sheet.cell(row, col+1).value = result['dates'][0-1].strftime(util.YY_MM_DD_format)
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    row = row + 1
    sheet.cell(row, col).value = 'Benchmark'
    sheet.cell(row, col+1).value = result['benchmark']['names'][0]
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    sheet.cell(row, col).value = 'Market'
    sheet.cell(row, col+1).value = result['market']['names'][0]
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')

    row = row + 2
    sheet.cell(row, col).value = 'Total Returns'
    sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['quarterly returns'].index):
        sheet.cell(row, col + ix + 1).value = y
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['quarterly returns'].index):
            sheet.cell(row, col + ix + 1).value = managed['quarterly returns'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    sheet.cell(row, col).value = bmk
    sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
    sheet.cell(row, col).fill = fill_pale_salmon
    for ix, y in enumerate(active['quarterly returns'].index):
        sheet.cell(row, col + ix + 1).value = benchmark['quarterly returns'].loc[y, name]
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_salmon

    row = row + 2
    sheet.cell(row, col).value = 'Alpha'
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['quarterly returns'].index):
        sheet.cell(row, col + ix + 1).value = y
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['quarterly returns'].index):
            sheet.cell(row, col + ix + 1).value = active['quarterly returns'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
            if active['quarterly returns'].loc[y, name] > 0.05:
                sheet.cell(row, col + ix + 1).fill = fill_pale_green
                sheet.cell(row, col + ix + 1).font = font_blue_bold
            elif active['quarterly returns'].loc[y, name] < -0.05:
                sheet.cell(row, col + ix + 1).fill = fill_yellow
                sheet.cell(row, col + ix + 1).font = font_red_bold
    row = row + 2
    sheet.cell(row, col).value = 'TE'
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['quarterly volatilities'].index):
        sheet.cell(row, col + ix + 1).value = y
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['quarterly volatilities'].index):
            sheet.cell(row, col + ix + 1).value = active['quarterly volatilities'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
    row = row + 2
    sheet.cell(row, col).value = 'IR'
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['quarterly information ratios'].index):
        sheet.cell(row, col + ix + 1).value = y
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['quarterly information ratios'].index):
            sheet.cell(row, col + ix + 1).value = active['quarterly information ratios'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
            if active['quarterly information ratios'].loc[y, name] > 1:
                sheet.cell(row, col + ix + 1).fill = fill_pale_green
                sheet.cell(row, col + ix + 1).font = font_blue_bold
            elif active['quarterly information ratios'].loc[y, name] < -1:
                sheet.cell(row, col + ix + 1).fill = fill_yellow
                sheet.cell(row, col + ix + 1).font = font_red_bold

    # month
    sheet = month
    display(f"exporting return analysis 'month'")
    row = 1
    col = 1
    sheet.cell(row, col).value = 'Return Summary'
    sheet.cell(row, col).font = font_blue_bold
    row = row + 1
    sheet.cell(row, col).value = 'Start Date'
    sheet.cell(row, col+1).value = result['dates'][0].strftime(util.YY_MM_DD_format)
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    sheet.cell(row, col).value = 'End Date'
    sheet.cell(row, col+1).value = result['dates'][0-1].strftime(util.YY_MM_DD_format)
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    row = row + 1
    sheet.cell(row, col).value = 'Benchmark'
    sheet.cell(row, col+1).value = result['benchmark']['names'][0]
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    sheet.cell(row, col).value = 'Market'
    sheet.cell(row, col+1).value = result['market']['names'][0]
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')

    row = row + 2
    sheet.cell(row, col).value = 'Total Returns'
    sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['monthly returns'].index):
        sheet.cell(row, col + ix + 1).value = calendar.month_name[y]
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['monthly returns'].index):
            sheet.cell(row, col + ix + 1).value = managed['monthly returns'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    sheet.cell(row, col).value = bmk
    sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
    sheet.cell(row, col).fill = fill_pale_salmon
    for ix, y in enumerate(active['monthly returns'].index):
        sheet.cell(row, col + ix + 1).value = benchmark['monthly returns'].loc[y, name]
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_salmon

    row = row + 2
    sheet.cell(row, col).value = 'Alpha'
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['monthly returns'].index):
        sheet.cell(row, col + ix + 1).value = calendar.month_name[y]
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['monthly returns'].index):
            sheet.cell(row, col + ix + 1).value = active['monthly returns'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
            if active['monthly returns'].loc[y, name] > 0.05:
                sheet.cell(row, col + ix + 1).fill = fill_pale_green
                sheet.cell(row, col + ix + 1).font = font_blue_bold
            elif active['monthly returns'].loc[y, name] < -0.05:
                sheet.cell(row, col + ix + 1).fill = fill_yellow
                sheet.cell(row, col + ix + 1).font = font_red_bold
    row = row + 2
    sheet.cell(row, col).value = 'TE'
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['monthly volatilities'].index):
        sheet.cell(row, col + ix + 1).value = calendar.month_name[y]
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['monthly volatilities'].index):
            sheet.cell(row, col + ix + 1).value = active['monthly volatilities'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
    row = row + 2
    sheet.cell(row, col).value = 'IR'
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['monthly information ratios'].index):
        sheet.cell(row, col + ix + 1).value = calendar.month_name[y]
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['monthly information ratios'].index):
            sheet.cell(row, col + ix + 1).value = active['monthly information ratios'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
            if active['monthly information ratios'].loc[y, name] > 1:
                sheet.cell(row, col + ix + 1).fill = fill_pale_green
                sheet.cell(row, col + ix + 1).font = font_blue_bold
            elif active['monthly information ratios'].loc[y, name] < -1:
                sheet.cell(row, col + ix + 1).fill = fill_yellow
                sheet.cell(row, col + ix + 1).font = font_red_bold

    # regime
    sheet = regime
    display(f"exporting return analysis 'regime'")
    row = 1
    col = 1
    sheet.cell(row, col).value = 'Return Summary'
    sheet.cell(row, col).font = font_blue_bold
    row = row + 1
    sheet.cell(row, col).value = 'Start Date'
    sheet.cell(row, col+1).value = result['dates'][0].strftime(util.YY_MM_DD_format)
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    sheet.cell(row, col).value = 'End Date'
    sheet.cell(row, col+1).value = result['dates'][0-1].strftime(util.YY_MM_DD_format)
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    row = row + 1
    sheet.cell(row, col).value = 'Benchmark'
    sheet.cell(row, col+1).value = result['benchmark']['names'][0]
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    sheet.cell(row, col).value = 'Market'
    sheet.cell(row, col+1).value = result['market']['names'][0]
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 2
    sheet.cell(row, col).value = 'Regime'
    sheet.cell(row, col+1).value = indicator
    sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')

    row = row + 2
    sheet.cell(row, col).value = 'Total Returns'
    sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['regime returns'].index):
        sheet.cell(row, col + ix + 1).value = md.get_regime_name(y, indicator)
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['regime returns'].index):
            sheet.cell(row, col + ix + 1).value = managed['regime returns'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    sheet.cell(row, col).value = bmk
    sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
    sheet.cell(row, col).fill = fill_pale_salmon
    for ix, y in enumerate(benchmark['regime returns'].index):
        sheet.cell(row, col + ix + 1).value = benchmark['regime returns'].iloc[ix, 0]
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_salmon
    row = row + 1
    sheet.cell(row, col).value = mkt
    sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
    sheet.cell(row, col).fill = fill_pale_salmon
    for ix, y in enumerate(market['regime returns'].index):
        sheet.cell(row, col + ix + 1).value = market['regime returns'].iloc[ix, 0]
        sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_salmon
    row = row + 2
    sheet.cell(row, col).value = 'Alpha'
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['regime returns'].index):
        sheet.cell(row, col + ix + 1).value = md.get_regime_name(y, indicator)
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['regime returns'].index):
            sheet.cell(row, col + ix + 1).value = active['regime returns'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
            if active['regime returns'].loc[y, name] > 0.05:
                sheet.cell(row, col + ix + 1).fill = fill_pale_green
                sheet.cell(row, col + ix + 1).font = font_blue_bold
            elif active['regime returns'].loc[y, name] < -0.05:
                sheet.cell(row, col + ix + 1).fill = fill_yellow
                sheet.cell(row, col + ix + 1).font = font_red_bold

    row = row + 2
    sheet.cell(row, col).value = 'TE'
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['regime volatilities'].index):
        sheet.cell(row, col + ix + 1).value = md.get_regime_name(y, indicator)
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['regime volatilities'].index):
            sheet.cell(row, col + ix + 1).value = active['regime volatilities'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
    row = row + 2
    sheet.cell(row, col).value = 'IR'
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['regime information ratios'].index):
        sheet.cell(row, col + ix + 1).value = md.get_regime_name(y, indicator)
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['regime information ratios'].index):
            sheet.cell(row, col + ix + 1).value = active['regime information ratios'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
            if active['regime information ratios'].loc[y, name] > 1:
                sheet.cell(row, col + ix + 1).fill = fill_pale_green
                sheet.cell(row, col + ix + 1).font = font_blue_bold
            elif active['regime information ratios'].loc[y, name] < -1:
                sheet.cell(row, col + ix + 1).fill = fill_yellow
                sheet.cell(row, col + ix + 1).font = font_red_bold

    row = row + 2
    sheet.cell(row, col).value = 'Beta'
    sheet.cell(row, col).fill = fill_pale_blue
    for ix, y in enumerate(active['regime betas'].index):
        sheet.cell(row, col + ix + 1).value = md.get_regime_name(y, indicator)
        sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + ix + 1).fill = fill_pale_blue
    for name in active['names']:
        row = row + 1
        sheet.cell(row, col).value = name
        sheet.cell(row, col).alignment = styles.Alignment(horizontal='left')
        for ix, y in enumerate(active['regime information ratios'].index):
            sheet.cell(row, col + ix + 1).value = managed['regime betas'].loc[y, name]
            sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_NUMBER_00
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')

    if include_ts:
        # TS
        ts = result['ts']
        types = ['managed', 'benchmark', 'active']
        for tp in types:
            td = ts[tp]
            if tp == 'managed':
                sheet = tsm
            elif tp == 'benchmark':
                sheet = tsb
                td = td[[td.columns[0]]]
            else:
                sheet = tsa
            display(f"exporting {tp} return timeseries")
            row = 1
            col = 1
            sheet.cell(row, col).value = 'Return Summary'
            sheet.cell(row, col).font = font_blue_bold
            row = row + 1
            sheet.cell(row, col).value = 'Start Date'
            sheet.cell(row, col+1).value = result['dates'][0].strftime(util.YY_MM_DD_format)
            sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
            row = row + 1
            sheet.cell(row, col).value = 'End Date'
            sheet.cell(row, col+1).value = result['dates'][0-1].strftime(util.YY_MM_DD_format)
            sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
            row = row + 1
            row = row + 1
            sheet.cell(row, col).value = 'Benchmark'
            sheet.cell(row, col+1).value = result['benchmark']['names'][0]
            sheet.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
            row = row + 1
            sheet.cell(row, col).value = 'Market'
            sheet.cell(row, col + 1).value = result['market']['names'][0]
            sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')

            row = row + 2
            sheet.cell(row, col).value = 'Date'
            for ix, y in enumerate(td.columns):
                sheet.cell(row, col + ix + 1).value = y
                sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
            for d in td.index:
                row = row + 1
                sheet.cell(row, col).value = d
                for ix, name in enumerate(td.columns):
                    sheet.cell(row, col + ix + 1).value = td.loc[d, name]
                    sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
                    sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='left')
        mts = rt.daily_to_period(ts['managed'], calendar_str=calendar_str)
        sheet = tsy
        display(f"exporting monthly return timeseries")
        row = 1
        col = 1
        sheet.cell(row, col).value = 'Return Summary'
        sheet.cell(row, col).font = font_blue_bold
        row = row + 1
        sheet.cell(row, col).value = 'Start Date'
        sheet.cell(row, col + 1).value = result['dates'][0].strftime(util.YY_MM_DD_format)
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = 'End Date'
        sheet.cell(row, col + 1).value = result['dates'][0 - 1].strftime(util.YY_MM_DD_format)
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        row = row + 1
        sheet.cell(row, col).value = 'Benchmark'
        sheet.cell(row, col + 1).value = result['benchmark']['names'][0]
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = 'Market'
        sheet.cell(row, col + 1).value = result['market']['names'][0]
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')

        row = row + 2
        sheet.cell(row, col).value = 'Date'
        for ix, y in enumerate(mts.columns):
            sheet.cell(row, col + ix + 1).value = y
            sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='center')
        for d in mts.index:
            row = row + 1
            sheet.cell(row, col).value = d
            for ix, name in enumerate(mts.columns):
                sheet.cell(row, col + ix + 1).value = mts.loc[d, name]
                sheet.cell(row, col + ix + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
                sheet.cell(row, col + ix + 1).alignment = styles.Alignment(horizontal='left')

    wb.save(file)
    display(f"Successfully generated return statistics report to file \n{file}")


def substitute(names, mapping):
    substituted = names
    for i, n in enumerate(names):
        ix = np.where(mapping.iloc[:, 0] == n)[0]
        if len(ix) == 0:
            display(f"{n} not replaced: no mapping found")
            continue
        if len(ix) > 1:
            display(f"{n} has {len(ix)} mapped names: using the first")
        substituted[i] = mapping.iloc[ix[0], 1]
    return substituted


def get_iq_indices():
    file = os.path.join(util.default_output_location('reports'), 'iq_indices.xlsx')
    data = pd.read_excel(file)
    indices = data['Indices'].to_numpy()
    return indices


def export_indices_performance(bus_day, indices=None, calendar_str='US', location=None):
    if indices is None:
        indices = get_iq_indices()
    benchmarks = np.full((len(indices),), None)
    names = np.full((len(indices),), None)
    types = np.full((len(indices),), None)
    for ix, ind in enumerate(indices):
        o = root.load_object(ind)
        benchmarks[ix] = o.benchmark
        types[ix] = o.benchmark_security_type
        names[ix] = o.description

    bus_days = util.load_business_days(calendar_str, None, bus_day)
    mends = util.load_business_days(calendar_str, None, bus_day, 'MONTHEND')
    mends = mends[mends < bus_days[-1]]
    yends = util.load_business_days(calendar_str, None, bus_day, 'YEAREND')
    yends = yends[yends < bus_days[-1]]

    # index returns
    sec_type = 'QSR'
    ret = md.get_returns(bus_days[-252*10], bus_days[-1], indices, calendar_str, security_type=sec_type)
    display(f"Loaded {len(ret.index)} X {len(indices)} index returns")
    df = pd.DataFrame(index=indices, columns=['1 day', '5 day', 'MTD', '1 month', '3 months', '6 months',
                                              '9 months', 'YTD', '1 yr', '3 yr', '5 yr', '7 yr', '10 yr', 'Turnover',
                                              'Dividend Yield', 'Holdings', 'P/B', 'P/E'])
    sec_ids = ret.columns.to_numpy()
    div = root.load_object('DividendYield')
    bp = root.load_object('B2P')
    ep = root.load_object('E2P')
    for sid in sec_ids:
        try:
            # dividend
            t_data = div.load_values('DESCRIPTOR', bus_days[-1], bus_days[-1], sid, composite_flag=True)
            df.loc[sid, 'Dividend Yield'] = t_data.loc[t_data.index[0], sid]
            # b2p
            t_data = bp.load_values('positive median', bus_days[-1], bus_days[-1], sid, composite_flag=True)
            df.loc[sid, 'P/B'] = 1. / t_data.loc[t_data.index[0], sid]
            # e2p
            t_data = ep.load_values('positive median', bus_days[-1], bus_days[-1], sid, composite_flag=True)
            df.loc[sid, 'P/E'] = 1. / t_data.loc[t_data.index[0], sid]
            # holdings
            p = port.get_portfolio_weights(bus_days[-1], bus_days[-1], sid)
            df.loc[sid, 'Holdings'] = (p > 1e-6).sum(axis=1).sum()
        except ValueError as ve:
            display(ve)
            continue
        except Exception as ee:
            display(ee)
            continue
    for sid in sec_ids:
        try:

            # turnover
            obj = root.load_object(sid)
            t = obj.load_turnover()
            if t.empty:
                continue
            freq = obj.descriptor_frequency
            if hasattr(obj, 'rebalance_frequency'):
                freq = obj.rebalance_frequency
            if freq in ('MONTHLY', 'MONTHEND'):
                multiplier = 6
            elif freq == 'DAILY':
                multiplier = 126
            elif freq == 'WEEKLY':
                multiplier = 252 / 5
            elif freq in ('QUARTERLY', 'QUARTEREND'):
                multiplier = 2
            elif freq in ('SEMI-ANNUAL', 'SEMI_ANNUALLY', 'SEMIANNUAL', 'SEMIANNUALLY', 'HALF-YEAR', 'HALFYEAR'):
                multiplier = 1
            elif freq in ('ANNUAL', 'ANNUALLY', 'YEARLY', 'YEAREND', 'YEAR'):
                multiplier = 0.5
            else:
                raise ValueError('Unrecognized frequency')

            df.loc[sid, 'Turnover'] = np.nanmean(t.to_numpy()) * multiplier
        except Exception as es:
            display(es)
            display(f"{sid}: unable to get turnover")
    df.loc[sec_ids, '1 day'] = ret.loc[ret.index[-1], sec_ids]
    df.loc[sec_ids, '5 day'] = np.prod(1 + ret.loc[ret.index[-5:], sec_ids]) - 1
    ia = np.argmax(ret.index>mends[-1])
    df.loc[sec_ids, 'MTD'] = np.prod(1 + ret.loc[ret.index[ia:], sec_ids]) - 1
    df.loc[sec_ids, '1 month'] = np.prod(1 + ret.loc[ret.index[-21:], sec_ids]) - 1
    df.loc[sec_ids, '3 months'] = np.prod(1 + ret.loc[ret.index[-21 * 3:], sec_ids]) - 1
    df.loc[sec_ids, '6 months'] = np.prod(1 + ret.loc[ret.index[-21 * 6:], sec_ids]) - 1
    df.loc[sec_ids, '9 months'] = np.prod(1 + ret.loc[ret.index[-21 * 9:], sec_ids]) - 1
    ib = np.argmax(ret.index > yends[-1])
    df.loc[sec_ids, 'YTD'] = np.prod(1 + ret.loc[ret.index[ib:], sec_ids]) - 1
    df.loc[sec_ids, '1 yr'] = np.prod(1 + ret.loc[ret.index[-21 * 12:], sec_ids]) - 1
    df.loc[sec_ids, '3 yr'] = np.prod(1 + ret.loc[ret.index[-21 * 36:], sec_ids])**(1/3.) - 1
    df.loc[sec_ids, '5 yr'] = np.prod(1 + ret.loc[ret.index[-21 * 60:], sec_ids])**(1/5.) - 1
    df.loc[sec_ids, '7 yr'] = np.prod(1 + ret.loc[ret.index[-21 * 84:], sec_ids])**(1/7.) - 1
    df.loc[sec_ids, '10 yr'] = np.prod(1 + ret.loc[ret.index, sec_ids])**(1/10.) - 1

    # benchmark returns
    unique_benchmarks = np.unique(benchmarks)
    ub_types = np.full((len(unique_benchmarks),), 'FUND')
    c, i1, i2=intersect(unique_benchmarks, benchmarks)
    ub_types[i1] = types[i2]
    unique_types = np.unique(ub_types)
    bret = pd.DataFrame(index=ret.index, columns=unique_benchmarks, dtype='float64')
    for t in unique_types:
        try:
            ix = np.where(ub_types == t)[0]
            r = md.get_returns(ret.index[0], ret.index[-1], unique_benchmarks[ix], calendar_str,
                               security_type=t)
            bret.loc[r.index, r.columns] = r
        except ValueError as ve:
            display(ve)
    bf = pd.DataFrame(index=benchmarks, columns=df.columns)
    display(f"Loaded  {len(bret.index)} X {len(benchmarks)} unique benchmark returns")
    sids = bret.columns.to_numpy()
    bf.loc[sids, '1 day'] = bret.loc[bret.index[-1], sids]
    bf.loc[sids, '5 day'] = np.prod(1 + bret.loc[bret.index[-5:], sids]) - 1
    bf.loc[sids, 'MTD'] = np.prod(1 + bret.loc[bret.index[ia:], sids]) - 1
    bf.loc[sids, '1 month'] = np.prod(1 + bret.loc[bret.index[-21:], sids]) - 1
    bf.loc[sids, '3 months'] = np.prod(1 + bret.loc[bret.index[-21 * 3:], sids]) - 1
    bf.loc[sids, '6 months'] = np.prod(1 + bret.loc[bret.index[-21 * 6:], sids]) - 1
    bf.loc[sids, '9 months'] = np.prod(1 + bret.loc[bret.index[-21 * 9:], sids]) - 1
    bf.loc[sids, 'YTD'] = np.prod(1 + bret.loc[bret.index[ib:], sids]) - 1
    bf.loc[sids, '1 yr'] = np.prod(1 + bret.loc[bret.index[-21 * 12:], sids]) - 1
    bf.loc[sids, '3 yr'] = np.prod(1 + bret.loc[bret.index[-21 * 36:], sids])**(1/3.) - 1
    bf.loc[sids, '5 yr'] = np.prod(1 + bret.loc[bret.index[-21 * 60:], sids])**(1/5.) - 1
    bf.loc[sids, '7 yr'] = np.prod(1 + bret.loc[bret.index[-21 * 84:], sids])**(1/7.) - 1
    bf.loc[sids, '10 yr'] = np.prod(1 + bret.loc[bret.index, sids])**(1/10.) - 1

    af = pd.DataFrame(df.to_numpy() - bf.to_numpy(), index=df.index, columns=df.columns)

    # volatilities
    vf = pd.DataFrame(index=indices, columns=df.columns)
    vf.loc[sec_ids, '5 day'] = np.nanstd(ret.loc[ret.index[-5:], sec_ids], axis=0) * np.sqrt(252)
    vf.loc[sec_ids, '1 month'] = np.nanstd(ret.loc[ret.index[-21:], sec_ids], axis=0) * np.sqrt(252)
    if len(ret.index[ia:]) < 10:
        vf.loc[sec_ids, 'MTD'] = vf.loc[sec_ids, '1 month']
    else:
        vf.loc[sec_ids, 'MTD'] = np.nanstd(ret.loc[ret.index[ia:], sec_ids], axis=0) * np.sqrt(252)
    vf.loc[sec_ids, '3 months'] = np.nanstd(ret.loc[ret.index[-21 * 3:], sec_ids], axis=0) * np.sqrt(252)
    vf.loc[sec_ids, '6 months'] = np.nanstd(ret.loc[ret.index[-21 * 6:], sec_ids], axis=0) * np.sqrt(252)
    vf.loc[sec_ids, '9 months'] = np.nanstd(ret.loc[ret.index[-21 * 9:], sec_ids], axis=0) * np.sqrt(252)
    vf.loc[sec_ids, '1 yr'] = np.nanstd(ret.loc[ret.index[-21 * 12:], sec_ids], axis=0) * np.sqrt(252)
    if len(ret.index[ib:]) < 10:
        vf.loc[sec_ids, 'YTD'] = vf.loc[sec_ids, '1 yr']
    else:
        vf.loc[sec_ids, 'YTD'] = np.nanstd(ret.loc[ret.index[ib:], sec_ids], axis=0) * np.sqrt(252)
    vf.loc[sec_ids, '3 yr'] = np.nanstd(ret.loc[ret.index[-21 * 36:], sec_ids], axis=0) * np.sqrt(252)
    vf.loc[sec_ids, '5 yr'] = np.nanstd(ret.loc[ret.index[-21 * 60:], sec_ids], axis=0) * np.sqrt(252)
    vf.loc[sec_ids, '7 yr'] = np.nanstd(ret.loc[ret.index[-21 * 84:], sec_ids], axis=0) * np.sqrt(252)
    vf.loc[sec_ids, '10 yr'] = np.nanstd(ret.loc[ret.index, sec_ids], axis=0) * np.sqrt(252)

    # tracking errors
    aret = pd.DataFrame(index=ret.index, columns=indices, dtype='float64')
    for ix, s in enumerate(aret.columns):
        if s not in ret.columns:
            continue
        b = benchmarks[ix]
        if b not in bret.columns:
            continue
        aret.iloc[:, ix] = bret[b].to_numpy() - ret[s].to_numpy()
    te = pd.DataFrame(index=indices, columns=df.columns, dtype='float64')
    te.loc[sec_ids, '5 day'] = np.nanstd(aret.loc[aret.index[-5:], sec_ids], axis=0) * np.sqrt(252)
    te.loc[sec_ids, '1 month'] = np.nanstd(aret.loc[aret.index[-21:], sec_ids], axis=0) * np.sqrt(252)
    if len(aret.index[ia:]) < 10:
        te.loc[sec_ids, 'MTD'] = te.loc[sec_ids, '1 month']
    else:
        te.loc[sec_ids, 'MTD'] = np.nanstd(aret.loc[aret.index[ia:], sec_ids], axis=0) * np.sqrt(252)
    te.loc[sec_ids, '3 months'] = np.nanstd(aret.loc[aret.index[-21 * 3:], sec_ids], axis=0) * np.sqrt(252)
    te.loc[sec_ids, '6 months'] = np.nanstd(aret.loc[aret.index[-21 * 6:], sec_ids], axis=0) * np.sqrt(252)
    te.loc[sec_ids, '9 months'] = np.nanstd(aret.loc[aret.index[-21 * 9:], sec_ids], axis=0) * np.sqrt(252)
    te.loc[sec_ids, '1 yr'] = np.nanstd(aret.loc[aret.index[-21 * 12:], sec_ids], axis=0) * np.sqrt(252)
    if len(aret.index[ib:]) < 10:
        te.loc[sec_ids, 'YTD'] = te.loc[sec_ids, '1 yr']
    else:
        te.loc[sec_ids, 'YTD'] = np.nanstd(aret.loc[aret.index[ib:], sec_ids], axis=0) * np.sqrt(252)
    te.loc[sec_ids, '3 yr'] = np.nanstd(aret.loc[aret.index[-21 * 36:], sec_ids], axis=0) * np.sqrt(252)
    te.loc[sec_ids, '5 yr'] = np.nanstd(aret.loc[aret.index[-21 * 60:], sec_ids], axis=0) * np.sqrt(252)
    te.loc[sec_ids, '7 yr'] = np.nanstd(aret.loc[aret.index[-21 * 84:], sec_ids], axis=0) * np.sqrt(252)
    te.loc[sec_ids, '10 yr'] = np.nanstd(aret.loc[aret.index, sec_ids], axis=0) * np.sqrt(252)

    vf[vf == 0] = np.nan
    sharpe = df / vf
    ir = af / te
    if location is None:
        location = os.path.join(util.default_output_location('reports'), 'tmp',
                                f"{bus_days[-1].strftime(util.yyyymmdd_format)}")
    if not os.path.exists(location):
        os.makedirs(location)
    file = os.path.join(location,
                        f"Indices.Performance."
                        f"{bus_days[-1].strftime(util.yyyymmdd_format)}.xlsx")

    with pd.ExcelWriter(file) as writer:
        df.to_excel(writer, sheet_name='Performance')
        af.to_excel(writer, sheet_name='Alpha')
        bf.to_excel(writer, sheet_name='Benchmark Returns')
        vf.to_excel(writer, sheet_name='Volatilities')
        te.to_excel(writer, sheet_name='Tracking Errors')
        sharpe.to_excel(writer, sheet_name='Sharpe Ratios')
        ir.to_excel(writer, sheet_name='Information Ratios')
        ret.to_excel(writer, sheet_name='Quasar Return Series')
        bret.to_excel(writer, sheet_name='Benchmark Return Series')
    display(f"Computed and saved {len(df.index)} indices returns for {bus_days[-1]}")
    display(file)
    return df, vf, ir


def export_index_data(bus_day, include_ts=False):
    indices = get_iq_indices()
    bus_days = util.load_business_days('US', None, bus_day)
    location = os.path.join(util.default_output_location('reports'), 'tmp',
                            f"{bus_days[-1].strftime(util.yyyymmdd_format)}")
    if not os.path.exists(location):
        os.makedirs(location)
        display(f"Successfully created: {location}")
    export_indices_performance(bus_day, indices, location=location)
    for index in indices:
        try:
            obj = root.load_object(index)
            composite_flag = False
            if hasattr(obj, 'security_type'):
                if obj.security_type.upper().strip() == 'PORTFOLIO':
                    composite_flag = True
            if composite_flag:
                export_portfolio(bus_day, index, location=location, sheet='Strategy')
                export_portfolio(bus_day, index, location=location, recurse=True, sheet='Allocation')
                export_portfolio(bus_day, index, location=location, recurse=True, deep=True, sheet='Security')
            else:
                export_portfolio(bus_day, index, location=location, recurse=True)
        except Exception as ee:
            display(ee)
            display(f"Unable to print index: {index}")
            continue


def barclay_report(start_date=20020101, end_date=None, file=None, look_back=24,
                   reg_factors=['MediumMomentum', 'FCFYield', 'BookYield', 'LongTermGrowth', 'Beta', 'ResidualRisk',
                                'FinancialLeverage'], market='M75BNK-R', print_report=True, output_file=None):
    if file is None:
        file = os.path.join(util.default_output_location('reports', 'DEV'), 'tmp', 'barclay.csv')
    if not util.exists(file):
        display(f"No valid file to load Barclay Hedge Return series: file missing: \n{file}")
        return None
    calendar_str = 'GL'
    if end_date is None:
        end_date = util.load_month_ends(calendar_str, None, util.today())
        end_date = end_date[-1]
    days = util.load_business_days(calendar_str, start_date, end_date)
    mends = util.load_business_days(calendar_str, start_date, end_date, freq='MONTHEND')
    fg = root.load_object('COSMOS_US_FACTOR_GROUP')
    data = pd.read_csv(file, index_col=0, header=[0, 1, 2])
    data['Date'] = util.parse_date(data.index.to_numpy())
    ix = np.where(np.logical_and(data['Date'] >= days[0], data['Date'] <= days[-1]))[0]
    data = data.iloc[ix]
    for s in data.index:
        ix = np.where(mends <= data.loc[s, 'Date'].values[0])[0][-1]
        data.loc[s, 'Date'] = mends[ix]
    p_m = util.previous_day(data.loc[data.index[0], 'Date'].values[0], calendar_str, 'MONTHEND')
    s_date = util.next_business_days(p_m, calendar_str)
    e_date = data.iloc[-1, -1]
    data.set_index('Date', inplace=True)
    data = data.iloc[1:]
    num = int(len(data.columns) / 3)
    funds = np.array([None]*num)
    mnemonics = np.array([None]*num)
    for i in range(num):
        funds[i] = data.columns[i*3][0]
        mnemonics[i] = data.columns[i*3][1]
    ret = pd.DataFrame(np.nan, index=data.index, columns=funds)
    for i in range(num):
        ret.iloc[:, i] = data.iloc[:, i*3 + 2]
    display(f"Loading market returns between {s_date} and {e_date}")
    market_name = md.get_names(market)
    m = md.get_returns(s_date, e_date, market, calendar_str)
    m.rename(columns={m.columns[0]: market_name.loc[market, 'name']}, inplace=True)
    mkt = rt.daily_to_period(m, 'MONTHEND', calendar_str)
    display(f"Loading factor returns: {fg.name}: {s_date} - {e_date}")
    fr = fg.load_factor_returns_from_cache(s_date, e_date)
    fr = fr['values'][0]['values']
    fm = rt.daily_to_period(fr, 'MONTHEND', calendar_str)
    display(f"Time Series Exposure regressions")
    ind = fm[reg_factors]
    reg = fp.regression_ts(ret, ind, look_back=look_back)
    exposures = {}
    for ix, k in enumerate(mnemonics):
        df = pd.DataFrame(np.nan, index=ret.index, columns=ind.columns)
        df.iloc[look_back - 1:, :] = reg[ix]
        exposures[k] = df.copy()
    diag = fp.return_statistics(ret, None, mkt, calendar_str, freq='MONTHEND')
    result = {'funds': funds, 'tickers': mnemonics, 'return': ret, 'factor_returns': fr,
              'factor_monthly': fm, 'exposure_ts': exposures}
    if print_report:
        if output_file is None or not isinstance(output_file, str):
            output_file = os.path.join(util.default_output_location('reports', 'DEV'), 'tmp',
                                       f"BARCLAY_{s_date.strftime(util.yyyymmdd_format)}_"
                                       f"{e_date.strftime(util.yyyymmdd_format)}.xlsx")
        export_return_analysis(diag, output_file, True, calendar_str)
        reference = pd.DataFrame(mnemonics, index=funds, columns=['Short Name'])
        reference.index.name = 'Fund Name'
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
            reference.to_excel(writer, sheet_name='Reference')
        for k in exposures.keys():
            exist = util.exists(output_file)
            if exist:
                with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    exposures[k].to_excel(writer, sheet_name=k)
            else:
                with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
                    exposures[k].to_excel(writer, sheet_name=k)
        display(f"Generated {num} funds exposure to factors time series to {output_file}")
    return result


fill_pale_blue = styles.PatternFill("solid", start_color='BBFFFF')
fill_paler_blue = styles.PatternFill("solid", start_color='CCFFFF')
fill_pale_green = styles.PatternFill("solid", start_color='EEFFEE')
fill_pale_salmon = styles.PatternFill("solid", start_color='FFE4E1')
fill_yellow = styles.PatternFill("solid", start_color='FFFF66')
fill_green = styles.PatternFill("solid", start_color='99FF99')
font_blue = styles.Font(color=styles.colors.BLUE)
font_blue_bold = styles.Font(bold=True, color=styles.colors.BLUE)
font_red_bold = styles.Font(bold=True, color='CC0000')
border_bottom = styles.Border(bottom=styles.Side(color='000000', border_style='double'))
thin_border = styles.Border(left=styles.Side(style='thin'),
                            right=styles.Side(style='thin'),
                            top=styles.Side(style='thin'),
                            bottom=styles.Side(style='thin'))
