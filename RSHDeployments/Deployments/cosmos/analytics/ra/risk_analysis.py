#
# risk analysis
#
# Author: Yun Chen
# Copyright: Indigo Dao, LLC
# Date: 2022
#
import numpy as np
import openpyxl as pxl
import openpyxl.styles as styles
from openpyxl.styles import Border, Side, PatternFill, Font, colors
import pandas as pd

import factors.portfolio as pt
import util.utilities
from util.utilities import display
from classes.root import *
from scipy.special import ndtri
import warnings
from util.intersect import *
import dataloader.portfolio as port
import classes.root as root


def risk_func(por, B, fcov, dcov, mktpor=None, snapshots=None):

    # por = por['values'][0]
    result = {'factors': []}
    result['factor_types'] = []
    result['factor_themes'] = []
    result['models'] = []
    result['sec_ids'] = []
    result['weights'] = []
    result['volatility'] = []
    result['variance'] = []
    result['beta'] = []
    result['div_yield'] = []
    result['dispersion_measure'] = []
    result['acte_factor'] = []
    result['acte_residual'] = []
    result['rcte_factor'] = []
    result['rcte_residual'] = []
    result['acte_by_factor'] = []
    result['rcte_by_factor'] = []
    result['acte_by_security'] = []
    result['rcte_by_security'] = []
    result['beta_by_security'] = []
    result['div_yield_by_security'] = []
    result['volatility_by_security'] = []
    result['residual_volatility_by_security'] = []
    hrzns = [1, 5, 21, 63, 126, 252]
    result['value_at_risk'] = {'horizons': hrzns * 2}
    result['value_at_risk']['values'] = [np.nan] * (len(hrzns) * 2)
    result['value_at_risk']['probabilities'] = [0.01] * len(hrzns) + [0.05] * len(hrzns)
    del hrzns
    result['variance_by_security'] = []
    result['exposures'] = []
    result['security_rcte_by_factor'] = []
    result['security_rcte_by_residual'] = []

    if por is None:
        return result
    if por.empty:
        return result

    try:
        # factor covariances
        ia = np.where(~np.isnan(np.diag(fcov['values'])))[0]
        fcov['factors'] = fcov['factors'][ia]
        fcov['values'] = fcov['values'][np.ix_(ia, ia)]
        if 'name' in fcov:
            fcov['name'] = fcov['name'][ia]
        if 'models' in fcov:
            fcov['models'] = fcov['models'][ia]

        B, fcov = align_exposure_factor_covariance(B, fcov)
        result['factors'] = B.columns.to_numpy()
        if snapshots is not None:
            t_factors = snapshots['factors']
            types = snapshots['factor_types']
            themes = snapshots['factor_themes']
            result['factor_types'] = np.full_like(result['factors'], 'NA')
            result['factor_themes'] = np.full_like(result['factors'], 'NA')
            c, i1, i2=intersect(result['factors'], t_factors)
            result['factor_types'][i1] = types[i2]
            result['factor_themes'][i1] = themes[i2]
            del (t_factors, types, themes)

        result['models'] = fcov['models']
        factor_covariance = fcov['values']
        K = len(result['factors'])

        # exposure and residual
        good_index = np.where(np.sum(np.isnan(B), axis=1) == 0)[0]
        B = B.iloc[good_index, :]
        del good_index

        good_index = np.where(pd.notnull(np.diag(dcov)))[0]
        dcov = dcov.iloc[good_index, :]
        dcov = dcov.iloc[:, good_index]
        del good_index

        sec_ids = np.sort(np.intersect1d(dcov.index.to_numpy(),
                                         np.intersect1d(B.index.to_numpy(), por.columns.to_numpy())))
        N = len(sec_ids)
        result['sec_ids'] = sec_ids

        w = np.zeros((N, 1))
        b = np.zeros((N, K))
        d = np.zeros((N, N))

        c, ia, ib = intersect(sec_ids, por.columns.to_numpy())
        # if por['values'].ndim == 1:
        #     por['values'] = np.atleast_2d(por['values'])
        # if por['values'].shape[0] == 1:
        #     por['values'] = por['values'].T
        w[ia] = por[c].transpose()
        del (c, ia, ib)
        n_index = np.where(np.isnan(w))[0]
        if np.size(n_index) > 0:
            warnings.warn('%d securities having NaN for weights, set to zero' % len(n_index))
            w[n_index] = 0
        result['weights'] = w.flatten()

        c, ia, ib = intersect(sec_ids, B.index.to_numpy())
        b[ia,] = B.loc[c]
        del (c, ia, ib)
        wb = b.T @ w

        c, ia, ib = intersect(sec_ids, dcov.index)
        d[np.ix_(ia, ia)] = dcov.to_numpy()[np.ix_(ib, ib)]
        del (c, ia, ib)

        result['factor_exposure_by_security'] = b
        result['acte_by_security'] = b @ factor_covariance @ wb + d @ w
        result['variance'] = w.T @ result['acte_by_security']
        if result['variance'].ndim == 2:
            result['variance'] = result['variance'][0][0]
        result['volatility'] = np.sqrt(result['variance'])
        vol = result['volatility']

        result['acte_by_security'] = result['acte_by_security'] / vol
        result['rcte_by_security'] = w * result['acte_by_security'] / vol

        result['acte_by_factor'] = factor_covariance @ wb / vol
        result['rcte_by_factor'] = wb * result['acte_by_factor'] / vol

        result['acte_factor'] = wb.T @ factor_covariance @ wb / vol
        result['rcte_factor'] = result['acte_factor'] / vol
        result['acte_residual'] = w.T @ d @ w / vol
        result['rcte_residual'] = result['acte_residual'] / vol

        result['variance_by_security'] = np.diag(b @ factor_covariance @ b.T) + np.diag(d)
        result['volatility_by_security'] = np.sqrt(result['variance_by_security'])
        result['residual_volatility_by_security'] = np.sqrt(np.diag(d))
        result['dispersion_measure'] = (w.T @ result['volatility_by_security']) ** 2
        result['dispersion_measure'] = 1 - (result['volatility'] * result['volatility']) / result['dispersion_measure']

        result['exposures'] = wb

        result['security_rcte_by_factor'] = b * (factor_covariance @ b.T).T
        result['security_rcte_by_factor'] = result['security_rcte_by_factor'] / \
                                            np.tile(result['variance_by_security'], (len(result['factors']), 1)).T
        result['security_rcte_by_residual'] = np.diag(d) / result['variance_by_security']

        result['rcte_by_security_by_factor'] = np.tile(w, (1, b.shape[1])) * b * \
                                               np.tile((factor_covariance @ wb).T, (b.shape[0], 1))
        result['rcte_by_security_by_residual'] = w * (d * w)
        result['rcte_by_security_by_factor'] = result['rcte_by_security_by_factor'] / result['variance']
        result['rcte_by_security_by_residual'] = result['rcte_by_security_by_residual'] / result['variance']

        if mktpor is not None and isinstance(mktpor, pd.DataFrame):
            c, ia, ib = intersect(mktpor.columns.to_numpy(), B.index.to_numpy())
            mb = np.nansum(np.tile(mktpor[c].to_numpy().transpose(),
                                   (1, len(B.columns))) * B.loc[c], axis=0)
            del (c, ia, ib)

            c, ia, ib = intersect(mktpor.columns.to_numpy(), dcov.columns.to_numpy())
            m_d = mktpor[c].to_numpy() @ dcov.to_numpy()[np.ix_(ib, ib)] @ mktpor[c].T.to_numpy()
            del (c, ia, ib)

            market_variance = mb @ factor_covariance @ mb.T + m_d
            factor_betas = b @ factor_covariance @ mb.T

            delta_cov = np.zeros((len(sec_ids), len(mktpor.columns)))
            c, ia, ib = intersect(sec_ids, dcov.columns)
            c, ic, ie = intersect(mktpor.columns, dcov.columns)
            delta_cov[np.ix_(ia, ic)] = dcov.to_numpy()[np.ix_(ib, ie)]
            del (c, ia, ib, ic, ie)

            delta_betas = (delta_cov @ mktpor.to_numpy().transpose()).flatten()
            result['beta_by_security'] = (factor_betas + delta_betas) / market_variance
            result['beta'] = np.nansum(result['beta_by_security'] * result['weights'])
        else:
            result['beta_by_security'] = np.array([np.nan] * len(result['sec_ids']))
            result['beta'] = np.nan
        del (w, wb, b, d)

        for i in range(len(result['value_at_risk']['horizons'])):
            svar = ndtri(result['value_at_risk']['probabilities'][i])
            result['value_at_risk']['values'][i] = svar * result['volatility'] / \
                                                    np.sqrt(252 / result['value_at_risk']['horizons'][i])
            result['value_at_risk']['values'][i] = np.exp(result['value_at_risk']['values'][i]) - 1

    except ValueError:
        raise Exception(f'Unable to calculate risk on {por.index[0]}')
    return result


def align_exposure_factor_covariance(B, fcov, exclude_nans=False):
    if exclude_nans is None or not isinstance(exclude_nans, bool):
        exclude_nans = False
    if 'models' in B:
        unique_models = np.unique(B['models'])

        ia = np.array([], dtype=int)
        ib = np.array([], dtype=int)

        for i in range(len(unique_models)):
            bindex = np.where(B['models'] == unique_models[i])[0]
            findex = np.where(fcov['models'] == unique_models[i])[0]
            if np.size(bindex) == 0:
                continue
            if np.size(findex) == 0:
                continue
            cfactors, i1, i2 = intersect(B['factors'][bindex], fcov['factors'][findex])
            if np.size(cfactors) == 0:
                warnings.warn('%s: factor set from exposures and factor covariance do not agree' % unique_models[i])
                continue
            ia = np.concatenate((ia, bindex[i1]))
            ib = np.concatenate((ib, findex[i2]))

        B['factors'] = B['factors'][ia]
        B['factor_types'] = B['factor_types'][ia]
        B['factor_groups'] = B['factor_groups'][ia]
        if 'factor_themes' in B:
            B['factor_themes'] = B['factor_themes'][ia]
        else:
            B['factor_themes'] = np.array(['NA'] * len(B['factors']))
        B['class_names'] = B['class_names'][ia]
        B['models'] = B['models'][ia]
        B['values'] = B['values'][:, ia]

        fcov['factors'] = fcov['factors'][ib]
        fcov['models'] = fcov['models'][ib]
        fcov['values'] = fcov['values'][np.ix_(ib, ib)]

        if exclude_nans:
            good_index = np.where(~np.isnan(np.diag(fcov['values'])))[0]
            if len(good_index) < len(fcov['factors']):
                warnings.warn('%d factor covariances are NaN: Excluded' % (len(fcov['factors']) - len(good_index)))
            B['factors'] = B['factors'][good_index]
            B['factor_types'] = B['factor_types'][good_index]
            B['factor_groups'] = B['factor_groups'][good_index]
            if 'factor_themes' in B:
                B['factor_themes'] = B['factor_themes'][good_index]
            else:
                B['factor_themes'] = np.array(['NA'] * len(B['factors']))
            B['class_names'] = B['class_names'][good_index]
            B['models'] = B['models'][good_index]
            B['values'] = B['values'][:, good_index]

            fcov['factors'] = fcov['factors'][good_index]
            fcov['models'] = fcov['models'][good_index]
            fcov['values'] = fcov['values'][np.ix_(good_index, good_index)]
    else:
        c, ia, ib = intersect(B.columns.to_numpy(), fcov['factors'])
        B = B[c]
        fcov['factors'] = fcov['factors'][ib]
        fcov['values'] = fcov['values'][np.ix_(ib, ib)]
        if 'models' in fcov:
            fcov['models'] = fcov['models'][ib]
        del (c, ia, ib)

        if exclude_nans:
            good_index = np.where(~np.isnan(np.diag(fcov['values'])))[0]
            if len(good_index) < len(fcov['factors']):
                warnings.warn('%d factor covariances are NaN: Excluded' % (len(fcov['factors']) - len(good_index)))
            B = B.iloc[:, good_index]
            fcov['factors'] = fcov['factors'][good_index]
            if 'models' in fcov:
                fcov['models'] = fcov['models'][good_index]
            fcov['values'] = fcov['values'][np.ix_(good_index, good_index)]
    return B, fcov


def risk_analysis(bus_day, portfolio, benchmark, risk_model, calendar_str=None, print_report=None,
                  email_add=None, market=None, weighting_method=None,
                  benchmark_weighting_method=None, market_weighting_method=None, base_currency=None,
                  residual_fill_na=False, benchmark_day=None, industry_classification='COSMOS', prod=False):
    """
    risk analysis
    :param bus_day:
    :param portfolio:
    :param benchmark:
    :param risk_model:
    :param calendar_str:
    :param print_report:
    :param email_add:
    :param market:
    :param weighting_method:
    :param benchmark_weighting_method:
    :param market_weighting_method:
    :param base_currency:
    :param residual_fill_na:
    :param benchmark_day:
    :param industry_classification:
    :param prod:
    :return:

    Author    : Yun Chen
    Copyright : Indigo Dao, LLC
    Date      : August 1, 2022
    """
    result = {'dates': []}
    result['risk_model'] = []
    result['factor_group'] = []
    result['volatility'] = []
    result['tracking_error'] = []
    result['beta'] = []
    result['managed_portfolio'] = []
    result['managed_weighting_method'] = []
    result['benchmark_portfolio'] = []
    result['benchmark_weighting_method'] = []
    result['market_portfolio'] = []
    result['market_weighting_method'] = []
    result['managed_risk'] = []
    result['benchmark_risk'] = []
    result['active_risk'] = []
    result['market_risk'] = []
    result['master_sec_ids'] = []
    result['calendar'] = []
    result['classification'] = []
    result['active_share'] = []

    if benchmark is None:
        warnings.warn('No benchmark provided')
        benchmark = []

    if risk_model is None:
        raise Exception('No valid risk model provided')

    if not isinstance(residual_fill_na, bool):
        residual_fill_na = False

    result['managed_portfolio'] = portfolio
    result['benchmark_portfolio'] = benchmark

    risk_model = root.load_object(risk_model)
    result['risk_model'] = risk_model.name
    if hasattr(risk_model, 'factor_groups') and risk_model.factor_groups is not None:
        result['factor_group'] = risk_model.factor_groups
    else:
        result['factor_group'] = None
    if market is None or not isinstance(market, (str, int)):
        warnings.warn('No market provided: default of model used')
        market = risk_model.universe
        if market is None:
            warnings.warn('Model univers is not specified: using benchmark')
            market = benchmark
    result['market_portfolio'] = market

    if calendar_str is None or not isinstance(calendar_str, str):
        calendar_str = risk_model.calendar

    if industry_classification is not None and isinstance(industry_classification, str):
        result['classification'] = industry_classification
    bus_day = util.load_business_days(calendar_str, None, bus_day)
    bus_day = bus_day[-1]

    if print_report is None or not isinstance(print_report, bool):
        print_report = False
    if email_add is None or not isinstance(email_add, (str, list)):
        email_add = None
    if email_add is not None:
        print_report = True
    if weighting_method is None or not isinstance(weighting_method, str):
        if isinstance(portfolio, int):
            weighting_method = 'NAV'
        else:
            p_obj = root.load_object(portfolio)
            if p_obj is None:
                weighting_method = 'NAV'
            elif isinstance(p_obj, pt.Portfolio):
                weighting_method = p_obj.weighting_method
            else:
                weighting_method = 'NAV'
            del p_obj
    result['managed_weighting_method'] = weighting_method
    if benchmark_weighting_method is None or not isinstance(benchmark_weighting_method, str):
        if benchmark is None or isinstance(benchmark, int):
            benchmark_weighting_method = 'NAV'
        else:
            p_obj = root.load_object(benchmark)
            if p_obj is None:
                benchmark_weighting_method = 'NAV'
            elif isinstance(p_obj, pt.Portfolio):
                benchmark_weighting_method = p_obj.weighting_method
            else:
                benchmark_weighting_method = 'NAV'
            del p_obj
    result['benchmark_weighting_method'] = benchmark_weighting_method
    if market_weighting_method is None or not isinstance(market_weighting_method, str):
        if market is None or isinstance(market, int):
            market_weighting_method = 'NAV'
        else:
            p_obj = root.load_object(market)
            if p_obj is None:
                market_weighting_method = 'NAV'
            elif isinstance(p_obj, pt.Portfolio):
                market_weighting_method = p_obj.weighting_method
            else:
                market_weighting_method = 'NAV'
            del p_obj

    result['market_weighting_method'] = market_weighting_method
    if base_currency is None or not isinstance(base_currency, str):
        base_currency = risk_model.base_currency

    result['dates'] = bus_day
    result['base_currency'] = base_currency
    master_sec_ids = np.array([])
    cash = md.get_cash_securities()

    if result['factor_group'] is not None:
        fg = root.load_object(result['factor_group'])
        snapshots = fg.snapshot(bus_day, expand_flag=True)
    else:
        snapshots = None

    kickout = {'sec_ids': []}
    kickout['reason'] = []
    display(f"loading managed portfolio {portfolio} weights...")
    try:
        managed = port.get_portfolio_weights(bus_day, bus_day, portfolio, calendar_str=calendar_str,
                                             weight_flag=weighting_method, recurse=True, deep=True)
    except ValueError:
        raise Exception('Unable to load managed portfolio')
    master_sec_ids = np.union1d(master_sec_ids, managed.columns.to_numpy())

    breakdown_ls = False
    # if more than -1% shorts in non-cash assets, call it long-short
    long, short = port.split_long_short_portfolio(managed, exclude_cash=True)
    if not short.empty:
        sum_short = short.sum(axis=0).sum()
        if sum_short >= 0.01:
            breakdown_ls = True
        else:
            breakdown_ls = False

    is_split = False
    bench = []
    if not breakdown_ls:
        if benchmark is not None:
            if isinstance(benchmark, (int, str)):
                display(f"loading benchmark portfolio {benchmark} weights...")
                if benchmark in cash:
                    bench = pd.DataFrame(1.0, index=[bus_day], columns=[benchmark])
                else:
                    if benchmark_day is None:
                        b_day = bus_day
                    else:
                        b_day = util.most_recent_business_day(benchmark_day, calendar_str)
                    bench = port.get_portfolio_weights(b_day, b_day, benchmark,
                                                       weight_flag=benchmark_weighting_method,
                                                       calendar_str=calendar_str)
                    if benchmark_day is not None:
                        if bench is not None:
                            bench.index = managed.index
                master_sec_ids = np.union1d(master_sec_ids, bench.columns.to_numpy())
    else:
        managed = long
        bench = short
        is_split = True
    result['long short split'] = is_split

    mkt = []
    if market:
        try:
            if market == benchmark:
                display(f"using benchmark {benchmark} for market portfolio")
                mkt = bench
            else:
                display(f"loading market portfolio {market} weights...")
                mkt = port.get_portfolio_weights(bus_day, bus_day, market,
                                                 weight_flag=market_weighting_method,
                                                 calendar_str=calendar_str)
                master_sec_ids = np.union1d(master_sec_ids, mkt.columns.to_numpy())
        except ValueError:
            raise Exception('Unable to load market positions')

    if len(master_sec_ids) == 0:
        warnings.warn('No securities in the portfolio')
        return result

    result['master_sec_ids'] = master_sec_ids
    result['calendar'] = calendar_str
    active = port.active_portfolio(managed, bench)
    result['active_share'] = active.abs().sum(axis=1).sum()/2
    display(f"Active Share: {result['active_share']:.1%}")
    display(f"loading {risk_model.name} factor exposures..")
    B = risk_model.load_exposures(bus_day, master_sec_ids, None, calendar_str)
    bad_index = np.where(pd.isnull(B).sum(axis=1) > 0)[0]
    if np.size(bad_index) > 0:
        warnings.warn(f'{len(bad_index)} out of {len(master_sec_ids)} securities have NaN '
                      f'exposures; kicked out of analysis')
        kickout['sec_ids'] = np.concatenate((kickout['sec_ids'], B.index[bad_index]))
        kickout['reason'] = np.concatenate(kickout['reason'], ['missing valid exposures'] * len(bad_index))
    del bad_index
    display(f"loading {risk_model.name} factor covariance..")
    factor_cov = risk_model.load_factor_covariance(bus_day)
    display(f"loading {risk_model.name} residual covariance..")
    residual_cov = risk_model.load_residual_covariance(bus_day, master_sec_ids,
                                                       universe=None, calendar_str=calendar_str,
                                                       matrix_flag=True, fill_na=residual_fill_na)
    bad_index = np.where(pd.isnull(np.diag(residual_cov)))[0]
    if np.size(bad_index) > 0:
        warnings.warn(f'{len(bad_index)} out of {len(master_sec_ids)} securities missing '
                      f'residual var; kicked out of analysis')
        kickout['sec_ids'] = np.concatenate((kickout['sec_ids'], residual_cov.index[bad_index]))
        kickout['reason'] = np.concatenate((kickout['reason'], ['missing valid residuals'] * len(bad_index)))
        display(f"including: {kickout['sec_ids'][0]} ...")
    del bad_index
    result['kickout'] = kickout

    try:
        display(f"computing portfolio {portfolio} risk using {risk_model.name} on {bus_day}..")
        result['managed_risk'] = risk_func(managed, B, factor_cov, residual_cov, mkt, snapshots)
        result['volatility'] = result['managed_risk']['volatility']
    except ValueError:
        raise Exception('Unable to compute risk for managed portfolio')

    try:
        display(f"computing benchmark {benchmark} risk using {risk_model.name} on {bus_day}..")
        result['benchmark_risk'] = risk_func(bench, B, factor_cov, residual_cov, mkt, snapshots)
    except ValueError:
        raise Exception('Unable to compute risk for benchmark portfolio')

    try:
        display(f"computing active {portfolio} vs {benchmark} risk using {risk_model.name} on {bus_day}..")
        result['active_risk'] = risk_func(active, B, factor_cov, residual_cov, mkt, snapshots=snapshots)
        # result['active_risk']['cross_section'] = aggregate_rcte_by_group_vs_factor(result['dates'],
        #                                                                            result['active_risk'],
        #                                                                            risk_model.name)
        # skip aggregate by country vs factor/sector
    except ValueError:
        raise Exception('Unable to compute risk for active portfolio')

    if not market:
        result['market_risk'] = []
    else:
        try:
            if market == benchmark:
                display(f"using benchmark {benchmark} for market risk with {risk_model.name} on {bus_day}..")
                result['market_risk'] = result['benchmark_risk']
            else:
                display(f"computing market {market} risk using {risk_model.name} on {bus_day}..")
                result['market_risk'] = risk_func(mkt, B, factor_cov, residual_cov, mkt, snapshots=snapshots)
        except ValueError:
            raise Exception('Unable to compute risk for market portfolio')

    result['tracking_error'] = result['active_risk']['volatility']

    display(f"Tracking error {bus_day} {portfolio} vs {benchmark}: "
            f"predicted {result['tracking_error']*100:.1f} %")

    # if print_report:
    if print_report:
        if prod:
            env = 'PROD'
        else:
            env = 'DEV'
        output_location = os.path.join(util.default_output_location('reports', env), 'ra')
        if not os.path.exists(output_location):
            os.makedirs(output_location)
        file = os.path.join(output_location, f"{portfolio}.{benchmark}."
                                             f"risk.analysis.{risk_model.name}."
                                             f"{bus_day.strftime(util.yyyymmdd_format)}.xlsx")
        util.preserve_file(file)
        export_risk_report(result, file)
        display(f"{portfolio} vs. {benchmark} {risk_model.name} on {bus_day} Risk Report Saved to {file}")
        if email_add is not None:
            display(f"to send report by email")
    return result


def aggregate_rcte_by_group_vs_factor(bus_day, risk, group_by_themes):
    style_factors = np.where((risk['factor_types'] == 'RETURN') | (risk['factor_types'] == 'CONTROL'))[0]
    result = {'factors': risk['factors'][style_factors]}
    result['group_types'] = ['INDUSTRY']
    result['groups'] = [None]
    result['groups'][0] = {'group_factors': np.unique(risk['factors'][risk['factor_types'] == 'INDUSTRY'])}
    N = len(result['groups'][0]['group_factors'])
    result['groups'][0]['group_weights'] = np.array([np.nan] * N)
    result['groups'][0]['rcte'] = np.full((N, len(result['factors'])), np.nan)
    result['groups'][0]['weighted_exposure'] = np.full((N, len(result['factors'])), np.nan)

    B = risk['factor_exposure_by_security'][:, style_factors]
    rcte = risk['rcte_by_security_by_factor'][:, style_factors]

    for i in range(N):
        sec_i = (risk['factors'] == result['groups'][0]['group_factors'][i]) & (risk['factor_types'] == 'INDUSTRY')
        w = np.sum(risk['factor_exposure_by_security'][:, sec_i], axis=1)
        result['groups'][0]['group_weights'] = np.nansum( w * risk['weights'])
        S = len(style_factors)
        result['groups'][0]['rcte'][i,] = np.nansum(np.tile(w, (S, 1)).T * rcte)
        wb = np.tile(w, (S, 1)).T * np.tile(risk['weights'], (S, 1)).T * B
        result['groups'][0]['weighted_exposure'][i,] = np.nansum(wb, axis=0)

    # return if no factors are consolidated into themes
    if group_by_themes is None or not isinstance(group_by_themes, str):
        return result

    # consolidate output factors into themes
    themes = get_factor_themes(group_by_themes)
    T = len(themes['names'])
    for i in range(len(result['groups'])):
        new_rcte = np.zeros((N, T))
        new_weighted_exposure = np.zeros((N, T))
        for j in range(T):
            tf, _ = ismember(result['factors'], themes['values'][j]['factors'])
            new_rcte[:, j] = np.nansum(result['groups'][i]['rcte'][:, tf], axis=1)
            new_weighted_exposure[:, j] = np.nansum(result['groups'][i]['weighted_exposure'][:, tf], axis=1)
        result['groups'][i]['rcte'] = new_rcte
        result['groups'][i]['weighted_exposure'] = new_weighted_exposure
    result['factors'] = themes['names']
    return result


def export_risk_report(result, file):
    wb = pxl.Workbook()
    summary = wb.active
    summary.title = 'Summary'
    detail = wb.create_sheet('Detail', 1)
    classification = result['classification']
    class_map = md.get_classification_meta_map(classification)
    fg = root.load_object(result['factor_group'])
    # summary
    display(f"exporting risk analysis 'summary'")
    # managed_risk = result['managed_risk']
    # benchmark_risk = result['benchmark_risk']
    # active_risk = result['active_risk']
    # market_risk = result['market_risk']
    por_types = ['Portfolio', 'Benchmark', 'Active', 'Market']
    portfolios = [result['managed_portfolio'], result['benchmark_portfolio'], '-', result['market_portfolio']]
    if result['long short split']:
        portfolios = ['long', 'short', '-', result['market_portfolio']]
    sec_type = None
    if isinstance(result['managed_portfolio'], str):
        man_por_obj = root.load_object(result['managed_portfolio'])
        if hasattr(man_por_obj, 'security_type'):
            sec_type = man_por_obj.security_type
    else:
        m_ref = md.get_portfolio_references(result['managed_portfolio'])
        portfolios[0] = m_ref['ReferenceCode'].iloc[0]
    if not isinstance(result['benchmark_portfolio'], str):
        b_ref = md.get_portfolio_references(result['benchmark_portfolio'])
        portfolios[1] = b_ref['ReferenceCode'].iloc[0]
    if not isinstance(result['market_portfolio'], str):
        k_ref = md.get_portfolio_references(result['market_portfolio'])
        portfolios[3] = k_ref['ReferenceCode'].iloc[0]
    risks = ['managed_risk', 'benchmark_risk', 'active_risk', 'market_risk']
    wt_flags = [result['managed_weighting_method'], result['benchmark_weighting_method'], '-',
                result['market_weighting_method']]
    row = 1
    col = 1
    summary.cell(row, col).value = 'Risk Summary'
    summary.cell(row, col).font = styles.Font(bold=True, color=colors.BLUE)
    row = row + 1
    summary.cell(row, col).value = 'Date'
    summary.cell(row, col+1).value = result['dates'].strftime(util.YY_MM_DD_format)
    summary.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    summary.cell(row, col).value = 'Risk Model'
    summary.cell(row, col+1).value = result['risk_model']
    summary.cell(row, col+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    summary.cell(row, col).value = 'Managed'
    summary.cell(row, col + 1).value = portfolios[0]
    summary.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    summary.cell(row, col).value = 'Benchmark'
    summary.cell(row, col + 1).value = portfolios[1]
    summary.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    summary.cell(row, col).value = 'Market'
    summary.cell(row, col + 1).value = portfolios[3]
    summary.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    summary.cell(row, col).value = 'Long Short Split'
    if result['long short split']:
        summary.cell(row, col + 1).value = 'True'
    else:
        summary.cell(row, col + 1).value = '---'
    summary.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')

    row = row + 2
    summary.cell(row, col).value = 'Risk Summary Statistics'
    summary.cell(row, col).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col).font = styles.Font(bold=True, color=colors.BLUE)
    summary.cell(row, col).border = thin_border
    summary.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+4)
    row = row + 1
    for idx, p in enumerate(por_types):
        summary.cell(row, col+idx+1).value = p
        summary.cell(row, col+idx+1).alignment = styles.Alignment(horizontal='left')
    row = row + 1
    summary.cell(row, col).value = 'ID'
    for idx, p in enumerate(portfolios):
        summary.cell(row, col+idx+1).value = p
        summary.cell(row, col+idx+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    summary.cell(row, col).value = 'Weighting'
    for idx, w in enumerate(wt_flags):
        summary.cell(row, col+idx+1).value = w
        summary.cell(row, col+idx+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    summary.cell(row, col).value = 'Vol / TE'
    for idx, p in enumerate(risks):
        summary.cell(row, col+idx+1).value = result[p]['volatility']
        summary.cell(row, col+idx+1).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col+idx+1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        if p == 'active_risk' and result[p]['volatility'] > 0.05:
            summary.cell(row, col + idx + 1).font = font_red_bold
            summary.cell(row, col + idx + 1).fill = fill_yellow
    row = row + 1
    summary.cell(row, col).value = 'Beta'
    for idx, p in enumerate(risks):
        summary.cell(row, col+idx+1).value = result[p]['beta']
        summary.cell(row, col+idx+1).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col+idx+1).number_format = styles.numbers.FORMAT_NUMBER_00
    row = row + 1
    summary.cell(row, col).value = 'Dispersion'
    for idx, p in enumerate(risks):
        if p.lower().strip() == 'active_risk':
            summary.cell(row, col + idx + 1).value = '-'
        else:
            summary.cell(row, col+idx+1).value = result[p]['dispersion_measure'][0]
            summary.cell(row, col+idx+1).number_format = styles.numbers.FORMAT_PERCENTAGE
        summary.cell(row, col+idx+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    summary.cell(row, col).value = 'Active Share'
    summary.cell(row, col+3).value = result['active_share']
    summary.cell(row, col+3).number_format = styles.numbers.FORMAT_PERCENTAGE

    row = row + 2
    summary.cell(row, col).value = 'Relative Contribution To Tracking Errors (RCTE)'
    summary.cell(row, col).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col).font = styles.Font(bold=True, color=colors.BLUE)
    summary.cell(row, col).border = thin_border
    summary.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+4)
    row = row + 1
    for idx, p in enumerate(por_types):
        summary.cell(row, col+idx+1).value = p
        summary.cell(row, col+idx+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    summary.cell(row, col).value = 'ID'
    for idx, p in enumerate(portfolios):
        summary.cell(row, col+idx+1).value = p
        summary.cell(row, col+idx+1).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    summary.cell(row, col).value = 'Factor'
    for idx, p in enumerate(risks):
        summary.cell(row, col+idx+1).value = result[p]['rcte_factor'][0][0]
        summary.cell(row, col+idx+1).number_format = styles.numbers.FORMAT_PERCENTAGE
        summary.cell(row, col+idx+1).alignment = styles.Alignment(horizontal='center')
    for style_type in ['RETURN', 'CONTROL', 'INDUSTRY', 'COUNTRY', 'CURRENCY']:
        row = row + 1
        summary.cell(row, col).value = style_type.capitalize()
        summary.cell(row, col).alignment = styles.Alignment(indent=1)
        for idx, p in enumerate(risks):
            sx = np.where(result[p]['factor_types'] == style_type)[0]
            s_rcte = result[p]['rcte_by_factor'][sx].sum()
            # summary.cell(row, col+idx+1).value = result[p]['rcte_by_factor'][result[p]['factor_types']
            #                                                                  == style_type].sum()
            summary.cell(row, col+idx+1).value = s_rcte
            summary.cell(row, col+idx+1).number_format = styles.numbers.FORMAT_PERCENTAGE
            summary.cell(row, col+idx+1).alignment = styles.Alignment(horizontal='center')
            s_flag = False
            if p == 'active_risk':
                if style_type == 'RETURN' and s_rcte > 0.20:
                    summary.cell(row, col + idx + 1).font = font_blue_bold
                    summary.cell(row, col + idx + 1).fill = fill_green
                if style_type in ('INDUSTRY', 'COUNTRY', 'CURRENCY', 'CONTROL') and s_rcte > 0.20:
                    summary.cell(row, col + idx + 1).font = font_red_bold
                    summary.cell(row, col + idx + 1).fill = fill_yellow
    row = row + 1
    summary.cell(row, col).value = 'Residual'
    for idx, p in enumerate(risks):
        s_rcte = result[p]['rcte_residual'][0][0]
        summary.cell(row, col+idx+1).value = s_rcte
        summary.cell(row, col+idx+1).number_format = styles.numbers.FORMAT_PERCENTAGE
        summary.cell(row, col+idx+1).alignment = styles.Alignment(horizontal='center')
        if s_rcte > 0.4:
            summary.cell(row, col + idx + 1).font = font_red_bold
            summary.cell(row, col + idx + 1).fill = fill_yellow
    row = row + 1
    summary.cell(row, col).value = 'Total'
    for idx, p in enumerate(risks):
        summary.cell(row, col+idx+1).value = 1
        summary.cell(row, col+idx+1).number_format = styles.numbers.FORMAT_PERCENTAGE
        summary.cell(row, col+idx+1).alignment = styles.Alignment(horizontal='center')

    # --------------------------------------------------
    # sector
    # RCTE by sector
    row = row + 3
    summary.cell(row, col).value = 'RCTE by Sector'
    summary.cell(row, col).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col).font = styles.Font(bold=True, color=colors.BLUE)
    summary.cell(row, col).border = thin_border
    summary.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+8)
    row = row + 1
    for idx, por in enumerate(por_types):
        summary.cell(row, col+2*idx+1).value = por
        summary.cell(row, col+2*idx+1).alignment = styles.Alignment(horizontal='center')
        summary.merge_cells(start_row=row, start_column=col+2*idx+1, end_row=row, end_column=col+2*idx+2)
    row = row + 1
    for idx, por in enumerate(portfolios):
        summary.cell(row, col+2*idx+1).value = por
        summary.cell(row, col+2*idx+1).alignment = styles.Alignment(horizontal='center')
        summary.merge_cells(start_row=row, start_column=col+2*idx+1, end_row=row, end_column=col+2*idx+2)

    row = row + 1
    summary.cell(row, col).value = 'Sectors'
    for idx, por in enumerate(por_types):
        summary.cell(row, col+2*idx+1).value = 'Weight'
        summary.cell(row, col+2*idx+1).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col+2*idx+2).value = 'RCTE'
        summary.cell(row, col+2*idx+2).alignment = styles.Alignment(horizontal='center')

    unique_sectors = np.unique(class_map['sector'])
    sec_ref = md.get_classification(result['active_risk']['sec_ids'], 'sector', vector_flag=True)
    for s in unique_sectors:
        row = row + 1
        summary.cell(row, col).value = s
        summary.cell(row, col).alignment = styles.Alignment(indent=1)
        ix = np.where(sec_ref['values'] == s)[0]
        if len(ix) == 0:
            continue
        sids = sec_ref.index[ix].to_numpy()
        for pdx, p in enumerate(risks):
            kix = np.where(np.isin(result[p]['sec_ids'], sids))[0]
            if len(kix) == 0:
                continue
            summary.cell(row, col + 2 * pdx + 1).value = result[p]['weights'][kix].sum()
            summary.cell(row, col + 2 * pdx + 1).number_format = styles.numbers.FORMAT_PERCENTAGE
            s_rcte = result[p]['rcte_by_security'][kix].sum()
            s_flag = False
            if s_rcte > 0.3:
                s_flag = True
            summary.cell(row, col + 2 * pdx + 2).value = s_rcte
            summary.cell(row, col + 2 * pdx + 2).number_format = styles.numbers.FORMAT_PERCENTAGE
            if s_flag:
                summary.cell(row, col + 2 * pdx + 2).font = font_red_bold
                summary.cell(row, col + 2 * pdx + 2).fill = fill_yellow
    # --------------------------------------------------
    # industry group
    # RCTE by indgrp
    row = row + 3
    summary.cell(row, col).value = 'RCTE by Industry Group'
    summary.cell(row, col).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col).font = styles.Font(bold=True, color=colors.BLUE)
    summary.cell(row, col).border = thin_border
    summary.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+8)
    row = row + 1
    for idx, por in enumerate(por_types):
        summary.cell(row, col+2*idx+1).value = por
        summary.cell(row, col+2*idx+1).alignment = styles.Alignment(horizontal='center')
        summary.merge_cells(start_row=row, start_column=col+2*idx+1, end_row=row, end_column=col+2*idx+2)
    row = row + 1
    for idx, por in enumerate(portfolios):
        summary.cell(row, col+2*idx+1).value = por
        summary.cell(row, col+2*idx+1).alignment = styles.Alignment(horizontal='center')
        summary.merge_cells(start_row=row, start_column=col+2*idx+1, end_row=row, end_column=col+2*idx+2)

    row = row + 1
    summary.cell(row, col).value = 'Industry Group'
    for idx, por in enumerate(por_types):
        summary.cell(row, col+2*idx+1).value = 'Weight'
        summary.cell(row, col+2*idx+1).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col+2*idx+2).value = 'RCTE'
        summary.cell(row, col+2*idx+2).alignment = styles.Alignment(horizontal='center')

    unique_sectors = np.unique(class_map['indgrp'])
    sec_ref = md.get_classification(result['active_risk']['sec_ids'], 'indgrp', vector_flag=True)
    for s in unique_sectors:
        row = row + 1
        summary.cell(row, col).value = s
        summary.cell(row, col).alignment = styles.Alignment(indent=1)
        ix = np.where(sec_ref['values'] == s)[0]
        if len(ix) == 0:
            continue
        sids = sec_ref.index[ix].to_numpy()
        for pdx, p in enumerate(risks):
            kix = np.where(np.isin(result[p]['sec_ids'], sids))[0]
            if len(kix) == 0:
                continue
            summary.cell(row, col + 2 * pdx + 1).value = result[p]['weights'][kix].sum()
            summary.cell(row, col + 2 * pdx + 1).number_format = styles.numbers.FORMAT_PERCENTAGE
            s_rcte = result[p]['rcte_by_security'][kix].sum()
            s_flag = False
            if s_rcte > 0.3:
                s_flag = True
            summary.cell(row, col + 2 * pdx + 2).value = s_rcte
            summary.cell(row, col + 2 * pdx + 2).number_format = styles.numbers.FORMAT_PERCENTAGE
            if s_flag:
                summary.cell(row, col + 2 * pdx + 2).font = font_red_bold
                summary.cell(row, col + 2 * pdx + 2).fill = fill_yellow

    # --------------------------------------------------
    # return themes
    # RCTE by themes
    row = row + 3
    summary.cell(row, col).value = 'RCTE by Styles'
    summary.cell(row, col).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col).font = styles.Font(bold=True, color=colors.BLUE)
    summary.cell(row, col).border = thin_border
    summary.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+8)
    row = row + 1
    for idx, por in enumerate(por_types):
        summary.cell(row, col+2*idx+1).value = por
        summary.cell(row, col+2*idx+1).alignment = styles.Alignment(horizontal='center')
        summary.merge_cells(start_row=row, start_column=col+2*idx+1, end_row=row, end_column=col+2*idx+2)
    row = row + 1
    for idx, por in enumerate(portfolios):
        summary.cell(row, col+2*idx+1).value = por
        summary.cell(row, col+2*idx+1).alignment = styles.Alignment(horizontal='center')
        summary.merge_cells(start_row=row, start_column=col+2*idx+1, end_row=row, end_column=col+2*idx+2)

    row = row + 1
    summary.cell(row, col).value = 'Themes'
    for idx, por in enumerate(por_types):
        summary.cell(row, col+2*idx+1).value = 'Weight'
        summary.cell(row, col+2*idx+1).alignment = styles.Alignment(horizontal='center')
        summary.cell(row, col+2*idx+2).value = 'RCTE'
        summary.cell(row, col+2*idx+2).alignment = styles.Alignment(horizontal='center')

    row = row + 1
    r_themes = get_factor_themes(fg, 'RETURN')
    summary.cell(row, col).value = 'RETURN (Alpha) Themes'
    summary.cell(row, col).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col).font = styles.Font(bold=True, color='009900')
    summary.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+8)
    for theme in r_themes:
        tf = r_themes[theme]
        row = row + 1
        summary.cell(row, col).value = theme
        summary.cell(row, col).alignment = styles.Alignment(indent=1)
        for pdx, p in enumerate(risks):
            tx = np.where(np.isin(result[p]['factors'], tf))[0]
            s_exp = np.nansum(result[p]['exposures'][tx])
            s_rcte = np.nansum(result[p]['rcte_by_factor'][tx])
            summary.cell(row, col + 2 * pdx + 1).value = s_exp
            summary.cell(row, col + 2 * pdx + 1).number_format = styles.numbers.FORMAT_NUMBER_00
            s_flag = False
            if s_rcte > 0.05:
                s_flag = True
            summary.cell(row, col + 2 * pdx + 2).value = s_rcte
            summary.cell(row, col + 2 * pdx + 2).number_format = styles.numbers.FORMAT_PERCENTAGE
            if s_flag:
                summary.cell(row, col + 2 * pdx + 2).font = font_blue_bold
                summary.cell(row, col + 2 * pdx + 2).fill = fill_green
            if s_rcte < -0.05:
                summary.cell(row, col + 2 * pdx + 2).font = font_red_bold
                summary.cell(row, col + 2 * pdx + 2).fill = fill_yellow
        s_row = row
        for ff in tf:
            row = row + 1
            summary.cell(row, col).value = ff
            summary.cell(row, col).alignment = styles.Alignment(indent=2)
            for pdx, p in enumerate(risks):
                tx = np.where(result[p]['factors'] == ff)[0]
                s_exp = np.nansum(result[p]['exposures'][tx])
                s_rcte = np.nansum(result[p]['rcte_by_factor'][tx])
                summary.cell(row, col + 2 * pdx + 1).value = s_exp
                summary.cell(row, col + 2 * pdx + 1).number_format = styles.numbers.FORMAT_NUMBER_00
                s_flag = False
                if s_rcte > 0.05:
                    s_flag = True
                summary.cell(row, col + 2 * pdx + 2).value = s_rcte
                summary.cell(row, col + 2 * pdx + 2).number_format = styles.numbers.FORMAT_PERCENTAGE
                if s_flag:
                    summary.cell(row, col + 2 * pdx + 2).font = font_blue_bold
                    summary.cell(row, col + 2 * pdx + 2).fill = fill_green
                if s_rcte < -0.05:
                    summary.cell(row, col + 2 * pdx + 2).font = font_red_bold
                    summary.cell(row, col + 2 * pdx + 2).fill = fill_yellow
        summary.row_dimensions.group(s_row+1, row, hidden=True)

    row = row + 1
    r_themes = get_factor_themes(fg, 'CONTROL')
    summary.cell(row, col).value = 'CONTROL (Risk) Themes'
    summary.cell(row, col).alignment = styles.Alignment(horizontal='center')
    summary.cell(row, col).font = styles.Font(bold=True, color='CC3333')
    summary.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+8)
    for theme in r_themes:
        tf = r_themes[theme]
        row = row + 1
        summary.cell(row, col).value = theme
        summary.cell(row, col).alignment = styles.Alignment(indent=1)
        for pdx, p in enumerate(risks):
            tx = np.where(np.isin(result[p]['factors'], tf))[0]
            s_exp = np.nansum(result[p]['exposures'][tx])
            s_rcte = np.nansum(result[p]['rcte_by_factor'][tx])
            summary.cell(row, col + 2 * pdx + 1).value = s_exp
            summary.cell(row, col + 2 * pdx + 1).number_format = styles.numbers.FORMAT_NUMBER_00
            s_flag = False
            if s_rcte > 0.05:
                s_flag = True
            summary.cell(row, col + 2 * pdx + 2).value = s_rcte
            summary.cell(row, col + 2 * pdx + 2).number_format = styles.numbers.FORMAT_PERCENTAGE
            if s_flag:
                summary.cell(row, col + 2 * pdx + 2).font = font_red_bold
                summary.cell(row, col + 2 * pdx + 2).fill = fill_yellow
            if s_rcte < -0.05:
                summary.cell(row, col + 2 * pdx + 2).font = font_blue_bold
                summary.cell(row, col + 2 * pdx + 2).fill = fill_green
        s_row = row
        for ff in tf:
            row = row + 1
            summary.cell(row, col).value = ff
            summary.cell(row, col).alignment = styles.Alignment(indent=2)
            for pdx, p in enumerate(risks):
                tx = np.where(result[p]['factors'] == ff)[0]
                s_exp = np.nansum(result[p]['exposures'][tx])
                s_rcte = np.nansum(result[p]['rcte_by_factor'][tx])
                summary.cell(row, col + 2 * pdx + 1).value = s_exp
                summary.cell(row, col + 2 * pdx + 1).number_format = styles.numbers.FORMAT_NUMBER_00
                s_flag = False
                if s_rcte > 0.05:
                    s_flag = True
                summary.cell(row, col + 2 * pdx + 2).value = s_rcte
                summary.cell(row, col + 2 * pdx + 2).number_format = styles.numbers.FORMAT_PERCENTAGE
                if s_flag:
                    summary.cell(row, col + 2 * pdx + 2).font = font_red_bold
                    summary.cell(row, col + 2 * pdx + 2).fill = fill_yellow
                if s_rcte < -0.05:
                    summary.cell(row, col + 2 * pdx + 2).font = font_blue_bold
                    summary.cell(row, col + 2 * pdx + 2).fill = fill_green
            summary.row_dimensions[row].hidden = True
        summary.row_dimensions.group(s_row + 1, row, hidden=True)
    # # --------------------------------------------------
    # # industry group
    # row = row + 3
    # summary.cell(row, col).value = 'Industry Group'
    # for idx, por in enumerate(por_types):
    #     summary.cell(row, col+2*idx+1).value = 'Weight'
    #     summary.cell(row, col+2*idx+1).alignment = styles.Alignment(horizontal='center')
    #     summary.cell(row, col+2*idx+2).value = 'RCTE'
    #     summary.cell(row, col+2*idx+2).alignment = styles.Alignment(horizontal='center')
    # index = np.where(result['active_risk']['factor_types'] == 'INDUSTRY')[0]
    # if len(index) == 0:
    #     display(f"!! No industry factors found !!")
    # else:
    #     industries = result['active_risk']['factors'][index]
    #     for idx, ind in enumerate(industries):
    #         row = row + 1
    #         summary.cell(row, col).value = ind
    #         summary.cell(row, col).alignment = styles.Alignment(indent=1)
    #         gdx = index[idx]
    #         for pdx, p in enumerate(risks):
    #             kix = np.where(result[p]['factor_exposure_by_security'][:, gdx] != 0)[0]
    #             if len(kix) == 0:
    #                 continue
    #             summary.cell(row, col+2*pdx+1).value = result[p]['weights'][kix].sum()
    #             summary.cell(row, col+2*pdx+1).number_format = styles.numbers.FORMAT_PERCENTAGE
    #             summary.cell(row, col+2*pdx+2).value = result[p]['rcte_by_security'][kix].sum()
    #             summary.cell(row, col+2*pdx+2).number_format = styles.numbers.FORMAT_PERCENTAGE

    # -------------------------------------------
    # detail
    # -------------------------------------------
    # ===========================================================================
    #
    # Details
    #
    # ===========================================================================
    # ===========================================================
    #
    # Details
    #
    # ===========================================================
    display(f"exporting risk analysis 'detail'")
    row = 1
    col = 1

    columns = ['ID', 'Ticker', 'Cusip', 'Sedol', 'Issuer', 'Domicile', 'Exchange', 'Currency', 'Sector', 'Industry',
               'Beta', 'Volatility', 'Residual', 'Weight', 'RCTE', 'Weight', 'RCTE', 'Weight', 'RCTE']
    col = col + len(columns) - 6
    detail.cell(row, col).value = portfolios[0]   # result['managed_portfolio']
    detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
    detail.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    col = col + 2
    detail.cell(row, col).value = portfolios[1]   # result['benchmark_portfolio']
    detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
    detail.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    col = col + 2
    detail.cell(row, col).value = 'Active'
    detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
    detail.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    col = col + 2
    detail.cell(row, col).value = 'Exposures'
    detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
    row = row + 1
    col = 1
    for p in columns:
        detail.cell(row, col).value = p
        detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
        col = col + 1
    act = result['active_risk']
    man = result['managed_risk']
    for fi, f in enumerate(act['factors']):
        if act['factor_types'][fi] in ['INDUSTRY']:
            continue
        detail.cell(row, col).value = f
        detail.cell(row, col).number_format = styles.numbers.FORMAT_NUMBER_00
        detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
        col = col + 1
    col = 1
    ref = md.get_references(act['sec_ids'], None, False, result['dates'])
    tickers = md.get_tickers(act['sec_ids'], None, True)
    cusips = md.get_cusips(act['sec_ids'], None, True)
    sedols = md.get_sedols(act['sec_ids'], None, True)
    sector = md.get_classification(act['sec_ids'], 'sector', source=classification, vector_flag=True)
    industry = md.get_classification(act['sec_ids'], 'industry', source=classification, vector_flag=True)
    for idx, sec in enumerate(act['sec_ids']):
        row = row + 1
        col = 1
        detail.cell(row, col).value = sec
        s_ref = md.get_references(sec, None, False, result['dates'])
        if s_ref.empty:
            display(f"{util.caller()}: {util.current_time()}: Unable to find reference for {sec}")
            continue
        t_ref = md.get_tickers(sec, result['dates'], True)
        c_ref = md.get_cusips(sec, result['dates'], True)
        d_ref = md.get_sedols(sec, result['dates'], True)
        if len(s_ref.index) > 0:
            # list active ones
            ai = np.where(np.logical_and(s_ref['is_active'] == 1, pd.notnull(s_ref['name']),
                                         pd.notnull(s_ref['entity_id'])))[0]
            if len(ai) == 0:
                sx = 0
            else:
                sx = ai[0]
        else:
            sx = 0
        if c_ref is not None:
            cus = c_ref.loc[c_ref.index[0], 'cusip']
        else:
            cus = ''
        if t_ref is not None:
            tik = t_ref.loc[t_ref.index[0], 'ticker_region']
        else:
            tik = ''
        if d_ref is not None:
            sed = d_ref.loc[d_ref.index[0], 'sedol']
        else:
            sed = ''
        col = col + 1
        # detail.cell(row, col).value = s_ref['ticker'].iloc[sx]
        detail.cell(row, col).value = tik
        col = col + 1
        # detail.cell(row, col).value = s_ref['cusip'].iloc[sx]
        detail.cell(row, col).value = cus
        col = col + 1
        # detail.cell(row, col).value = s_ref['sedol'].iloc[sx]
        detail.cell(row, col).value = sed
        col = col + 1
        detail.cell(row, col).value = s_ref['name'].iloc[sx]
        col = col + 1
        detail.cell(row, col).value = s_ref['domicile'].iloc[sx]
        col = col + 1
        detail.cell(row, col).value = s_ref['exchange'].iloc[sx]
        col = col + 1
        detail.cell(row, col).value = s_ref['currency'].iloc[sx]
        col = col + 1
        detail.cell(row, col).value = sector.loc[sec, 'values']
        col = col + 1
        detail.cell(row, col).value = industry.loc[sec, 'values']
        col = col + 1
        detail.cell(row, col).value = act['beta_by_security'][0][idx]
        detail.cell(row, col).number_format = styles.numbers.FORMAT_NUMBER_00
        col = col + 1
        detail.cell(row, col).value = act['volatility_by_security'][idx]
        detail.cell(row, col).number_format = styles.numbers.FORMAT_PERCENTAGE
        detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
        col = col + 1
        detail.cell(row, col).value = act['residual_volatility_by_security'][idx]
        detail.cell(row, col).number_format = styles.numbers.FORMAT_PERCENTAGE
        detail.cell(row, col).alignment = styles.Alignment(horizontal='center')

        for p in risks:
            if p.lower().strip() == 'market_risk':
                continue
            m_idx = np.where(result[p]['sec_ids'] == sec)[0]
            col = col + 1
            if len(m_idx) == 0:
                detail.cell(row, col).value = 0
                detail.cell(row, col).number_format = styles.numbers.FORMAT_PERCENTAGE
                detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
            else:
                detail.cell(row, col).value = result[p]['weights'][m_idx][0]
                detail.cell(row, col).number_format = styles.numbers.FORMAT_PERCENTAGE_00
                detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
            col = col + 1
            if len(m_idx) == 0:
                detail.cell(row, col).value = 0
                detail.cell(row, col).number_format = styles.numbers.FORMAT_PERCENTAGE
                detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
            else:
                detail.cell(row, col).value = result[p]['rcte_by_security'][m_idx][0][0]
                detail.cell(row, col).number_format = styles.numbers.FORMAT_PERCENTAGE_00
                detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
                if detail.cell(row, col).value > 0.1:
                    detail.cell(row, col).font = font_red_bold
                    detail.cell(row, col).fill = fill_yellow

        for fi, f in enumerate(act['factors']):
            if act['factor_types'][fi] in ['INDUSTRY']:
                continue
            col = col + 1
            detail.cell(row, col).value = act['factor_exposure_by_security'][idx, fi]
            detail.cell(row, col).number_format = styles.numbers.FORMAT_NUMBER_00
            detail.cell(row, col).alignment = styles.Alignment(horizontal='center')
    wb.save(file)

    display(f"Successfully export risk to {file}")


def get_factor_themes(fg, factor_type='RETURN'):
    f = root.load_object(fg)
    s = f.snapshot(expand_flag=True)
    ix = np.where(s['factor_types'] == 'RETURN')[0]
    r_themes = np.unique(s['factor_themes'][ix])
    rd = dict.fromkeys(r_themes)
    for r in r_themes:
        rx = np.where(np.logical_and(s['factor_themes'] == r, s['factor_types'] == 'RETURN'))[0]
        rd[r] = s['factors'][rx]
    ix = np.where(s['factor_types'] == 'CONTROL')[0]
    c_themes = np.unique(s['factor_themes'][ix])
    od = dict.fromkeys(c_themes)
    for c in c_themes:
        cx = np.where(np.logical_and(s['factor_themes'] == c, s['factor_types'] == 'CONTROL'))[0]
        od[c] = s['factors'][cx]
    if factor_type is None or not isinstance(factor_type, str):
        return rd
    if factor_type.upper().strip() == 'RETURN':
        return rd
    else:
        return od


# def aggregate_by_themes(result):

# styles

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

fill_yellow = PatternFill("solid", start_color='FFFF66')
fill_green = PatternFill("solid", start_color='99FF99')
font_blue = Font(color=colors.BLUE)
font_blue_bold = Font(bold=True, color=colors.BLUE)
font_red_bold = styles.Font(bold=True, color='CC0000')