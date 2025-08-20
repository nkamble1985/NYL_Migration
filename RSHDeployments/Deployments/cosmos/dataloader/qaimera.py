import util.utilities as util
from util.utilities import display
import pandas as pd
import numpy as np
import os
import dataloader.market_data as md
import classes.root as root
import sys
import openpyxl as pxl
import openpyxl.styles as styles
from openpyxl.worksheet.dimensions import ColumnDimension

cache = {}


def input_location():
    location = os.path.join(util.default_output_location('reports'), 'qaimera')
    return location


def process_factor_covariance(bus_day, model='COSMOS_US_RISK_MODEL'):
    """
    process factor covariance
    :param bus_day:
    :param model: default 'COSMOS_US_RISK_MODEL'
    :return:

    Example:
        Input:
            process_factor_covaiance(20220915)

    Author: Yun Chen
    Copyright: Indigo Dao, LLC
    Date: September 15, 2022
    """
    if model is None or not isinstance(model, str):
        display(f"Incorrect risk model name type: str expected")
        raise ValueError(f"No valid model name")
    model = model.strip()
    d = util.parse_date(bus_day)
    file = os.path.join(input_location(), f"{model}.factor_covariance.{d.strftime(util.yyyymmdd_format)}.csv")
    df = pd.read_csv(file)
    df.set_index(df.columns[0], inplace=True)
    df.index.name = None
    display(f"{d}: {len(df.index)} x {len(df.columns)} factor covariance")
    e, v = np.linalg.eig(df)
    display(f"energy: max {np.max(e):.2E}, min {np.min(e): .2E}; {np.sum(e<=0)} non-positive; "
            f"condition number: {np.max(e)/np.min(e):.2E}")

    obj = root.load_object('COSMOS_US_RISK_MODEL')
    file = os.path.join(obj.factor_cov_location, f"{d.strftime(util.yyyymmdd_format)}.qd")
    if util.exists(file):
        display(f"{obj.name} factor covariance file exists: {file}\nOverwriting")
    vd = {'factors': df.index.to_numpy(), 'values': df.to_numpy()}
    zf = {'dates': d, 'horizons': np.array([1]), 'values': np.array([vd]), 'risk_model': obj.name,
          'factor_covariance_model': obj.name }
    try:
        util.save_data(zf, file)
        display(f"{d}: {obj.name} {len(df.index)} X {len(df.columns)} "
                f"factor covariance \nsuccessfully saved to {file}")
    except IOError as ioe:
        display(ioe)
        display(f"{d}: factor covariance ({obj.name}) failed due to "
                f"I/O error")
        raise ioe
    except Exception as iee:
        display(iee)
        display(f"{d}: factor covariance ({obj.name}) failed due to "
                f"exception")
        raise iee
    return df


def process_related(bus_day, model='COSMOS_US_RISK_MODEL'):
    """
    process related security correlations
    :param bus_day:
    :param model: default 'COSMOS_US_RISK_MODEL'
    :return:

    Example:
        Input:
            process_related(20220915)

    Author: Yun Chen
    Copyright: Indigo Dao, LLC
    Date: September 15, 2022
    """
    if model is None or not isinstance(model, str):
        display(f"no valid risk model type: str expected")
        raise ValueError(f"Incorrect input type: str expected for risk model")
    d = util.parse_date(bus_day)
    file = os.path.join(input_location(), f"{model}.related.{d.strftime(util.yyyymmdd_format)}.csv")
    df = pd.read_csv(file)
    # map IDs
    rf = df[['row', 'cusip_row']].copy(deep=True)
    rf.rename(columns={'row': 'ticker', 'cusip_row': 'cusip'}, inplace=True)
    cf = df[['column', 'cusip_column']].copy(deep=True)
    cf.rename(columns={'column': 'ticker', 'cusip_column': 'cusip'}, inplace=True)
    zf = pd.concat((rf, cf), axis=0, ignore_index=True)
    zf.drop_duplicates(inplace=True)
    sf = map_to_sec_ids(zf.to_numpy(), as_of=d)

    zf = pd.DataFrame(columns=['row', 'column', 'value'])

    ef = pd.DataFrame()
    for i in df.index:
        r = rf.loc[i, 'ticker']
        c = cf.loc[i, 'ticker']
        v = df.loc[i, 'value']
        try:
            if np.isnan(v):
                ef = pd.concat((ef, df.loc[i]), axis=0)
                display(f"{r} - {c}: NaN skipping")
                continue
            if r == c and v == 1.0:
                # display(f"{d}: {r} - {c}: 100% skipping")
                continue
            if r in sf.index:
                rid = sf.loc[r, 'sec_id']
            else:
                display(f"{d}: row: {i}: row ticker {r} not mapped")
                rid = None
            if c in sf.index:
                cid = sf.loc[c, 'sec_id']
            else:
                display(f"{d}: row: {i}: column ticker {c} not mapped")
                cid = None

            if rid is None or cid is None:
                display(f"skipping row {i}")
                ef = pd.concat((ef, df.loc[i]), axis=0)
                continue
            kf = pd.DataFrame([[rid, cid, v]], columns=['row', 'column', 'value'])
            zf = pd.concat([zf, kf], axis=0, ignore_index=True)
            del rid, cid
        except ValueError as ve:
            display(ve)
            display(f"{d}: processing related error: value error")
        except Exception as ee:
            display(ee)
            display(f"{d}: processing related error: exception")
    display(f"{d}: {len(zf.index)} pairs of related securities processed")
    zf.set_index(['row', 'column'], inplace=True)
    obj = root.load_object('COSMOS_US_RISK_MODEL')
    r_file = os.path.join(obj.residual_cov_location, f"related.{d.strftime('%Y%m%d')}.qd")
    if util.exists(r_file):
        display(f"{d}: file exists; overwriting: \n{r_file}")
    try:
        util.save_data(zf, r_file)
        display(f"{d}: {obj.name}: {len(zf.index)} pairs of related "
                f"security correlations saved to\n{r_file}")
    except IOError as ioe:
        display(ioe)
        display(f"{d}: unable to save due to I/O Error\n{r_file}")
        raise ioe
    except Exception as eee:
        display(eee)
        display(f"{d}: unable to save due to Exception\n{r_file}")
        raise eee

    return zf


def process_exposures(bus_day, model='COSMOS_US_RISK_MODEL'):
    """
    process factor exposures
    :param bus_day:
    :param model: default 'COSMOS_US_RISK_MODEL'
    :return:

    Example:
        Input:
            process_exposures(20220915)

    Author: Yun Chen
    Copyright: Indigo Dao, LLC
    Date: September 15, 2022
    """
    if model is None or not isinstance(model, str):
        display(f"no valid risk model type: str expected")
        raise ValueError(f"Incorrect input type: str expected for risk model")
    d = util.parse_date(bus_day)
    file = os.path.join(input_location(), f"{model}.stock.{d.strftime(util.yyyymmdd_format)}.csv")
    rf = pd.read_csv(file, keep_default_na=False, na_values='')
    rf.rename(columns={rf.columns[0]: 'ticker'}, inplace=True)
    rf.drop_duplicates(inplace=True)
    df = rf.copy(deep=True)
    rf.set_index('ticker', inplace=True)
    ix = np.where(pd.notnull(df['issuer']))[0]
    df = df.iloc[ix]
    ix = np.where(pd.notnull(df['permno']))[0]
    df = df.iloc[ix]
    sf = map_to_sec_ids(df[['ticker', 'cusip']].to_numpy(), as_of=d)
    df.set_index('ticker', inplace=True)
    ids = np.intersect1d(df.index, sf.index)
    missing = np.setdiff1d(rf.index, ids)
    df = df.loc[ids]
    sf = sf.loc[ids]
    df = df.join(sf)
    vc = df['sec_id'].value_counts()
    vec = vc[vc > 1] # those with duplicate rows
    dup = vec.index.to_numpy()
    df.reset_index(inplace=True)
    df.set_index('sec_id', inplace=True)
    if len(dup) > 0:
        display(f"{d}: {len(dup)} out of {len(df.index)} rows duplicate identifiers in \n{file}")
        drop = np.array([], dtype=int)
        md.get_tickers(dup, d, True)
        for dp in dup:
            try:
                ix = np.where(df.index == dp)[0]
                tik = md.get_tickers(dp, d, True)
                if tik.empty:
                    display(f"{d}: {dp}: duplicate {len(ix)} rows; no valid tickers; removing all")
                    drop = np.union1d(drop, ix)
                    continue
                iy = np.where(df.loc[dp, 'ticker'] != tik.loc[tik.index[0], 'ticker_region'][:-3])[0]
                display(f"{d}: {dp}: duplicate {len(ix)} rows: removing {len(iy)} rows")
                drop = np.union1d(drop, ix[iy])
            except ValueError as vve:
                display(vve)
                display(f"failed to disambiguate {dp}")
                continue
        if len(drop) > 0:
            df.reset_index(inplace=True)
            df.drop(df.index[drop], inplace=True)
            df.set_index('sec_id', inplace=True)
        vc = df.index.value_counts()
        vec = vc[vc > 1]  # those with duplicate rows
        dup = vec.index.to_numpy()
        if len(dup) > 0:
            ids = np.setdiff1d(df.index, dup)
            df = df.loc[ids]
            display(f"{d}: removing ambiguous {len(dup)} securities")
            display(f"{d}: {len(df.index)} rows to be processed")
    ix = np.where(pd.notnull(df.index))[0]
    df = df.loc[df.index[ix]]
    fg = root.load_object('COSMOS_US_FACTOR_GROUP')
    snapshots = fg.snapshot(expand_flag=True)
    factors = snapshots['factors'][np.where(~np.isin(snapshots['factor_types'],
                                                     ['CURRENCY', 'COUNTRY', 'INDUSTRY']))[0]]
    if util.exists(fg.exposure_location):
        file = os.path.join(fg.exposure_location, f"{d.strftime(util.yyyymmdd_format)}.qd")
        util.save_data(df[factors], file)
        display(f"Factor Group Level exposure file: {d}: saved to \n{file}")
    for f in factors:
        if f not in df.columns:
            display(f"{fg.name} factor: {f} not found for {d}: {file}\n"
                    f"Skipping...")
            continue
        o = root.load_object(f)
        if o is None:
            display(f"{fg.name} factor: {f} not set up properly: {d}\n"
                    f"Skipping...")
            continue
        try:
            f_file = os.path.join(o.exposure_location, f"{d.strftime(util.yyyymmdd_format)}.qd")
            if not util.exists(o.exposure_location):
                util.makedirs(o.exposure_location)
                display(f"created: {o.exposure_location}")
            vf = df[[f]].copy(deep=True)
            vf.reset_index(inplace=True)
            vf.rename(columns={f: 'values', 'sec_id': 'sec_ids'}, inplace=True)
            vf['source'] = o.source
            util.save_data(vf, f_file)
            display(f"{f}: exposure {len(vf.index)} x 1: {d}: saved to {f_file}")
        except ValueError as ve:
            display(ve)
            display(f"{f}: {d}: unable to process: value error")
        except IOError as ive:
            display(ive)
            display(f"{f}: {d}: unable to process: I/O error")
        except Exception as ee:
            display(ee)
            display(f"{f}: {d}: unable to process: Exception")

    des = ['PredictedVolatility', 'PredictedBeta']
    for f in des:
        if f in df.columns:
            o = root.load_object(f)
            if not util.exists(o.descriptor_location):
                util.makedirs(o.descriptor_location)
                display(f"created {o.descriptor_location}")
            try:
                d_file = os.path.join(o.descriptor_location, f"{d.strftime(util.yyyymmdd_format)}.qd")
                vf = df[[f]].copy(deep=True)
                vf.reset_index(inplace=True)
                vf.rename(columns={f: 'values', 'sec_id': 'sec_ids'}, inplace=True)
                vf['source'] = o.source
                util.save_data(vf, d_file)
                display(f"{f}: descriptors {len(vf.index)} x 1: {d}: saved to {d_file}")
            except ValueError as ve:
                display(ve)
                display(f"{f}: {d}: unable to process: value error")
            except IOError as ive:
                display(ive)
                display(f"{f}: {d}: unable to process: I/O error")
            except Exception as ee:
                display(ee)
                display(f"{f}: {d}: unable to process: Exception")

    obj = root.load_object('COSMOS_US_RISK_MODEL')
    r_file = os.path.join(obj.residual_cov_location, f"{d.strftime(util.yyyymmdd_format)}.qd")
    if 'residual' in df.columns:
        vd = df[['residual']]
        zf = {'dates': d, 'horizons': np.array([1]), 'values': [vd], 'risk_model': obj.name,
              'factor_group': obj.factor_groups[0], 'residual_covariance_model': obj.name}
        try:
            util.save_data(zf, r_file)
            display(f"{d}: {len(df.index)} residuals ({obj.name})\n"
                    f"Successfully saved to {r_file}")
        except IOError as ioe:
            display(ioe)
            display(f"IO Error")
            raise ioe
    else:
        display(f"{d}: residual not found: \n{file}")
    e_file = os.path.join(input_location(), f"error.stock.{d.strftime(util.yyyymmdd_format)}.xlsx")
    ix = np.where(np.isin(sf.to_numpy(), dup))[0]
    dup_tickers = sf.index[ix].to_numpy()
    ef = rf.loc[np.union1d(missing, dup_tickers)]
    ef['issue'] = None
    ef.loc[missing, 'issue'] = 'Unrecognized identifier'
    ef.loc[dup_tickers, 'issue'] = 'Duplicates'
    ef.to_excel(e_file)
    display(f"{d}: stock exposures, residuals, {len(missing)} "
            f"unidentified, {len(dup)} duplicates, recorded to\n{e_file}")

    return df


def process_factor_model(start_date=None, end_date=None, skip_factor_returns=False, skip_residuals=False,
                         factor_group='COSMOS_US_FACTOR_GROUP'):
    """
    process factor model
    :param start_date:
    :param end_date:
    :param skip_factor_returns: default False
    :param skip_residuals: default False
    :param factor_group: default 'COSMOS_US_FACTOR_GROUP'
    :return:

    Example:
        Input:
            process_factor_model(20220910,20220915)

    Author: Yun Chen
    Copyright: Indigo Dao, LLC
    Date: September 15, 2022
    """
    fg = root.load_object(factor_group)
    days = util.load_business_days(fg.calendar, start_date, end_date)
    if len(days) == 0:
        raise ValueError(f"no valid business days: calendar {fg.calendar}")
    directory = input_location()
    if not skip_factor_returns:
        for d in days:
            # ===============================================
            #
            # factor returns
            #
            # ===============================================
            try:
                file = os.path.join(directory, f"{fg.name}.factor_returns.{d.strftime(util.yyyymmdd_format)}.xlsx")
                if not util.exists(file):
                    display(f"{fg.name}: {d}: input factor return file not found: {file}")
                    continue
                df = pd.read_excel(file, header=0)
                df.set_index(df.columns[0], inplace=True)
                df.index.name = None
                df.rename(columns={df.columns[0]: 1}, inplace=True)
                d_file = os.path.join(fg.regression_location, f"factor_returns.{d.strftime(util.yyyymmdd_format)}.qd")
                try:
                    util.save_data(df, d_file)
                    display(f"{fg.name}: {d}: {len(df.index)} factor returns saved to {d_file}")
                except IOError as ioe:
                    display(ioe)
                    display(f"{fg.name}: due to I/O error unable to process factor returns on {d}")
                    continue
                except Exception as ee:
                    display(ee)
                    display(f"{fg.name}: due to Exception unable to process factor returns on {d}")
                    continue
            except IOError as oe:
                display(oe)
                display(f"{fg.name}: due to I/O error unable to process factor returns on {d}")
                continue
            except Exception as eee:
                display(eee)
                display(f"{fg.name}: due to Exception unable to process factor returns on {d}")
                continue
    # ===============================================
    #
    # stock residual returns
    #
    # ===============================================
    if not skip_residuals:
        for d in days:
            try:
                file = os.path.join(directory, f"{fg.name}.residuals.{d.strftime(util.yyyymmdd_format)}.xlsx")
                if not util.exists(file):
                    display(f"{fg.name}: {d}: input residual return file not found: {file}")
                    continue
                rf = pd.read_excel(file, header=0)
                if 'cusips' in rf.columns:
                    rf.rename(columns={'cusips': 'cusip'}, inplace=True)
                rf = rf[['ticker', 'cusip', 'values', 'name']].copy(deep=True)
                rf.drop_duplicates(keep='last', inplace=True)
                ix = np.where(pd.notnull(rf['ticker']))[0]
                rf = rf.iloc[ix]
                df = rf.copy(deep=True)
                rf.set_index('ticker', inplace=True)
                sf = map_to_sec_ids(df[['ticker', 'cusip']].to_numpy(), as_of=d)
                df.set_index('ticker', inplace=True)
                ids = np.intersect1d(df.index, sf.index)
                missing = np.setdiff1d(rf.index, ids)
                df = df.loc[ids]
                sf = sf.loc[ids]
                df = df.join(sf)
                vc = df['sec_id'].value_counts()
                vec = vc[vc > 1]  # those with duplicate rows
                dup = vec.index.to_numpy()
                df.reset_index(inplace=True)
                df.set_index('sec_id', inplace=True)
                df.rename(columns={'values': 1}, inplace=True)
                # if len(dup) > 0:
                #     display(f"{d}: {len(dup)} out of {len(df.index)} rows duplicate identifiers in \n{file}")
                #     ids = np.setdiff1d(df.index, dup)
                #     df = df.loc[ids]
                #     display(f"{d}: {len(df.index)} rows to be processed")
                if len(dup) > 0:
                    display(f"{d}: {len(dup)} out of {len(df.index)} rows duplicate identifiers in \n{file}")
                    drop = np.array([], dtype=int)
                    md.get_tickers(dup, d, True)
                    for dp in dup:
                        try:
                            ix = np.where(df.index == dp)[0]
                            tik = md.get_tickers(dp, d, True)
                            if tik.empty:
                                display(f"{d}: {dp}: duplicate {len(ix)} rows; no valid tickers; removing all")
                                drop = np.union1d(drop, ix)
                                continue
                            iy = np.where(df.loc[dp, 'ticker'] != tik.loc[tik.index[0], 'ticker_region'][:-3])[0]
                            display(f"{d}: {dp}: duplicate {len(ix)} rows: removing {len(iy)} rows")
                            drop = np.union1d(drop, ix[iy])
                        except ValueError as vve:
                            display(vve)
                            display(f"failed to disambiguate {dp}")
                            continue
                    if len(drop) > 0:
                        df.reset_index(inplace=True)
                        df.drop(df.index[drop], inplace=True)
                        df.set_index('sec_id', inplace=True)
                    vc = df.index.value_counts()
                    vec = vc[vc > 1]  # those with duplicate rows
                    dup = vec.index.to_numpy()
                    if len(dup) > 0:
                        ids = np.setdiff1d(df.index, dup)
                        df = df.loc[ids]
                        display(f"{d}: removing ambiguous {len(dup)} securities")
                        display(f"{d}: {len(df.index)} rows to be processed")
                ix = np.where(pd.notnull(df.index))[0]
                df = df.loc[df.index[ix]]

                d_file = os.path.join(fg.regression_location, f"residuals.{d.strftime(util.yyyymmdd_format)}.qd")
                try:
                    save_factors = [1, 'ticker']
                    util.save_data(df[save_factors], d_file)
                    display(f"{fg.name}: {d}: {len(df.index)} residual returns saved to {d_file}")
                except IOError as ioe:
                    display(ioe)
                    display(f"{fg.name}: due to I/O error unable to process residual returns on {d}")
                except Exception as ee:
                    display(ee)
                    display(f"{fg.name}: due to Exception unable to process factor returns on {d}")

                e_file = os.path.join(directory, f"error.residual.{d.strftime(util.yyyymmdd_format)}.xlsx")
                ix = np.where(np.isin(sf.to_numpy(), dup))[0]
                dup_tickers = sf.index[ix].to_numpy()
                ef = rf.loc[np.union1d(missing, dup_tickers)]
                ef['issue'] = None
                ef.loc[missing, 'issue'] = 'Unrecognized identifier'
                ef.loc[dup_tickers, 'issue'] = 'Duplicates'
                ef.to_excel(e_file)
                display(f"{d}: stock residuals, {len(missing)} "
                        f"unidentified, {len(dup)} duplicates, recorded to\n{e_file}")
            except IOError as oe:
                display(oe)
                display(f"{fg.name}: due to I/O error unable to process residual returns on {d}")
            except Exception as eee:
                display(eee)
                display(f"{fg.name}: due to Exception unable to process residual returns on {d}")
    return df


def process_classification(bus_day):
    d = util.parse_date(bus_day)
    file = os.path.join(input_location(), 'cosmos.xlsx')
    if not util.exists(file):
        display(f"cosmos classification file not found; returning")
        return False
    df = pd.read_excel(file)
    if df.empty or len(df.index) == 0:
        display(f"cosmos classification: no new updates; returning")
        return False
    df = df[['ticker', 'sector', 'indgrp', 'industry', 'cusips']]
    ix = np.where(pd.notnull(df['ticker']))[0]
    df = df.iloc[ix]
    ids = map_to_sec_ids(df[['ticker', 'cusips']].to_numpy().astype('str'), as_of=d)
    ids = ids[['sec_id']]
    df = df.merge(ids, how='left', left_on='ticker', right_index=True)
    ix = np.where(pd.notnull(df['sec_id']))[0]
    df = df.iloc[ix]
    df.set_index('sec_id', inplace=True)
    df['from_dt'] = None
    df['to_dt'] = None
    if df.empty:
        display(f"No valid addition to classifications")
        return True
    # figure out from through dates
    tickers = md.get_tickers(df.index.to_numpy(), None, True)
    if tickers.empty:
        display(f"No valid tickers found to determine from and through date: skip processing classifications")
        return True
    tickers.set_index('sec_id', inplace=True)
    si = np.where(tickers.columns == 'start_date')[0][0]
    ei = np.where(tickers.columns == 'end_date')[0][0]
    for s in df.index:
        if df.loc[s, 'ticker'] is None:
            continue
        tc = df.loc[s, 'ticker'] + '-US'
        tk = tickers.loc[[s]]
        ix = np.where(tk['ticker_region'] == tc)[0]
        if len(ix) == 0:
            continue
        df.loc[s, 'from_dt'] = tk.iloc[ix[0], si]
        df.loc[s, 'to_dt'] = tk.iloc[ix[0], ei]
    gx = np.where(pd.notnull(df['to_dt']))[0]
    nx = np.where(pd.isnull(df['to_dt']))[0]
    if len(nx) > 0:
        display(f"{len(nx)} rows of classification mismatched with tickers")
    df = df.iloc[gx]
    ind = md.get_all_classifications('COSMOS')
    original = len(ind.index)
    si = np.where(ind.columns == 'from_dt')[0][0]
    ei = np.where(ind.columns == 'to_dt')[0][0]
    new = 0
    update = False
    modified = 0
    for s in df.index:
        try:
            if s not in ind.index:
                ind = pd.concat([ind, df.loc[[s]]], axis=0)
                new = new + 1
                update = True
            else:
                t = df.loc[s, 'ticker']
                tk = ind.loc[[s]]
                ix = np.where(tk['ticker'] == t)[0]
                if len(ix) == 0:
                    ind = pd.concat([ind, df.loc[[s]]], axis=0)
                    new = new + 1
                else:
                    tk.iloc[ix, si] = df.loc[s, 'from_dt']
                    tk.iloc[ix, ei] = df.loc[s, 'to_dt']
                    ind.loc[[s]] = tk
                    modified = modified + 1
                update = True
        except ValueError as ve:
            display(ve)
            display(f"{s}: merging into existing classification: value error")
            continue
        except Exception as ee:
            display(ee)
            display(f"{s}: merging into existing classification: exception")
            continue
    display(f"original {original}; new {new}; modified {modified}; now {len(ind.index)} rows of cosmos classification")
    if update:
        c_file = os.path.join('classifications', 'COSMOS', 'classification.qd')
        util.save_data(ind, c_file)
        display(f"{len(ind.index)}: {new} new, {modified} modified: COSMOS classification saved to {c_file}")
    return True


def process_us_model(start_date, end_date, skip_cov=False, skip_exp=False, skip_related=False, skip_factor_model=False,
                     skip_classification=False, skip_diagnostics=False):
    """
    process cosmos US models: factor covariance, factor exposures, factor returns, related security correlations
    :param start_date:
    :param end_date:
    :param skip_cov: default False
    :param skip_exp: default False
    :param skip_related: default False
    :param skip_factor_model: default False
    :param skip_classification: default False
    :param skip_diagnostics: default False
    :return:

    Example:
        Input:
            process_us_model(20220914, 20220915)

    Author: Yun Chen
    Copyright: Indigo Dao, LLC
    Date: September 15, 2022
    """
    days = util.load_business_days('US', start_date, end_date)
    for d in days:
        try:
            if not skip_cov:
                fc = process_factor_covariance(d)
            if not skip_exp:
                fb = process_exposures(d)
            if not skip_related:
                try:
                    fr = process_related(d)
                except Exception as see:
                    display(f"{see}")
                    display(f"{d}: related security correlation: exception")
            if not skip_factor_model:
                ff = process_factor_model(d, d)
            if not skip_classification:
                cc = process_classification(d)
            if not skip_diagnostics:
                df = run_diagnostics(d, save_flag=True)
        except ValueError as ve:
            display(ve)
            display(f"{util.caller()}: {d}: unable to process risk model: Value Error")
        except IOError as ie:
            display(ie)
            display(f"{util.caller()}: {d}: unable to process risk model: IOError")
        except Exception as ee:
            display(ee)
            display(f"{util.caller()}: {d}: unable to process risk model: Exception")
    return True


def process_indicators():
    file = os.path.join(input_location(), 'regimes.xlsx')
    if not util.exists(file):
        display(f"Indicators not found: {file}")
        return False
    df = pd.read_excel(file, header=0)
    df = df[['values', 'indicators', 'from_date', 'to_date']]
    i_file = os.path.join(util.default_output_location('macro'), 'all_regimes.qd')
    indicators = np.unique(df['indicators'].to_numpy())
    if not util.exists(i_file):
        util.save_data(df, i_file)
        display(f"First time macro regime: {len(indicators)} total {len(df.index)} rows")
        return df
    regimes = util.load_data(i_file)
    for ind in indicators:
        ix = np.where(df['indicators'] == ind)[0]
        if len(ix) == 0:
            continue
        if ind in regimes['indicators'].to_list():
            iy = np.where(regimes['indicators'] == ind)[0]
            regimes.drop(regimes.index[iy], inplace=True)
            display(f"{len(iy)} rows of {ind} expunged from macro regime: now total {len(regimes.index)} rows")
        regimes = pd.concat((regimes, df.iloc[ix]), axis=0)
        display(f"{len(ix)} rows {ind} appended to macro regime: now total {len(regimes.index)} rows")
    return regimes


def dates_to_differential(df, calendar_str='GL'):
    zf = pd.DataFrame()
    for ix, c in enumerate(df.index):
        if ix == 0:
            zf = df.iloc[[0]]
            continue
        iy = ix - 1
        lv = df.loc[df.index[iy], 'values']
        cv = df.loc[c, 'values']
        if lv != cv:
            zf = pd.concat((zf, df.iloc[[ix]]), axis=0, ignore_index=True)
            continue
        else:
            if zf.loc[zf.index[-1], 'to_date'] == df.loc[c, 'from_date']:
                zf.loc[zf.index[-1], 'to_date'] = df.loc[c, 'to_date']
            else:
                zf = pd.concat((zf, df.iloc[[ix]]), axis=0, ignore_index=True)
                continue
    d = util.most_recent_business_day(util.today(), calendar_str)
    if zf.loc[zf.index[-1], 'from_date'] >= d:
        zf.loc[zf.index[-1], 'to_date'] = util.parse_date(99991231)
    else:
        if zf.loc[zf.index[-1], 'to_date'] >= util.today():
            zf.loc[zf.index[-1], 'to_date'] = util.parse_date(99991231)
    return zf


def merge_differential(rf, xf):
    ix = np.where(rf['to_date'] < xf.loc[xf.index[0], 'from_date'])[0]
    zf = rf.iloc[ix]
    for ix, c in enumerate(xf.index):
        s = xf.loc[c, 'from_date']
        e = xf.loc[c, 'to_date']
        iy = np.where(np.logical_and(rf['from_date'] <= s, rf['to_date'] > s))[0]
        if len(iy) == 0:
            zf = pd.concat([zf, xf.loc[[c]]], axis=0, ignore_index=True)
        else:
            vf = rf.iloc[iy]
            vf.loc[vf.index[0], 'to_date'] = s
            zf = pd.concat([zf, vf], axis=0, ignore_index=True)
            zf = pd.concat([zf, xf.loc[[c]]], axis=0, ignore_index=True)
    zf = dates_to_differential(zf)
    return zf


def map_to_sec_ids(ticker_cusips, exclude_unrecognizable=True, include_recognized=True, as_of=None, save_flag=True):
    """
    primary ticker then cusips
    matched to sec_id

    :param ticker_cusips: N x 2, [Regional Ticker, Cusips]
    :param exclude_unrecognizable: default True
    :param include_recognized: default True
    :param as_of: default None
    :param save_flag: default True
    :return:

    Example:
        Input:
            map_to_sec_ids(['CAEJ-US', '169364106'])
        Output:
            array(['GD87LC-R'])

    Author: Yun Chen
    Copyright: Indigo Dao, LLC
    Date: August 25, 2022
    """
    ix = np.where(pd.notnull(ticker_cusips[:, 0]))[0]
    ticker_cusips = ticker_cusips[ix, :]
    tickers = ticker_cusips[:, 0]
    if ticker_cusips.ndim == 1 or ticker_cusips.shape[1] == 1:
        cusips = None
    else:
        cusips = ticker_cusips[:, 1]
    ad = None
    if as_of is not None:
        ad = util.parse_date(as_of)
    bad_file = os.path.join(util.default_output_location('reports'), 'unrecognized.qd')
    if util.exists(bad_file):
        bad = util.load_data(bad_file)
    else:
        bad = pd.DataFrame(columns=['sec_id', 'cusips'])
    missing = tickers
    good = pd.DataFrame()
    if 'security_map' in cache:
        sm = cache['security_map']
        gx = np.where(pd.notnull(sm['sec_id']))[0]
        good = sm.iloc[gx]
        missing = np.setdiff1d(tickers, good['ticker'].to_numpy())
    file = os.path.join(util.default_output_location('reports'), 'sec_map.qd')
    if util.exists(file) and include_recognized:
        pm = util.load_data(file)
        good = pd.concat([good, pm], axis=0)
        good.drop_duplicates(inplace=True)
        gx = np.where(pd.notnull(good['sec_id']))[0]
        good = good.iloc[gx]
        missing = np.setdiff1d(missing, good['ticker'].to_numpy())
    if len(bad) > 0 and exclude_unrecognizable:
        mb = np.intersect1d(missing, bad.index)
        if len(mb) > 0:
            display(f"{len(mb)} out of {len(missing)} missing are unrecognizable; excluded")
            missing = np.setdiff1d(missing, bad.index)
    if len(missing) > 0:
        cu = np.array([None]*len(missing))
        for i, m in enumerate(missing):
            ix = np.where(tickers == m)[0]
            cu[i] = cusips[ix[0]]
        sf = map_ticker_and_cusips(missing, cu)
        if sf is not None:
            if not sf.empty:
                good = pd.concat([good, sf], axis=0)
                good.drop_duplicates(keep='last', inplace=True)
    if ad is not None:
        ix = np.where(np.logical_and(good['start_date'] <= ad, good['end_date'] > ad))[0]
        good = good.iloc[ix]
    good.set_index('ticker', inplace=True)
    ids = np.intersect1d(tickers, good.index)
    return good.loc[ids]


# ---------------------------------
# process russell historical files
# ---------------------------------
def process_russell(start_date, end_date, calendar_str='US'):
    days = util.load_business_days(calendar_str, start_date, end_date)
    if len(days) == 0:
        display(f"No valid business days according to {calendar_str}")
        return None
    location = os.path.join(util.default_output_location('reports'), 'russell')
    indices = np.array(['R3', 'R3G', 'R3V', 'R1', 'R1G', 'R1V', 'R2', 'R2G', 'R2V',
                        'R25', 'R25G', 'R25V', 'RMID', 'MIDG', 'MIDV', 'RT2', 'RT2G',
                        'RT2V', 'RSSC', 'RSSCG', 'RSSCV'])
    for d in days:
        try:
            file = os.path.join(location, f"H_{d.strftime(util.yyyymmdd_format)}.csv")
            if not util.exists(file):
                display(f"{d}: NOT FOUND: {file}")
                display(f"Skipping ...")
                continue
            df = pd.read_csv(file)
            dates = util.parse_date(df['Date'])
            ix = np.where(pd.notnull(dates))[0]
            df = df.iloc[ix]
            ids = df[['Cusip', 'ISIN', 'Ticker', 'Exchange', 'Name']]
            sids = map_ticker_cusip_isin(ids)
            sids.set_index(['Cusip', 'ISIN', 'Ticker'], inplace=True)
            for ind in indices:
                try:
                    tag = f"WT-{ind.strip().upper()}"
                    if tag not in df:
                        display(f"{d}: {ind} not found: missing column {tag}")
                        continue
                    zf = df[['Cusip', 'ISIN', 'Ticker', tag]].copy()
                    zf.rename(columns={tag: 'weights'}, inplace=True)
                    zf.set_index(['Cusip', 'ISIN', 'Ticker'], inplace=True)
                    zf = zf.join(sids)
                    zf.reset_index(inplace=True)
                    zf.loc[zf.index, 'weights'] = pd.to_numeric(zf['weights'].to_numpy(), errors='coerce')
                    ix = np.where(pd.notnull(zf['weights']))[0]
                    zf = zf.iloc[ix]
                    ix = np.where(zf['weights'] != 0)[0]
                    zf = zf.iloc[ix]
                    loc = os.path.join(location, f"{ind}")
                    if not util.exists(loc):
                        util.makedirs(loc)
                        display(f"{ind}: output location created: {loc}")
                    i_file = os.path.join(loc, f"{d.strftime(util.yyyymmdd_format)}.csv")
                    zf.to_csv(i_file)
                    display(f"{d}: {ind}: {len(zf.index)} holdings saved to {i_file}")
                except ValueError as vie:
                    display(vie)
                    display(f"Unable to process russell {ind} due to value error: {d}")
                except IOError as iie:
                    display(iie)
                    display(f"Unable to process russell {ind} due to IO error: {d}")
                except Exception as iee:
                    display(iee)
                    display(f"Unable to process russell {ind} due to Exception: {d}")
        except ValueError as ve:
            display(ve)
            display(f"Unable to process russell file due to value error: {d}")
        except IOError as ie:
            display(ie)
            display(f"Unable to process russell file due to IO error: {d}")
        except Exception as ee:
            display(ee)
            display(f"Unable to process russell file due to Exception: {d}")


def map_ticker_cusip_isin(ids, day=None):
    file = os.path.join(util.default_output_location('reports'), 'tmp', 'russell_id_map.qd')
    if not util.exists(file):
        cached = pd.DataFrame()
    else:
        cached = util.load_data(file)
    sec_ids = pd.DataFrame(None, index=ids.index, columns=['sec_id'])
    sec_ids = sec_ids.join(ids)
    if cached.empty:
        missing = ids
        missing['sec_id'] = None
    else:
        cf = cached.set_index(['Cusip', 'ISIN', 'Ticker'])
        sf = sec_ids.set_index(['Cusip', 'ISIN', 'Ticker'])
        missing = sf.index.difference(cf.index)
        missing = pd.DataFrame(None, index=missing, columns=['sec_id'])
        missing.reset_index(inplace=True)
    if not missing.empty:
        display(f"{len(missing.index)} yet to be matched to unique IDs")
        tickers = [t + '-US' for t in missing['Ticker']]
        tids = md.get_sec_ids(tickers, currency='USD', region='AMER', day=day)
        cids = md.get_sec_ids(missing['Cusip'].to_numpy(), 'cusip', currency='USD', region='AMER', day=day)
        iids = md.get_sec_ids(missing['ISIN'].to_numpy(), 'ISIN', currency='USD', region='AMER', day=day)
        for r in missing.index:
            try:
                t = missing.loc[r, 'Ticker']
                c = missing.loc[r, 'Cusip']
                i = missing.loc[r, 'ISIN']
                tc = t + '-US'
                tx = np.where(tids['ticker_region'] == tc)[0]
                if len(tx) == 0:
                    display(f"ticker {t}: no matching sec_id found")
                    st = np.array([])
                else:
                    st = tids.loc[tids.index[tx[0]], 'sec_id']
                cx = np.where(cids['cusip'] == c)[0]
                if len(cx) == 0:
                    display(f"cusip {c}: no matching sec_id found")
                    ct = np.array([])
                else:
                    ct = cids.loc[cids.index[cx[0]], 'sec_id']
                ix = np.where(iids['isin'] == i)[0]
                if len(ix) == 0:
                    display(f"ISIN {i}: no matching sec_id found")
                    it = np.array([])
                else:
                    it = iids.loc[iids.index[ix[0]], 'sec_id']
                sid = np.intersect1d(it, np.intersect1d(ct, st))
                if len(sid) == 0:
                    display(f"ticker: {t}: cusip: {c}: ISIN {i}: no intersection <===========")
                    sid = np.intersect1d(ct, st)
                    if len(sid) == 0:
                        display(f"ticker: {t}: cusip: {c}:no intersection <===========")
                if len(sid) == 1:
                    missing.loc[r, 'sec_id'] = sid[0]
                elif len(sid) > 1:
                    ref = md.get_references(sid)
                    ia = np.where(ref['is_active'])[0]
                    ref = ref.iloc[ia]
                    if len(ref.index) == 0:
                        display(f"Strange error: No reference found for sec_ids matched for {t} and {c}")
                    else:
                        if len(ref.index) > 1:
                            display(f"More than 2 live securities found for {t} and {c}: picking the first")
                        missing.loc[r, 'sec_id'] = ref.index[0]
            except ValueError as ve:
                display(ve)
                display(f"{missing.loc[r, 'Ticker']} : failed to align id")

        good_index = np.where(pd.notnull(missing['sec_id']))[0]
        if len(good_index) > 0:
            cached = pd.concat([cached, missing.iloc[good_index]], axis=0, ignore_index=True)
            cached.drop_duplicates(keep='last', inplace=True)
            try:
                util.save_data(cached, file)
                display(f"Russell Index Mapping Updated (Total {len(cached.index)} symbols): {file}")
            except IOError as ioe:
                display(ioe)
                display(f"Unable to saved previously mapped securities to {file}")
    if len(cached.index) > 0:
        cf = cached.set_index(['Cusip', 'ISIN', 'Ticker'])
        sec_ids.set_index(['Cusip', 'ISIN', 'Ticker'], inplace=True)
        sym = sec_ids.index.intersection(cf.index)
        sec_ids.loc[sym, 'sec_id'] = cf.loc[sym, 'sec_id']
        sec_ids.reset_index(inplace=True)
    return sec_ids


def map_tickers(tickers):
    if isinstance(tickers, str):
        tickers = np.array([tickers])
    if isinstance(tickers, list):
        tickers = np.array(tickers)
    tickers = np.unique(tickers)
    ix = np.where(pd.notnull(tickers))[0]
    tickers = tickers[ix]
    c_tickers = [t + '-US' for t in tickers]
    ids = md.get_sec_ids(c_tickers, 'ticker')
    rids = np.unique(ids['sec_id'])
    # for k in ids.keys():
    #     if ids[k] is None:
    #         continue
    #     rids = np.union1d(rids, ids[k])
    if len(rids) == 0:
        df = pd.DataFrame(columns=['sec_ids', 'ticker_region', 'start_date', 'end_date'])
        return df
    df = md.get_tickers(rids, None, True)
    ix = np.where(np.isin(df['ticker_region'].to_numpy(), c_tickers))[0]
    df = df.iloc[ix]
    df.reset_index(inplace=True)
    df.drop(df.columns[0], axis=1, inplace=True)
    df['ticker'] = df['ticker_region']
    for s in df.index:
        df.loc[s, 'ticker'] = df.loc[s, 'ticker'][:-3]
    return df


def map_ticker_and_cusips(tickers, cusips):
    tf = map_tickers(tickers)
    if tf is None or tf.empty:
        return None
    cf = md.get_cusips(tf['sec_id'], None, True)
    ut = np.unique(tf['ticker'])
    tf['keep'] = True
    for t in ut:
        ix = np.where(tickers == t)[0]
        if len(ix) == 0:
            continue
        cu = cusips[ix[0]]
        if cu is None or not isinstance(cu, str):
            continue
        cu = cu.split(' ')
        iy = np.where(tf['ticker'] == t)[0]
        if len(iy) == 0:
            continue
        ids = tf['sec_id'].iloc[iy].to_numpy()
        iz = np.where(np.isin(cf['cusip'], cu))[0]
        cids = cf['sec_id'].iloc[iz].to_numpy()
        drop = np.setdiff1d(ids, cids)
        if len(drop) == 0:
            continue
        ik = np.where(np.isin(tf['sec_id'], drop))[0]
        dx = np.intersect1d(iy, ik)
        tf.loc[tf.index[dx], 'keep'] = False
    return tf


def run_diagnostics(bus_day, prior_day=None, calendar_str='US', save_flag=False):
    days = util.load_business_days(calendar_str, None, bus_day)
    d = days[-1]
    if prior_day is None:
        prev_day = days[-2]
    else:
        prev_day = util.prior_day(prior_day)
    df = pd.DataFrame(columns=[prev_day, d])

    # risk models
    risk_model = 'COSMOS_US_RISK_MODEL'
    risk = root.load_object(risk_model)
    risk_data = {d: None, prev_day: None}
    res_data = {d: None, prev_day: None}
    for rd in df.columns.to_numpy():
        fc = risk.load_factor_covariance(rd)
        ff = pd.DataFrame(fc['values'], index=fc['factors'], columns=fc['factors'])
        fv = pd.Series(np.sqrt(np.diag(ff)), index=fc['factors'])
        df.loc['valid factors', rd] = pd.notnull(fv).sum()
        df.loc['invalid factors', rd] = pd.isnull(fv).sum()
        df.loc['condition number', rd] = np.linalg.cond(fc['values'])
        df.loc['max volatility', rd] = np.nanmax(fv)
        ix = np.nanargmax(fv)
        df.loc['max factor', rd] = fv.index[ix]
        df.loc['min volatility', rd] = np.nanmin(fv)
        ix = np.nanargmin(fv)
        df.loc['min factor', rd] = fv.index[ix]
        risk_data[rd] = ff

        file = os.path.join(risk.residual_cov_location, f"{rd.strftime(util.yyyymmdd_format)}.qd")
        if not util.exists(file):
            display(f"{rd}: residual file not found: {file}")
            continue
        sc = util.load_data(file)
        sc = sc['values'][0]
        sc = pd.DataFrame(np.sqrt(sc.to_numpy()), index=sc.index, columns=[rd])
        res_data[rd] = sc
        df.loc['valid residuals', rd] = pd.notnull(sc).sum().sum()
        df.loc['invalid residuals', rd] = pd.isnull(sc).sum().sum()
        df.loc['max residual', rd] = np.nanmax(sc)
        ix = np.nanargmax(sc)
        df.loc['max security', rd] = sc.index[ix]
        df.loc['min residual', rd] = np.nanmin(sc)
        ix = np.nanargmin(sc)
        df.loc['min security', rd] = sc.index[ix]

    # factor vol comparisons
    cv = risk_data[d]
    pv = risk_data[prev_day]
    factors = np.intersect1d(cv.index, pv.index)
    cv = cv.loc[factors, factors]
    ct = pd.DataFrame(np.sqrt(np.diag(cv)), index=cv.index, columns=[d])
    pv = pv.loc[factors, factors]
    pt = pd.DataFrame(np.sqrt(np.diag(pv)), index=pv.index, columns=[prev_day])
    vol = pd.concat((ct, pt), axis=1)
    df.loc['factor day to day correlation'] = vol.corr().iloc[0, 1]
    vol['difference'] = vol[d] - vol[prev_day]
    df.loc['factor max increase', d] = np.nanmax(vol['difference'])
    ix = np.nanargmax(vol['difference'])
    df.loc['factor max increase factor', d] = vol.index[ix]
    df.loc['factor max decrease', d] = np.nanmin(vol['difference'])
    ix = np.nanargmin(vol['difference'])
    df.loc['factor max decrease factor', d] = vol.index[ix]

    # residual vol comparisons
    cv = res_data[d]
    pv = res_data[prev_day]
    rv = pd.concat((cv, pv), axis=1)
    rv['difference'] = rv[d] - rv[prev_day]
    df.loc['residual day to day correlation'] = rv[[d, prev_day]].corr().iloc[0, 1]
    df.loc['dropped residuals', d] = np.logical_and(pd.notnull(rv[d]), pd.isnull(rv[prev_day])).sum()
    df.loc['dropped residuals %', d] = df.loc['dropped residuals', d] / len(rv.index)
    df.loc['gained residuals', d] = np.logical_and(pd.notnull(rv[prev_day]), pd.isnull(rv[d])).sum()
    df.loc['gained residuals %', d] = df.loc['gained residuals', d] / len(rv.index)
    df.loc['max residual increase', d] = np.nanmax(rv['difference'])
    ix = np.nanargmax(rv['difference'])
    df.loc['max residual increase security', d] = rv.index[ix]
    df.loc['max residual decrease', d] = np.nanmin(rv['difference'])
    ix = np.nanargmin(rv['difference'])
    df.loc['max residual decrease security', d] = rv.index[ix]
    df.loc['increase by more than 5%', d] = (rv['difference'] > 0.05).sum()
    df.loc['decrease by more than 5%', d] = (rv['difference'] < -0.05).sum()
    if save_flag:
        wb = pxl.Workbook()
        output_file = os.path.join(input_location(), f"diagnostic.{d.strftime(util.yyyymmdd_format)}.xlsx")
        sheet = wb.active
        sheet.title = 'Summary'
        row = 1
        col = 1
        sheet.cell(row, col).value = 'Cosmos Daily Summary'
        sheet.cell(row, col).font = font_blue_bold
        row = row + 1
        sheet.cell(row, col).value = 'Date'
        sheet.cell(row, col + 1).value = d.strftime(util.YY_MM_DD_format)
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = 'Prior'
        sheet.cell(row, col + 1).value = prev_day.strftime(util.YY_MM_DD_format)
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')

        row = row + 2
        sheet.cell(row, col).value = 'Risk Model'
        sheet.cell(row, col + 1).value = risk_model
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')

        row = row + 1
        sheet.cell(row, col).value = '# of valid factors'
        v = df.loc['valid factors', d]
        sheet.cell(row, col + 1).value = v
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = '# of missing factors'
        v = df.loc['invalid factors', d]
        sheet.cell(row, col + 1).value = v
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        if v > 0:
            sheet.cell(row, col + 1).font = font_red_bold
            sheet.cell(row, col + 1).fill = fill_yellow
        row = row + 1
        sheet.cell(row, col).value = 'day to day factor vol correlation'
        v = df.loc['factor day to day correlation', d]
        sheet.cell(row, col + 1).value = v
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        if v < 0.99:
            sheet.cell(row, col + 1).font = font_red_bold
            sheet.cell(row, col + 1).fill = fill_yellow
        row = row + 1
        sheet.cell(row, col).value = 'condition number'
        v = df.loc['condition number', d]
        sheet.cell(row, col + 1).value = v
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        if v > 1e6:
            sheet.cell(row, col + 1).font = font_red_bold
            sheet.cell(row, col + 1).fill = fill_yellow
        row = row + 1
        sheet.cell(row, col).value = 'Max Factor Volatility'
        v = df.loc['max factor', d]
        sheet.cell(row, col + 1).value = v
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        v = df.loc['max volatility', d]
        sheet.cell(row, col + 2).value = v
        sheet.cell(row, col + 2).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + 2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        if v > 0.5:
            sheet.cell(row, col + 2).font = font_red_bold
        row = row + 1
        sheet.cell(row, col).value = 'Min Factor Volatility'
        v = df.loc['min factor', d]
        sheet.cell(row, col + 1).value = v
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        v = df.loc['min volatility', d]
        sheet.cell(row, col + 2).value = v
        sheet.cell(row, col + 2).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + 2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        if v == 0.0:
            sheet.cell(row, col + 1).font = font_red_bold

        row = row + 2
        sheet.cell(row, col).value = '# valid residuals'
        v = df.loc['valid residuals', d]
        sheet.cell(row, col + 1).value = v
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        row = row + 1
        sheet.cell(row, col).value = 'day to day residual vol correlation'
        v = df.loc['residual day to day correlation', d]
        sheet.cell(row, col + 1).value = v
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + 1).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        if v < 0.99:
            sheet.cell(row, col + 1).font = font_red_bold
            sheet.cell(row, col + 1).fill = fill_yellow
        row = row + 1
        sheet.cell(row, col).value = 'dropped residuals'
        v = df.loc['dropped residuals', d]
        sheet.cell(row, col + 1).value = v
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        if v > 500:
            sheet.cell(row, col + 1).font = font_red_bold
            sheet.cell(row, col + 1).fill = fill_yellow
        v = df.loc['dropped residuals %', d]
        sheet.cell(row, col + 2).value = v
        sheet.cell(row, col + 2).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + 2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        if v > 0.05:
            sheet.cell(row, col + 2).font = font_red_bold
            sheet.cell(row, col + 2).fill = fill_yellow
        row = row + 1
        sheet.cell(row, col).value = 'gained residuals'
        v = df.loc['gained residuals', d]
        sheet.cell(row, col + 1).value = v
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        if v > 500:
            sheet.cell(row, col + 1).font = font_red_bold
            sheet.cell(row, col + 1).fill = fill_yellow
        v = df.loc['gained residuals %', d]
        sheet.cell(row, col + 2).value = v
        sheet.cell(row, col + 2).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + 2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        if v > 0.05:
            sheet.cell(row, col + 2).font = font_red_bold
            sheet.cell(row, col + 2).fill = fill_yellow
        row = row + 1
        sheet.cell(row, col).value = 'Max Residual Volatility'
        v = df.loc['max security', d]
        sheet.cell(row, col + 1).value = v
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        v = df.loc['max residual', d]
        sheet.cell(row, col + 2).value = v
        sheet.cell(row, col + 2).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + 2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        if v > 1.0:
            sheet.cell(row, col + 1).font = font_red_bold
        row = row + 1
        sheet.cell(row, col).value = 'Max Increase Residual Volatility'
        v = df.loc['max residual increase security', d]
        sheet.cell(row, col + 1).value = v
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        v = df.loc['max residual increase', d]
        sheet.cell(row, col + 2).value = v
        sheet.cell(row, col + 2).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + 2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        if v > 0.2:
            sheet.cell(row, col + 1).font = font_red_bold
        row = row + 1
        sheet.cell(row, col).value = 'Min Residual Volatility'
        v = df.loc['min security', d]
        sheet.cell(row, col + 1).value = v
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        v = df.loc['min residual', d]
        sheet.cell(row, col + 2).value = v
        sheet.cell(row, col + 2).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + 2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        if v == 0.0:
            sheet.cell(row, col + 1).font = font_red_bold
        row = row + 1
        sheet.cell(row, col).value = 'Max Decrease Residual Volatility'
        v = df.loc['max residual decrease security', d]
        sheet.cell(row, col + 1).value = v
        sheet.cell(row, col + 1).alignment = styles.Alignment(horizontal='center')
        v = df.loc['max residual decrease', d]
        sheet.cell(row, col + 2).value = v
        sheet.cell(row, col + 2).alignment = styles.Alignment(horizontal='center')
        sheet.cell(row, col + 2).number_format = styles.numbers.FORMAT_PERCENTAGE_00
        if v < -0.2:
            sheet.cell(row, col + 1).font = font_red_bold
        ColumnDimension(sheet, bestFit=True)
        wb.save(output_file)
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name='Detail')
        display(f"output {d} cosmos diagnostic to {output_file}")
    return df, rv


if __name__ == "__main__":
    a = int(sys.argv[1])
    b = int(sys.argv[2])
    process_us_model(a, b)


fill_pale_blue = styles.PatternFill("solid", start_color='BBFFFF')
fill_paler_blue = styles.PatternFill("solid", start_color='CCFFFF')
fill_pale_green = styles.PatternFill("solid", start_color='EEFFEE')
fill_yellow = styles.PatternFill("solid", start_color='FFFF66')
fill_green = styles.PatternFill("solid", start_color='99FF99')
font_blue = styles.Font(color=styles.colors.BLUE)
font_blue_bold = styles.Font(bold=True, color=styles.colors.BLUE)
font_red_bold = styles.Font(bold=True, color='CC0000')
border_bottom = styles.Border(bottom=styles.Side(color='000000', border_style='double'))
thin_border = styles.Border(left=styles.Side(style='thin'),
                            right=styles.Side(style='thin'),
                            top=styles.Side(style='thin'),
                            bottom=styles.Side(style='thin'))
